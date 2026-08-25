"""SIPA — point d'entree et classe principale AuditIA_Ultimate.

ETAT DU CYCLE 8.0.0-alpha (mis a jour en continu, pas de fiche separee) :

  FAIT ET VERIFIE
    - Refonte visuelle : deux themes a fort contraste, zero gris terne.
      * « sombre » : rouge et noir profond (voir sipa_core/theme.py).
      * « clair »  : blanc, noir et accents vifs.
      Chaque role tient >= 4.5:1 sur son fond ; verifie par tests/run_all.py.
    - Audit de posture du poste local (sipa_core/feature_posture.py) : SMBv1,
      cache WDigest, RDP/NLA, dump LSASS oublie. Lit de vraies valeurs de
      registre et du disque ; statut FAIBLE / OK / INCONNU, jamais de verdict
      « machine saine ». C'est le volet defensif, oppose au pentest refuse.

  TODO: WIP — finitions UI pas encore faites (cf. discussion « rendre l'interface
  plus jolie »). Non simulees : elles ne sont pas cablees, donc l'interface se
  comporte comme en 7.x en attendant.
    - Hierarchie des boutons (action principale large, outils secondaires plus petits).
    - Panneau de constats separe du journal (une ligne + pastille de gravite).
    - Bandeau de bilan permanent (2 CRITIQUE / 1 ELEVE ...).
    - Echelle d'espacement unique, etats de survol/desactive, ecran d'accueil.

  TODO: WIP — fonctionnalites avancees demandees (securite, tracabilite, audit,
  diagnostic, base enrichie). Aucune n'est encore implementee. Regle du projet :
  tant qu'une capacite n'existe pas, elle ne doit rien afficher qui laisse croire
  le contraire — pas de coquille, pas de _not_implemented deguise en resultat.

  REFUSE — volet « pentest furtif » (Impacket/Mimikatz/BloodHound, evasion
  d'antivirus, dump LSASS, contournement de la surveillance LDAP). Hors sujet
  pour un outil d'audit defensif, et non implemente volontairement. L'inverse
  serait legitime : DETECTER ces techniques cote defense (acces anormal a LSASS,
  pic de requetes LDAP, noms de services connus de psexec). Non commence.
"""
import tkinter as tk
from tkinter import scrolledtext, messagebox, ttk, filedialog
import threading
import socket
import platform
import subprocess
import time
import random
import math
import os
import json
from collections import defaultdict
import ssl
import csv
import locale
from datetime import datetime, timedelta
from typing import Optional, Any
import hashlib
import sqlite3
import shlex
import struct
from http.server import BaseHTTPRequestHandler, HTTPServer
import queue
import logging
from logging.handlers import RotatingFileHandler

import re
import sys

# L'application ecrit des emojis et des caracteres semi-graphiques via print().
# Dans une console Windows interactive cela passe, mais des que la sortie est
# redirigee (fichier, pipe, CI) Python retombe sur cp1252 et le moindre print
# leve UnicodeEncodeError -- y compris dans les gestionnaires d'erreurs, ce qui
# masque l'erreur d'origine. On force donc UTF-8 avec remplacement.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# Echelle DPI et theme extraits dans sipa_core/theme.py
from sipa_core import APP_NAME, APP_VERSION
from sipa_core import secrets as _secrets
from sipa_core import theme as theme_module
from sipa_core import profils as _profils
from sipa_core.theme import ScaleFactor, scaled, THEME
# Widgets personnalises extraits dans sipa_core/widgets.py
from sipa_core.widgets import (
    HoverButton,
    DarkCombobox,
    StyledFrame,
    ProgressBar,
    DarkThemeManager,
    Tooltip,
)

# Lazy loaders
def _load_nmap():
    """Charger python-nmap dynamiquement (lazy load)."""
    global nmap
    if nmap is None:
        import importlib
        nmap = importlib.import_module("nmap")
    return nmap

PLOTLY_AVAILABLE = False
go = None
px = None

def _load_plotly():
    global go, px, PLOTLY_AVAILABLE
    if PLOTLY_AVAILABLE:
        return go, px
    try:
        import plotly.graph_objects as go_temp
        import plotly.express as px_temp
        go, px = go_temp, px_temp
        PLOTLY_AVAILABLE = True
        return go, px
    except ImportError:
        return None, None

# nmap reste charge paresseusement (via _load_nmap) : c'est le seul import
# reellement lourd et il n'est utile qu'au moment d'un scan.
nmap = None

# Ces noms etaient declares a None et JAMAIS importes : tout appel a
# requests.get(...), MIMEMultipart(...) ou smtplib.SMTP(...) levait donc une
# AttributeError/TypeError. Huit methodes etaient concernees (webhooks Slack et
# Discord, VirusTotal, reconnaissance web, MAC lookup, SearchSploit, scan
# broadcast, OSINT) ainsi que tout envoi d'email. On les importe reellement.
try:
    import requests
except ImportError:  # dependance listee dans requirements.txt mais absente
    requests = None

# stdlib : cout d'import negligeable, aucune raison de les differer
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Flags pour modules optionnels
TOAST_AVAILABLE = False
PIL_AVAILABLE = False
APSCHEDULER_AVAILABLE = False

try:
    from win10toast import ToastNotifier
    TOAST_AVAILABLE = True
except ImportError:
    ToastNotifier = None

try:
    from PIL import ImageGrab
    PIL_AVAILABLE = True
except ImportError:
    ImageGrab = None

# Flags pour autres modules optionnels
PDF_AVAILABLE = False
SCAPY_AVAILABLE = False
EXCEL_AVAILABLE = False
GVM_AVAILABLE = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    PDF_AVAILABLE = True
except ImportError:
    canvas = None

try:
    import scapy.all as scapy
    SCAPY_AVAILABLE = True
except ImportError:
    scapy = None

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    openpyxl = None

try:
    # `from gvm.protocols import Gvmd` n'a jamais existe dans python-gvm :
    # l'ImportError etait systematique, GVM_AVAILABLE restait donc toujours
    # False et le bouton OpenVAS reclamait d'installer une librairie presente.
    from gvm.connections import TLSConnection
    from gvm.protocols.gmp import Gmp
    GVM_AVAILABLE = True
except ImportError:
    TLSConnection = None
    Gvmd = None

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    APSCHEDULER_AVAILABLE = True
except ImportError:
    BackgroundScheduler = None
    CronTrigger = None

import io

def _load_matplotlib():
    """Charger matplotlib dynamiquement"""
    try:
        import matplotlib.pyplot as plt
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.animation import FuncAnimation
        return plt, Figure, FigureCanvasTkAgg, FuncAnimation
    except ImportError:
        return None, None, None, None

# Donnees de localisation extraites dans sipa_core/locales.py
from sipa_core.locales import (
    UI_LABELS,
    TRANSLATIONS,
    LANGUAGES,
    LICENSE_TEXT,
    BUTTON_DESCRIPTIONS,
)

# ========== TOOLTIP CLASS ==========
# =============================================================================
# [UI/UX MODERNIZATION] DARK THEME COMPONENTS - v6.0
# =============================================================================



# =============================================================================
# [PARTIE 17] BOOTSTRAP - SYSTÈME D'INSTALLATION AUTOMATIQUE (OFFLINE MODE)
# Utilise les installateurs INCLUS dans le .exe (PyInstaller)
# =============================================================================
import ctypes
import time

# Services autonomes extraits dans sipa_core/services.py
# Fenetre d'aide extraite dans sipa_core/gui_help.py (mixin)
from sipa_core.gui_help import HelpTabsMixin
# Console de commandes extraite dans sipa_core/gui_console.py (mixin)
from sipa_core.gui_console import CommandConsoleMixin
# Groupes de fonctionnalites extraits (mixins)
from sipa_core.feature_traffic import TrafficAnalysisMixin
from sipa_core.feature_forensics import ForensicsMixin
from sipa_core.feature_scans import AdvancedScansMixin
from sipa_core.feature_posture import PostureMixin
from sipa_core.gui_animations import AnimationsMixin
from sipa_core.services import (
    resource_path,
    ConfigManager,
    AuditLogger,
    EmailNotifier,
    WebhookNotifier,
    PDFReportGenerator,
    ScanScheduler,
    ScanQueue,
    AutoInstaller,
)

class AuditIA_Ultimate(
    HelpTabsMixin,
    CommandConsoleMixin,
    TrafficAnalysisMixin,
    ForensicsMixin,
    AdvancedScansMixin,
    PostureMixin,
    AnimationsMixin,
):
    def __init__(self, root):
        self.root = root
        self.root.title("NETWORK AUDIT")
        
        # Configuration de la fenêtre avec redimensionnement
        self.default_geometry = "1100x1900"
        self.window_geometry_file = "window_geometry.json"
        
        # Charger la géométrie sauvegardée ou utiliser les valeurs par défaut
        saved_geometry = self._load_window_geometry()
        if saved_geometry:
            self.root.geometry(saved_geometry)
        else:
            self.root.geometry(self.default_geometry)
        
        # Autoriser le redimensionnement
        self.root.resizable(True, True)
        self.root.configure(bg=THEME["bg"])
        
        # Sauvegarder la géométrie de la fenêtre avant la fermeture
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)

        # Langue par défaut
        self.current_language = 'FR'
        
        # CRITICAL: Initialize performance_mode BEFORE ANY LOGGING
        self.performance_mode = False
        
        # ========== THREAD-SAFE LOGGING SYSTEM (QUEUE) ==========
        # Renseigne depuis __main__ : permet a l'interface de piloter
        # la planification (le service existait mais n'etait jamais relie).
        self.scan_scheduler = None
        # Signal d'annulation : arme par le bouton STOP, verifie entre chaque
        # hote/port et avant chaque cible en multi-scan.
        import threading as _threading
        self.scan_cancel = _threading.Event()
        # Services renseignes depuis __main__ (ou le mode console).
        self.config_manager = None
        self.email_notifier = None
        self.webhook_notifier = None
        # Base de vulnerabilites locale (alimentee depuis NVD). Renseignee
        # depuis __main__ / le mode console.
        self.cve_db = None
        # Base de connaissances des appareils (registre fabricants + regles
        # de classification). Renseignee depuis __main__ / le mode console.
        self.knowledge = None
        self.msg_queue = queue.Queue()
        self._setup_logging()  # Configure logging structuré
        
        # Variables pour animations
        self.matrix_chars = []
        self.particle_list = []
        self.scan_angle = 0
        self.pulse_alpha = 0
        self.hover_effects = {}
        
        # Fonctionnalités V4.0+
        self.scan_history = []  # Historique des scans (MAX 10 en RAM)
        self.scan_results_cache = {}  # Cache des résultats
        self.auto_save_enabled = True
        self.notification_enabled = True
        self.logs_file = "network_audit_logs.txt"
        self.threat_scores = {}
        self.notification_queue = []
        self.temp_html_files = []  # Fichiers HTML temporaires Plotly à nettoyer
        
        # FONCTIONNALITÉS ENTREPRISE
        self.monitoring_active = False  # Mode surveillance 24/7
        self.stealth_mode = False  # Mode furtif
        self.scan_comparison_data = []  # Comparaison de scans
        self.network_baseline = {}  # Baseline réseau
        self.realtime_graph_data = {'time': [], 'threats': [], 'traffic': []}
        self.toaster = ToastNotifier() if TOAST_AVAILABLE else None
        self.sound_enabled = True
        self.email_config = {
            'smtp_server': 'smtp.gmail.com',
            'smtp_port': 587,
            'sender': '',
            'password': '',
            'recipients': []
        }
        self.scan_history = []  # Historique des scans
        self.scan_results_cache = {}  # Cache des résultats
        self.auto_save_enabled = True
        self.notification_enabled = True
        self.logs_file = "network_audit_logs.txt"
        self.threat_scores = {}
        self.notification_queue = []
        
        # Ports backdoor connus
        self.backdoor_ports = {
            31337: "Back Orifice", 12345: "NetBus", 54321: "Back Orifice 2000",
            1243: "SubSeven", 6666: "Beast", 6667: "Trinity", 6969: "GateCrasher",
            9999: "WinCrash", 10000: "OpwinTRojan", 20034: "NetBus Pro",
            21544: "GirlFriend", 23456: "Evil FTP", 26274: "Delta Source",
            30100: "NetSphere", 30103: "NetSphere", 40421: "Masters Paradise",
            50505: "Sockets de Troie", 65000: "Devil", 1337: "Elite Proxy"
        }
        self.problems_found = []
        self.status_alive = True

        self.setup_styles()
        self.build_header()
        self.build_target()
        self.build_controls()
        self.build_logs()
        self.build_progress()
        self.setup_responsive_design()  # Activer design responsif

        self.init_nmap()
        self.blink_status()
        self.start_matrix_rain()
        self.typewriter_log(f"{APP_NAME} {APP_VERSION} — session demarree",
                            speed=0.03, tag="title")
        self.typewriter_log("Interface prete.", speed=0.02, tag="info")
        
        # CORRECTION BUG #4: Précharger imports lourds en arrière-plan
        self._preload_heavy_libraries()
        
        # Messages de prérequis pour OpenVAS/SearchSploit - MIS À JOUR
        self.log("\n" + "="*80, tag="title")
        self.log(f"[*] {APP_NAME} — version {APP_VERSION}", tag="warn")
        self.log("="*80, tag="title")

        self.log("\nCE QUE FAIT CE LOGICIEL :", tag="accent")
        self.log("  • Scans reseau via Nmap (rapide, complet, CVE, backdoors)", tag="ok")
        self.log("  • Audit DNS, SSL/TLS et exposition de services", tag="ok")
        self.log("  • Recherche de rootkits, persistence et traces d'exploitation", tag="ok")
        self.log("  • Analyse de trafic (Scapy, sockets bruts ou netstat)", tag="ok")
        self.log("  • Rapports PDF, Excel, JSON, CSV et HTML", tag="ok")
        self.log("  • Alertes email, Slack et Discord ; planification de scans", tag="ok")
        self.log("  • Honeypot, regles de pare-feu et surveillance continue", tag="ok")

        self.log("\nNECESSITE UNE CLE API (sinon aucun resultat) :", tag="warn")
        self.log("  • SHODAN — https://account.shodan.io", tag="info")
        self.log("  • VirusTotal — https://www.virustotal.com/gui/my-apikey", tag="info")

        self.log("\nNON IMPLEMENTE (les boutons le signalent) :", tag="warn")
        self.log("  • Synchronisation Tenable, Qualys et Microsoft Defender", tag="info")
        self.log("  • Mode agent distribue, configuration proxy, sandbox", tag="info")

        self.log("\nPREREQUIS :", tag="accent")
        self.log("  • Nmap dans le PATH — https://nmap.org/download.html", tag="info")
        self.log("  • Docker Desktop, uniquement pour OpenVAS (optionnel)", tag="info")
        self.log("  • pip install -r requirements.txt", tag="info")

        self.log("\nPOUR COMMENCER :", tag="ok")
        self.log("  1. Saisissez une cible (IP, nom d'hote ou plage CIDR)", tag="info")
        self.log("  2. Lancez un SCAN RAPIDE depuis l'onglet Analyses", tag="info")
        self.log("  3. Exportez le rapport depuis l'onglet Rapports", tag="info")
        self.log("  4. Le bouton AIDE detaille chaque fonctionnalite", tag="info")

        self.log("="*80 + "\n", tag="title")
        
        self.animate_progress_pulse()  # Animation barre de progression
        self.init_database()  # Initialiser la base de données persistante
        self.check_queue() # Démarrer la surveillance de la queue thread-safe

    # ========================================================================
    # GESTION DE LA GÉOMÉTRIE DE LA FENÊTRE (Redimensionnement)
    # ========================================================================
    def _load_window_geometry(self) -> Optional[str]:
        """Charge la géométrie de la fenêtre depuis un fichier JSON"""
        try:
            if os.path.exists(self.window_geometry_file):
                with open(self.window_geometry_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    return data.get('geometry')
        except Exception as e:
            print(f"[WARN] Impossible de charger la géométrie: {e}")
        return None

    def _save_window_geometry(self):
        """Sauvegarde la géométrie actuelle de la fenêtre"""
        try:
            geometry = self.root.geometry()
            with open(self.window_geometry_file, 'w', encoding='utf-8') as f:
                json.dump({'geometry': geometry}, f, indent=2)
        except Exception as e:
            print(f"[WARN] Impossible de sauvegarder la géométrie: {e}")

    def _on_closing(self):
        """Gère la fermeture de la fenêtre"""
        self._save_window_geometry()
        self.root.quit()

    def reset_window_size(self):
        """Réinitialise la taille de la fenêtre aux dimensions par défaut"""
        self.root.geometry(self.default_geometry)
        self._save_window_geometry()

    # ========================================================================
    # THREAD-SAFE LOGGING SYSTEM avec queue.Queue + logging structuré
    # ========================================================================
    def _setup_logging(self):
        """Configure le système de logging structuré Python avec rotation"""
        # Créer un logger nommé
        self.logger = logging.getLogger('T800_NetworkAudit')
        self.logger.setLevel(logging.DEBUG)
        
        # Éviter les doublons de handlers
        if self.logger.hasHandlers():
            self.logger.handlers.clear()
        
        # Handler 1: Fichier avec rotation (10MB max, 5 fichiers)
        file_handler = RotatingFileHandler(
            'network_audit_logs.txt',
            maxBytes=10*1024*1024,  # 10 MB
            backupCount=5,
            encoding='utf-8'
        )
        file_handler.setLevel(logging.DEBUG)
        file_formatter = logging.Formatter(
            '%(asctime)s - %(name)s - [%(levelname)s] - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(file_formatter)
        self.logger.addHandler(file_handler)
        
        # Handler 2: Console (optionnel)
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_formatter = logging.Formatter('[%(levelname)s] %(message)s')
        console_handler.setFormatter(console_formatter)
        self.logger.addHandler(console_handler)
        
        self.logger.info("="*60)
        self.logger.info(f"{APP_NAME} {APP_VERSION} - Session demarree")
        self.logger.info("="*60)
    
    def _preload_heavy_libraries(self):
        """BUG #4: Précharge Matplotlib/Plotly/Scapy en thread pour éviter gel UI"""
        def preload_worker():
            try:
                time.sleep(3)  # Laisser UI s'afficher d'abord
                self.log("[*] Préchargement librairies graphiques...", tag="info")
                
                # Charger Matplotlib
                _load_matplotlib()
                self.log("Matplotlib prêt", tag="ok")
                
                # Charger Plotly
                _load_plotly()
                self.log("Plotly prêt", tag="ok")
                
                # Scapy optionnel (très lourd)
                # _load_scapy()
                
                self.log("Toutes les librairies chargées (pas de gel UI)", tag="success")
            except Exception as e:
                self.log(f"[!] Préchargement partiel: {e}", tag="warn")
        
        threading.Thread(target=preload_worker, daemon=True, name="LibraryPreloader").start()
    
    def check_queue(self):
        """
        THREAD-SAFE: Vérifie la queue et met à jour l'interface depuis le thread principal
        Cette méthode est appelée périodiquement par root.after()
"""
        try:
            while not self.msg_queue.empty():
                msg, tag = self.msg_queue.get_nowait()
                
                # Mise à jour SÉCURISÉE du widget Tkinter (thread principal uniquement)
                self.text_area.config(state="normal")
                self.text_area.insert(tk.END, msg + "\n", tag)
                self.text_area.see(tk.END)
                self.text_area.config(state="normal")
        except queue.Empty:
            pass
        except Exception as e:
            # Fallback: logger l'erreur sans planter
            try:
                self.logger.error(f"Erreur check_queue: {e}")
            except:
                pass
        finally:
            # Se rappelle toutes les 100ms (équilibre entre réactivité et CPU)
            self.root.after(100, self.check_queue)
    
    def set_progress_real(self, value, maximum=100):
        """BUG #3: Progression réelle au lieu de ping-pong indéterminé"""
        try:
            self.progress["mode"] = "determinate"
            self.progress["maximum"] = maximum
            self.progress["value"] = value
            self.root.update_idletasks()
        except Exception as e:
            pass

    def start_progress_tracking(self, total_steps=5):
        """BUG #3: Initialise progression réelle basée sur étapes"""
        self.current_progress = 0
        self.total_progress_steps = total_steps
        self.set_progress_real(0, 100)

    def increment_progress(self, message=""):
        """BUG #3: Incrémente progression avec feedback utilisateur"""
        self.current_progress += 1
        percentage = int((self.current_progress / self.total_progress_steps) * 100)
        self.set_progress_real(percentage, 100)
        if message:
            self.log(f"[{percentage}%] {message}", tag="info")
    
    def _cleanup_temp_html_files(self):
        """BUG #2: Nettoyer fichiers HTML temporaires de Plotly"""
        try:
            import os
            cleaned = 0
            for temp_file in self.temp_html_files:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    cleaned += 1
            self.temp_html_files.clear()
            if cleaned > 0:
                self.log(f"[*] {cleaned} fichiers HTML temporaires nettoyés", tag="info")
        except Exception as e:
            self.log(f"[!] Erreur nettoyage HTML: {e}", tag="warn")

    # -------------------------------------------------------------------------
    # [REAL] PERSISTENCE ENGINE (SQLite)
    # Windows gère les fichiers .db avec une élégance que EXT4 ne comprendra jamais.
    # -------------------------------------------------------------------------
    def init_database(self):
        """Initialise une mémoire persistante réelle pour stocker les audits."""
        try:
            # check_same_thread=False permet d'appeler depuis les threads de
            # scan, mais ne rend pas la connexion thread-safe pour autant :
            # ce verrou protège les écritures concurrentes (voir save_scan_result_real).
            self.db_lock = threading.Lock()
            self.db_conn = sqlite3.connect('t800_neural_net.db', check_same_thread=False)
            self.cursor = self.db_conn.cursor()
            
            # Création des tables si elles n'existent pas
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS scan_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT,
                    target_ip TEXT,
                    scan_type TEXT,
                    vulns_found INTEGER
                )
            ''')
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS vulnerabilities (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    scan_id INTEGER,
                    host TEXT,
                    port INTEGER,
                    service TEXT,
                    cve_id TEXT,
                    severity TEXT,
                    details TEXT,
                    FOREIGN KEY(scan_id) REFERENCES scan_history(id)
                )
            ''')
            self.db_conn.commit()
            self.log("[INIT] Base de données locale (SQLite) connectée.", tag="success")
        except Exception as e:
            self.log(f"[ERREUR DB] Échec initialisation mémoire: {e}", tag="error")

    def save_scan_result_real(self, target, scan_type, vulns):
        """Sauvegarde réelle des résultats pour analyse historique (Delta)."""
        try:
            # Plusieurs scans peuvent se terminer en même temps sur des threads
            # différents : sans verrou, les écritures/lastrowid peuvent
            # s'entrelacer et corrompre l'association scan_id <-> vulns.
            with self.db_lock:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.cursor.execute("INSERT INTO scan_history (timestamp, target_ip, scan_type, vulns_found) VALUES (?, ?, ?, ?)",
                                    (timestamp, target, scan_type, len(vulns)))
                scan_id = self.cursor.lastrowid

                for v in vulns:
                    # Les deux dernieres colonnes recevaient les mauvaises valeurs :
                    # `cve_id` heritait du texte de detail et `severity` du type de
                    # constat. Tout l'historique en etait fausse.
                    self.cursor.execute(
                        "INSERT INTO vulnerabilities "
                        "(scan_id, host, port, service, cve_id, severity, details) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (scan_id, v.get('host'), v.get('port'), v.get('service'),
                         v.get('type'), self.classify_severity(v), v.get('details')))
                self.db_conn.commit()
            self.log("[DB] Données tactiques archivées sur le disque.", tag="ok")
        except Exception as e:
            self.log(f"[DB SAVE ERROR] {e}", tag="error")

    # ---------- SAFE COMMAND ----------
    def run_safe_command(self, cmd_list, timeout=10):
        """Exécute une commande système sans faire planter l'interface avec Unicode"""
        try:
            # IMPORTANT : text=False pour récupérer les données brutes (bytes)
            result = subprocess.run(cmd_list, capture_output=True, text=False, timeout=timeout)
            # On décode manuellement en ignorant les erreurs (cp850 est l'encodage console Windows)
            return result.stdout.decode('cp850', errors='ignore')
        except Exception:
            return ""

    # ---------- UI BUILD ----------
    def setup_styles(self):
        """Configure les styles modernes pour l'interface - v6.0 Dark Theme"""
        style = ttk.Style()
        style.theme_use('clam')  # Thème plus moderne que 'default'
        
        # === PROGRESSBAR MODERNE ===
        style.configure("TProgressbar", 
                       thickness=8,
                       background=THEME["fg"],
                       troughcolor=THEME["bg_secondary"],
                       lightcolor=THEME["fg_bright"],
                       darkcolor=THEME["fg_dim"])
        
        # === SCROLLBAR MODERNE ===
        style.configure("Vertical.TScrollbar",
                       background=THEME["bg_secondary"],
                       troughcolor=THEME["bg"],
                       bordercolor=THEME["bg"],
                       arrowcolor=THEME["fg"],
                       lightcolor=THEME["bg_secondary"],
                       darkcolor=THEME["bg_secondary"])
        style.map("Vertical.TScrollbar",
                 background=[("active", THEME["fg"]), ("!active", THEME["bg_secondary"])])
        
        # === COMBOBOX SOMBRE (DARK MODE) ===
        style.configure("TCombobox",
                       fieldbackground=THEME["bg"],
                       background=THEME["bg_secondary"],
                       foreground=THEME["accent"],
                       borderwidth=1,
                       relief="solid")
        style.map("TCombobox",
                 fieldbackground=[("readonly", THEME["bg"]), ("disabled", THEME["bg_tertiary"])],
                 foreground=[("readonly", THEME["accent"]), ("disabled", THEME["accent_dim"])],
                 background=[("active", THEME["bg_secondary"]), ("!active", THEME["bg"])])
        
        # === NOTEBOOK MODERNE AVEC HOVER ===
        style.configure('TNotebook',
                       background=THEME["bg"],
                       borderwidth=0,
                       padding=2)
        style.configure('TNotebook.Tab',
                       padding=[16, 10],
                       background=THEME["bg_secondary"],
                       foreground=THEME["accent"],
                       font=THEME["font_button"])
        style.map('TNotebook.Tab',
                 background=[('selected', THEME["fg"]),
                            ('active', THEME["fg_bright"]),
                            ('!selected', THEME["bg_secondary"])],
                 foreground=[('selected', THEME["bg"]),
                            ('active', THEME["bg"]),
                            ('!selected', THEME["accent"])])
        
        # === BUTTON STYLE ===
        style.configure("TButton",
                       background=THEME["bg_secondary"],
                       foreground="#FFFFFF",         # Texte BLANC au lieu de rouge
                       borderwidth=1,
                       relief="solid",
                       font=THEME["font_button"])
        style.map("TButton",
                 background=[("active", THEME["fg"]), ("!active", THEME["bg_secondary"])],
                 foreground=[("active", "#000000"), ("!active", "#FFFFFF")])  # Noir sur rouge, blanc normal
        
        # === LABEL STYLE ===
        style.configure("TLabel",
                       background=THEME["bg"],
                       foreground="#FFFFFF",         # Texte BLANC au lieu de rouge
                       font=THEME["font_label"])
        
        # === FRAME STYLE ===
        style.configure("TFrame", background=THEME["bg"], borderwidth=0)

        # === LISTES DEROULANTES (sous-categories) ===
        # ttk gardait le rendu clair de Windows : champ blanc, texte noir, et
        # une liste deroulante blanche illisible sur fond sombre.
        style.configure("TCombobox",
                        fieldbackground=THEME["bg_input"],
                        background=THEME["bg_secondary"],
                        foreground=THEME["fg_text"],
                        arrowcolor=THEME["fg"],
                        bordercolor=THEME["fg_dim"],
                        lightcolor=THEME["bg_secondary"],
                        darkcolor=THEME["bg_secondary"],
                        selectbackground=THEME["fg"],
                        selectforeground=THEME["bg"],
                        padding=scaled(5))
        style.map("TCombobox",
                  fieldbackground=[("readonly", THEME["bg_input"])],
                  foreground=[("readonly", THEME["fg_text"]),
                              ("disabled", THEME["btn_muted"])],
                  bordercolor=[("focus", THEME["fg"])],
                  arrowcolor=[("active", THEME["fg_text"])])
        # La liste deroulante elle-meme est une widget Tk classique, stylable
        # uniquement via la base de donnees d'options.
        self.root.option_add("*TCombobox*Listbox.background", THEME["bg_input"])
        self.root.option_add("*TCombobox*Listbox.foreground", THEME["fg_text"])
        self.root.option_add("*TCombobox*Listbox.selectBackground", THEME["fg"])
        self.root.option_add("*TCombobox*Listbox.selectForeground", THEME["bg"])
        self.root.option_add("*TCombobox*Listbox.font", THEME["font_label"])

        # === TABLEAUX (inventaire reseau) ===
        style.configure("Treeview",
                        background=THEME["bg_input"],
                        fieldbackground=THEME["bg_input"],
                        foreground=THEME["fg_text"],
                        bordercolor=THEME["bg_secondary"],
                        borderwidth=0,
                        rowheight=scaled(24),
                        font=THEME["font_label"])
        style.configure("Treeview.Heading",
                        background=THEME["bg_secondary"],
                        foreground=THEME["fg"],
                        relief="flat",
                        font=THEME["font_button"])
        style.map("Treeview.Heading",
                  background=[("active", THEME["fg_dim"])],
                  foreground=[("active", THEME["fg_text"])])
        # Selection rouge pleine, texte noir : lisible et coherent avec le
        # survol des boutons.
        style.map("Treeview",
                  background=[("selected", THEME["fg"])],
                  foreground=[("selected", THEME["bg"])])

        # === ONGLETS : marquer nettement l'onglet actif ===
        style.configure("TNotebook.Tab", padding=[scaled(16), scaled(10)],
                        font=THEME["font_button"])
        style.map("TNotebook.Tab",
                  background=[("selected", THEME["fg"]),
                              ("active", THEME["fg_dim"]),
                              ("!selected", THEME["bg_secondary"])],
                  foreground=[("selected", THEME["bg"]),
                              ("active", THEME["fg_text"]),
                              ("!selected", THEME["btn_tool"])])

    def build_header(self):
        """En-tete de la fenetre : nom, langue, voyant d'etat et horloge."""
        frame = tk.Frame(self.root, bg=THEME["bg"], bd=0, relief="flat")
        frame.pack(fill="x", padx=0, pady=0)
        
        # Un seul filet d'accent (sous l'en-tete) : deux barres epaisses
        # encadraient l'en-tete et alourdissaient l'ecran.
        
        # Contenu header
        inner = tk.Frame(frame, bg=THEME["bg_secondary"])
        inner.pack(fill="x", padx=THEME["padding_lg"], pady=THEME["padding_lg"])

        # Gauche : nom et version
        left = tk.Frame(inner, bg=THEME["bg_secondary"])
        left.pack(side=tk.LEFT, padx=THEME["padding_std"])
        
        # Titre typographique : les blocs ASCII "▓▓▓" faisaient daté et
        # n'apportaient aucune information.
        self.title_label = tk.Label(left, text=APP_NAME.upper(),
                        font=THEME["font_header"], bg=THEME["bg_secondary"],
                        fg=THEME["fg_text"])
        self.title_label.pack(anchor="w")

        subtitle = tk.Label(left, text=f"Programme unique intégré pour les audits · v{APP_VERSION}",
                           font=THEME["font_label"], bg=THEME["bg_secondary"],
                           fg=THEME["btn_muted"])
        subtitle.pack(anchor="w", pady=(2, 0))

        # Centre - Langue
        center = tk.Frame(inner, bg=THEME["bg_secondary"])
        center.pack(side=tk.LEFT, padx=20, expand=True)
        
        tk.Label(center, text="Langue", font=THEME["font_label"],
                bg=THEME["bg_secondary"], fg=THEME["btn_muted"]).pack(anchor="w")
        
        # Utiliser le Combobox sombre personnalisé
        try:
            self.lang_combo = DarkCombobox(center, values=list(LANGUAGES.values()), 
                                          state="readonly", width=14, font=THEME["font_mono"])
        except:
            # Fallback sur ttk.Combobox standard
            self.lang_combo = ttk.Combobox(center, values=list(LANGUAGES.values()), 
                                          state="readonly", width=14, font=THEME["font_label"])
        self.lang_combo.set(LANGUAGES['FR'])
        self.lang_combo.pack(pady=2)
        self.lang_combo.bind("<<ComboboxSelected>>", self._change_language)

        # Droite - Status LED
        right = tk.Frame(inner, bg=THEME["bg_secondary"])
        right.pack(side=tk.RIGHT, padx=THEME["padding_std"])
        
        self.status_led = tk.Label(right, text="●", fg=THEME["fg"], 
                                  bg=THEME["bg_secondary"], font=("Arial", 16, "bold"))
        self.status_led.pack(anchor="e")
        
        self.lbl_time = tk.Label(right, text="--:--:--", 
                                bg=THEME["bg_secondary"], fg=THEME["accent"], 
                                font=THEME["font_label"])
        self.lbl_time.pack(anchor="e", pady=2)
        self.update_time()
        
        # Ligne inférieure glow (rouge)
        glow_bottom = tk.Frame(frame, bg=THEME["fg"], height=2)
        glow_bottom.pack(fill="x", padx=0, pady=0)

    #: Texte d'aide affiche dans le champ cible tant qu'il est vide.
    TARGET_PLACEHOLDER = "Adresse IP, nom d'hôte ou plage CIDR (ex. 192.168.1.0/24)"

    def build_target(self):
        frame = tk.LabelFrame(self.root, text=self.tr(" Cible "), font=THEME["font_title"],
                              bg=THEME["bg"], fg=THEME["fg"], bd=1, relief="solid")
        frame.pack(padx=THEME["padding_lg"], pady=THEME["padding_std"], fill="x")

        self.entry_ip = tk.Entry(frame, font=THEME["font_mono"], bg=THEME["bg_input"],
                                 fg=THEME["accent"], insertbackground=THEME["fg"],
                                 bd=0, relief="flat")
        self.entry_ip.pack(side=tk.LEFT, fill="x", expand=True,
                           padx=(THEME["padding_lg"], THEME["padding_std"]),
                           pady=THEME["padding_std"], ipady=scaled(6))

        # Indice contextuel : le champ etait vide et sans explication, l'utilisateur
        # ne pouvait pas deviner les formats acceptes.
        # get_target() renvoie "" quand le champ ne contient que le texte
        # d'aide. Comparer get_target() au texte d'aide etait donc toujours
        # faux : le focus ne l'effacait jamais, chaque perte de focus le
        # re-inserait, et deux exemplaires colles cessaient d'etre reconnus
        # et partaient comme cible de scan. On suit l'etat par un drapeau.
        self._placeholder_visible = False

        def show_placeholder():
            if not self._placeholder_visible and not self.get_target():
                self.entry_ip.insert(0, self.TARGET_PLACEHOLDER)
                self.entry_ip.config(fg=THEME["btn_muted"])
                self._placeholder_visible = True

        def clear_placeholder(_event=None):
            if self._placeholder_visible:
                self.entry_ip.delete(0, tk.END)
                self.entry_ip.config(fg=THEME["accent"])
                self._placeholder_visible = False

        self.entry_ip.bind("<FocusIn>", clear_placeholder)
        self.entry_ip.bind("<FocusOut>", lambda e: show_placeholder())
        show_placeholder()

        self.make_button(frame, "Détecter", self.auto_detect, width=14).pack(
            side=tk.LEFT, padx=(0, THEME["padding_lg"]), pady=THEME["padding_std"])

    def animate_widget(self, widget, duration=500, effect="fade"):
        """Anime un widget (fade-in/slide)"""
        if effect == "fade":
            start_alpha = 0
            end_alpha = 1
            step = end_alpha / (duration // 16)
            
            def fade_step(alpha):
                if alpha < end_alpha:
                    widget.pack_configure(padx=10, pady=10)
                    self.root.after(16, fade_step, alpha + step)
        elif effect == "slide":
            start_y = -50
            end_y = 0
            step = 50 / (duration // 16)
            
            def slide_step(y):
                if y < end_y:
                    widget.pack_configure(pady=abs(int(y)))
                    self.root.after(16, slide_step, y + step)
        elif effect == "glow":
            # Animation glow (pulse) pour les onglets
            colors = [THEME["fg"], THEME["fg_bright"], THEME["fg"], THEME["bg_secondary"]]
            
            def glow_step(index):
                try:
                    if hasattr(widget, 'config'):
                        widget.config(fg=colors[index % len(colors)])
                    self.root.after(100, glow_step, index + 1)
                except:
                    pass
            
            glow_step(0)
    
    def animate_tab_glow(self, notebook):
        """Ajoute un effet glow à l'onglet sélectionné"""
        def on_tab_change(event):
            selected_tab = notebook.select()
            # Appliquer glow à l'onglet actif
            try:
                # Récupérer l'index de l'onglet
                tab_index = notebook.index(selected_tab)
                # Style déjà configuré avec map() - pas besoin de reconfigurer
                pass
            except:
                pass
        
        notebook.bind("<<NotebookTabChanged>>", on_tab_change)

    def build_controls(self):
        """Interface compacte: 5 onglets avec sous-catégories (2 niveaux)"""
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook', background=THEME["bg"], borderwidth=0)
        style.configure('TNotebook.Tab', padding=[14, 8])
        style.map('TNotebook.Tab',
                  background=[('selected', THEME["fg"]), ('!selected', THEME["bg_secondary"])],
                  foreground=[('selected', '#FFFFFF'), ('!selected', THEME["accent"])])

        notebook = ttk.Notebook(self.root)
        notebook.pack(padx=10, pady=10, fill="both", expand=True)
        self.animate_tab_glow(notebook)

        def render_button_grid(container, buttons):
            """Affiche une grille de boutons à partir d'une liste"""
            for child in container.winfo_children():
                child.destroy()
            # Les boutons sont enregistres pour que disable_buttons() puisse
            # reellement les griser pendant un scan.
            if not hasattr(self, "_boutons_generes"):
                self._boutons_generes = []
            cols = 3
            for idx, btn_def in enumerate(buttons):
                label = btn_def[0]
                command = btn_def[1]
                color = btn_def[2] if len(btn_def) > 2 else None
                fg = btn_def[3] if len(btn_def) > 3 else None
                width = btn_def[4] if len(btn_def) > 4 else 18
                btn = self.make_button(container, label, command, width=width, color=color, fg=fg)
                btn.grid(row=idx // cols, column=idx % cols, sticky="ew", padx=5, pady=5)
                self._boutons_generes.append(btn)
            for col in range(cols):
                container.columnconfigure(col, weight=1)

        def build_tab_with_combo(tab_frame, title, options):
            """Onglet avec ses sous-categories en boutons segmentes.

            Une liste deroulante cachait les categories derriere un clic alors
            qu'il n'y en a que deux par onglet : on les montre cote a cote, la
            selection etant marquee par un aplat rouge et un soulignement.
            """
            header = tk.Frame(tab_frame, bg=THEME["bg"])
            header.pack(fill="x", padx=THEME["padding_std"],
                        pady=(THEME["padding_std"], scaled(6)))
            tk.Label(header, text=self.tr(title), font=THEME["font_title"],
                     bg=THEME["bg"], fg=THEME["fg"]).pack(side=tk.LEFT,
                                                          padx=(0, scaled(16)))

            switcher = tk.Frame(header, bg=THEME["bg"])
            switcher.pack(side=tk.LEFT)

            content = tk.Frame(tab_frame, bg=THEME["bg"])
            content.pack(fill="both", expand=True,
                         padx=THEME["padding_std"], pady=THEME["padding_std"])

            # Fin trait sous l'en-tete : separe visuellement le choix de
            # categorie de la grille de boutons.
            tk.Frame(tab_frame, bg=THEME["fg_dim"], height=1).pack(
                fill="x", padx=THEME["padding_std"], before=content)

            segments = {}

            def select(name):
                for key, (button, underline) in segments.items():
                    active = (key == name)
                    button.config(
                        bg=THEME["fg"] if active else THEME["bg_secondary"],
                        fg=THEME["bg"] if active else THEME["btn_tool"])
                    underline.config(
                        bg=THEME["fg"] if active else THEME["bg_secondary"])
                render_button_grid(content, options.get(name, []))

            for name in options:
                cell = tk.Frame(switcher, bg=THEME["bg"])
                cell.pack(side=tk.LEFT, padx=(0, scaled(6)))
                button = tk.Button(
                    cell, text=self.tr(name).upper(), font=THEME["font_button"],
                    bg=THEME["bg_secondary"], fg=THEME["btn_tool"],
                    activebackground=THEME["fg"], activeforeground=THEME["bg"],
                    relief="flat", bd=0, cursor="hand2",
                    padx=scaled(16), pady=scaled(7),
                    command=lambda n=name: select(n))
                button.pack(fill="x")
                underline = tk.Frame(cell, bg=THEME["bg_secondary"], height=scaled(3))
                underline.pack(fill="x")
                segments[name] = (button, underline)

                # Survol : seulement si la categorie n'est pas deja active.
                def enter(_e, b=button, n=name):
                    if segments[n][1].cget("bg") != THEME["fg"]:
                        b.config(fg=THEME["fg_text"])

                def leave(_e, b=button, n=name):
                    if segments[n][1].cget("bg") != THEME["fg"]:
                        b.config(fg=THEME["btn_tool"])

                button.bind("<Enter>", enter)
                button.bind("<Leave>", leave)

            if options:
                select(next(iter(options)))
            return segments, content

        # ---------- TAB 1: ANALYSES ----------
        # Deux categories seulement par onglet : la liste deroulante reste
        # lisible et chaque groupe correspond a une intention claire.
        analyses_options = {
            "Scans réseau": [
                ("Scan rapide", lambda: self.start_scan("fast"), THEME["btn_action"]),
                ("Scan complet", lambda: self.start_scan("full"), THEME["btn_action"]),
                ("Scan CVE", lambda: self.start_scan("vuln"), THEME["btn_action"]),
                ("Portes dérobées", lambda: self.start_scan("backdoor"), THEME["btn_critical"]),
                ("Audit des services", self.run_service_exposure_check, THEME["btn_action"]),
                ("Santé du réseau", self.check_network_health, THEME["btn_action"]),
                ("Analyse DNS", self.analyze_dns, THEME["btn_tool"]),
                ("SSL / TLS", self.scan_ssl_tls, THEME["btn_tool"]),
                ("Chaîne de certificats", self.scan_ssl_chain, THEME["btn_tool"]),
            ],
            "Menaces et forensique": [
                ("Exploits connus", self.check_exploitation, THEME["btn_critical"]),
                ("Exploits avancés", self.detect_advanced_exploits, THEME["btn_critical"]),
                ("Rootkits", self.detect_rootkits, THEME["btn_critical"]),
                ("Mécanismes de persistance", self.check_persistence, THEME["btn_critical"]),
                ("Attaques par force brute", self.detect_brute_force, THEME["btn_critical"]),
                ("Analyse forensique", self.forensic_analysis, THEME["btn_action"]),
                ("Audit du système", self.audit_system_internals, THEME["btn_action"]),
                ("Détection d'anomalies", self.detect_anomalies_ai, THEME["btn_action"]),
                ("Secrets dans vos fichiers", self.scan_credentials, THEME["btn_tool"]),
            ],
        }

        tab_analyses = tk.Frame(notebook, bg=THEME["bg"])
        notebook.add(tab_analyses, text=self.tr("Analyses"))
        build_tab_with_combo(tab_analyses, "Analyses", analyses_options)

        # ---------- TAB 2: RÉSEAU & RENSEIGNEMENT ----------
        reseau_options = {
            "Cartographie et trafic": [
                ("Analyse du trafic", self.analyze_traffic, THEME["btn_action"]),
                ("Graphiques en temps réel", self.show_realtime_graphs, THEME["btn_action"]),
                ("Graphe du réseau", self.visualize_network, THEME["btn_tool"]),
                ("Carte du réseau", self.show_network_map, THEME["btn_tool"]),
                ("Changements depuis le dernier scan", self.detect_network_changes, THEME["btn_tool"]),
                ("Intégrité ARP", self.check_arp_integrity, THEME["btn_tool"]),
                ("Fabricant (adresse MAC)", self.tool_mac_lookup, THEME["btn_muted"]),
                ("Réseaux Wi-Fi", self.scan_wifi, THEME["btn_muted"]),
                ("Découverte par diffusion", self.scan_broadcast, THEME["btn_muted"]),
            ],
            "Renseignement externe": [
                ("Shodan", self.shodan_lookup, THEME["btn_action"]),
                ("VirusTotal", self.virustotal_scan, THEME["btn_action"]),
                ("Réputation d'une IP", self.check_threat_intelligence, THEME["btn_action"]),
                ("SearchSploit", self.run_real_searchsploit, THEME["btn_critical"]),
                ("Renseignement public (OSINT)", self.osint_research, THEME["btn_tool"]),
                ("Active Directory", self.audit_active_directory, THEME["btn_tool"]),
                ("Géolocalisation", self.geolocate_ips, THEME["btn_tool"]),
                ("Technologies web", self.scan_web_technologies, THEME["btn_tool"]),
            ],
        }

        tab_reseau = tk.Frame(notebook, bg=THEME["bg"])
        notebook.add(tab_reseau, text=self.tr("Réseau et renseignement"))
        build_tab_with_combo(tab_reseau, "Réseau", reseau_options)

        # ---------- TAB 3: RAPPORTS & OUTILS ----------
        reports_options = {
            "Rapports et historique": [
                ("Exporter (PDF, Excel, JSON)", self.export_reports, THEME["btn_action"]),
                ("Rapport HTML", self.generate_html_report, THEME["btn_action"]),
                ("Tableau de bord", self.show_dashboard, THEME["btn_tool"]),
                ("Tableau de bord 3D", self.show_3d_dashboard, THEME["btn_tool"]),
                ("Appareils détectés", self.show_inventory_window, THEME["btn_action"]),
                ("Historique des scans", self.show_scan_history, THEME["btn_tool"]),
                ("Comparer deux scans", self.compare_scans, THEME["btn_tool"]),
                ("Rejouer un scan", self.replay_scans, THEME["btn_muted"]),
                ("Journal d'audit", self.show_logs_viewer, THEME["btn_muted"]),
                ("Capture d'écran", self.capture_screenshot, THEME["btn_muted"]),
            ],
            "Défense et surveillance": [
                ("Audit de posture", self.lancer_audit_posture, THEME["btn_action"]),
                ("Surveillance continue", self.toggle_monitoring, THEME["btn_action"]),
                ("Règles du pare-feu", self.generate_firewall_rules, THEME["btn_critical"]),
                ("Leurre réseau", self.deploy_honeypot_plus, THEME["btn_critical"]),
                ("Mode discret", self.toggle_stealth_mode, THEME["btn_tool"]),
                ("Processus en cours", self.analyze_processes, THEME["btn_tool"]),
                ("Bac à sable", self.open_sandbox, THEME["btn_muted"]),
            ],
        }

        tab_reports = tk.Frame(notebook, bg=THEME["bg"])
        notebook.add(tab_reports, text=self.tr("Rapports et défense"))
        build_tab_with_combo(tab_reports, "Rapports", reports_options)

        # ---------- TAB 4: COMMANDES (inchangé) ----------
        tab11 = tk.Frame(notebook, bg=THEME["bg"])
        notebook.add(tab11, text=self.tr("Commandes"))
        frame11 = tk.Frame(tab11, bg=THEME["bg"])
        frame11.pack(padx=10, pady=10, fill="both", expand=True)
        
        title_cmd = tk.Label(frame11, text="Commandes manuelles", 
                             font=("Arial", 11, "bold"), bg=THEME["bg"], fg=THEME["fg"])
        title_cmd.pack(pady=5)
        
        cmd_section = tk.LabelFrame(frame11, text=" COMMANDE & OPTIONS ", bg=THEME["bg"], 
                                   fg=THEME["accent"], font=THEME["font_label"])
        cmd_section.pack(fill="x", padx=5, pady=3)
        
        cmd_input_frame = tk.Frame(cmd_section, bg=THEME["bg"])
        cmd_input_frame.pack(fill="x", padx=5, pady=3)
        
        tk.Label(cmd_input_frame, text="CMD:", font=("Arial", 8), 
                bg=THEME["bg"], fg=THEME["accent"]).pack(side=tk.LEFT, padx=2)
        
        self.cmd_entry = tk.Entry(cmd_input_frame, width=40, font=THEME["font_mono"], 
                                  bg="#1a1a1a", fg=THEME["fg"], insertbackground=THEME["fg"],
                                  bd=1, relief="solid")
        self.cmd_entry.pack(side=tk.LEFT, padx=2, fill="x", expand=True)
        self.cmd_entry.bind("<Return>", lambda e: self.execute_custom_command())
        self.cmd_entry.bind("<Up>", lambda e: self.history_previous())
        self.cmd_entry.bind("<Down>", lambda e: self.history_next())
        self.cmd_entry.bind("<Tab>", lambda e: self.autocomplete_command(e))
        
        btn_exec = self.make_button(cmd_input_frame, "▶", 
                                   lambda: self.execute_custom_command(), width=3, color=THEME["fg"])
        btn_exec.pack(side=tk.LEFT, padx=1)
        
        btn_fav = self.make_button(cmd_input_frame, "", 
                                  lambda: self.add_favorite_command(), width=3)
        btn_fav.pack(side=tk.LEFT, padx=1)
        
        btn_copy = self.make_button(cmd_input_frame, "", 
                                   lambda: self.copy_results(), width=3)
        btn_copy.pack(side=tk.LEFT, padx=1)
        
        btn_save = self.make_button(cmd_input_frame, "", 
                                   lambda: self.save_results(), width=3)
        btn_save.pack(side=tk.LEFT, padx=1)
        
        result_frame = tk.Frame(frame11, bg=THEME["bg"])
        result_frame.pack(fill="both", expand=True, padx=5, pady=3)
        
        scrollbar_cmd = ttk.Scrollbar(result_frame, style="Vertical.TScrollbar")
        scrollbar_cmd.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.cmd_result = tk.Text(result_frame, width=100, height=16,
                                 bg="#0a0a0a", fg=THEME["accent"],
                                 font=THEME["font_mono"],
                                 bd=1, relief="solid",
                                 insertbackground=THEME["fg"],
                                 selectbackground=THEME["fg_bright"],
                                 yscrollcommand=scrollbar_cmd.set)
        self.cmd_result.pack(side=tk.LEFT, fill="both", expand=True)
        scrollbar_cmd.config(command=self.cmd_result.yview)
        
        self.cmd_result.tag_config("error", foreground="#880000")
        self.cmd_result.tag_config("success", foreground=THEME["btn_tool"])
        self.cmd_result.tag_config("command", foreground=THEME["fg"], font=THEME["font_button"])
        
        self.command_history = []
        self.command_favorites = []
        self.history_position = -1

        # Barre d'outils de la console. Ces widgets etaient references par le
        # code (self.timeout_var, self.hist_combo, self.fav_combo) mais n'etaient
        # crees nulle part : la console echouait a chaque commande et les
        # favoris s'ajoutaient a une liste que rien n'affichait.
        tools = tk.Frame(frame11, bg=THEME["bg"])
        tools.pack(fill="x", padx=5, pady=(0, 4))

        tk.Label(tools, text="Délai (s) :", font=THEME["font_label"],
                 bg=THEME["bg"], fg=THEME["btn_muted"]).pack(side=tk.LEFT)
        self.timeout_var = tk.IntVar(value=15)
        tk.Spinbox(tools, from_=1, to=300, width=5, textvariable=self.timeout_var,
                   font=THEME["font_mono"], bg=THEME["bg_input"],
                   fg=THEME["accent"], buttonbackground=THEME["bg_secondary"],
                   insertbackground=THEME["fg"], relief="flat",
                   justify="center").pack(side=tk.LEFT, padx=(4, 14))

        tk.Label(tools, text="Historique :", font=THEME["font_label"],
                 bg=THEME["bg"], fg=THEME["btn_muted"]).pack(side=tk.LEFT)
        self.hist_combo = ttk.Combobox(tools, state="readonly", width=26,
                                       font=THEME["font_label"])
        self.hist_combo.pack(side=tk.LEFT, padx=(4, 14))
        self.hist_combo.bind("<<ComboboxSelected>>",
                             lambda e: self.load_from_history(self.hist_combo))

        tk.Label(tools, text="Favoris :", font=THEME["font_label"],
                 bg=THEME["bg"], fg=THEME["btn_muted"]).pack(side=tk.LEFT)
        self.fav_combo = ttk.Combobox(tools, state="readonly", width=26,
                                      font=THEME["font_label"])
        self.fav_combo.pack(side=tk.LEFT, padx=4)
        self.fav_combo.bind("<<ComboboxSelected>>",
                            lambda e: self.load_from_favorites(self.fav_combo))

        # ---------- TAB 5: PARAMÈTRES ----------
        settings_options = {
            "Intégrations": [
                ("OpenVAS (Windows)", self.run_openvas_windows, THEME["btn_action"]),
                ("OpenVAS (Docker)", self.run_openvas_bridge, THEME["btn_action"]),
                ("Serveur API REST", self.start_api_server, THEME["btn_action"]),
                ("Scan multi-cibles", self.scan_multi_target, THEME["btn_tool"]),
                ("Planifier des scans", self.schedule_scans, THEME["btn_action"]),
                ("Slack", self.setup_slack_notifications, THEME["btn_tool"]),
                ("Webhook", self.setup_webhooks, THEME["btn_tool"]),
                ("Tenable", self.sync_tenable, THEME["btn_muted"]),
                ("Qualys", self.sync_qualys, THEME["btn_muted"]),
                ("Microsoft Defender", self.sync_defender, THEME["btn_muted"]),
                ("Mode agent", self.start_agent_mode, THEME["btn_muted"]),
            ],
            "Préférences et alertes": [
                ("Aide", self.show_help, THEME["btn_action"]),
                ("Configurer l'e-mail", self.configure_email, THEME["btn_tool"]),
                ("Notifications", self.test_notification, THEME["btn_tool"]),
                ("Sons d'alerte", self.toggle_sound_alerts, THEME["btn_tool"]),
                ("Mode performance", self.toggle_performance, THEME["btn_tool"]),
                ("Thème sombre", lambda: self.toggle_dark_mode(True), THEME["btn_muted"]),
                ("Thème clair", lambda: self.toggle_dark_mode(False), THEME["btn_muted"]),
                ("Palette active", self.show_current_theme, THEME["btn_muted"]),
            ],
        }

        tab_settings = tk.Frame(notebook, bg=THEME["bg"])
        notebook.add(tab_settings, text=self.tr("Paramètres"))
        build_tab_with_combo(tab_settings, "Paramètres", settings_options)

        # Barre d'etat : la place etait occupee par un bandeau rouge pleine
        # largeur "PANIC BUTTON (WIPE ALL)", visuellement dominant alors qu'il
        # s'agit d'une action rare. On garde l'action, discrete, a droite.
        status_bar = tk.Frame(self.root, bg=THEME["bg_secondary"])
        status_bar.pack(side=tk.BOTTOM, fill="x")

        tk.Frame(status_bar, bg=THEME["fg_dim"], height=1).pack(fill="x")

        inner_status = tk.Frame(status_bar, bg=THEME["bg_secondary"])
        inner_status.pack(fill="x", padx=THEME["padding_lg"], pady=scaled(6))

        self.lbl_status = tk.Label(inner_status, text="Prêt", anchor="w",
                                   bg=THEME["bg_secondary"], fg=THEME["btn_muted"],
                                   font=THEME["font_label"])
        self.lbl_status.pack(side=tk.LEFT)

        self.make_button(inner_status, "EFFACEMENT D'URGENCE", self.emergency_wipe,
                         width=24, color=THEME["btn_critical"]).pack(side=tk.RIGHT)
        self.make_button(inner_status, "⏹ ARRÊTER LE SCAN", self.request_cancel,
                         width=20, color=THEME["btn_muted"]).pack(side=tk.RIGHT, padx=8)


    def setup_responsive_design(self):
        """Configure le layout adaptatif en fonction de la taille de fenêtre"""
        def on_window_resize(event=None):
            width = self.root.winfo_width()
            height = self.root.winfo_height()
            
            # Adapter les marges et polices selon la taille
            if width < 800:
                # Petit écran: réduire marges et polices
                padding = 5
                font_size_small = 8
                font_size_normal = 9
            elif width < 1200:
                # Écran moyen
                padding = 10
                font_size_small = 9
                font_size_normal = 10
            else:
                # Grand écran
                padding = 15
                font_size_small = 10
                font_size_normal = 11
            
            # Sauvegarder pour utilisation dans les widgets
            self.responsive_padding = padding
            self.responsive_font_small = ("Arial", font_size_small)
            self.responsive_font_normal = ("Arial", font_size_normal)
        
        # Bind resize event
        self.root.bind("<Configure>", on_window_resize)
        
        # Initialiser avec taille par défaut
        on_window_resize()

    # =========================================================================
    # INVENTAIRE RÉSEAU : fenêtre dédiée, ouverte à la fin d'une analyse
    # =========================================================================

    #: Date de premiere et derniere detection, par hote.
    SEEN_FILE = "appareils_vus.json"

    def _load_seen(self):
        try:
            with open(self.SEEN_FILE, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            return data if isinstance(data, dict) else {}
        except FileNotFoundError:
            return {}
        except Exception:
            return {}

    def record_seen(self, hosts):
        """Note la date de detection de chaque hote (premiere et derniere)."""
        seen = self._load_seen()
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        for host in hosts:
            entry = seen.setdefault(host, {"premiere": stamp})
            entry["derniere"] = stamp
        try:
            with open(self.SEEN_FILE, "w", encoding="utf-8") as handle:
                json.dump(seen, handle, indent=2, ensure_ascii=False)
        except Exception as exc:
            self.log(f"[INVENTAIRE] Historique de détection non écrit : {exc}",
                     tag="warn")

    def _host_os(self, host):
        """Systeme devine par nmap (necessite -O), ou chaine vide."""
        try:
            matches = self.nm[host].get("osmatch") or []
        except Exception:
            return ""
        if not matches:
            return ""
        best = matches[0]
        name = (best.get("name") or "").strip()
        accuracy = best.get("accuracy")
        # En dessous de 85 % nmap se trompe souvent : on le signale.
        if accuracy and int(accuracy) < 85:
            return f"{name} (~{accuracy}%)"
        return name

    def query_announcements(self, timeout=3.0):
        """Demande aux appareils du reseau de se presenter (SSDP, mDNS, NetBIOS).

        Ce sont des requetes de decouverte standard : on interroge, l'appareil
        repond ce qu'il diffuse deja publiquement. Aucun trafic tiers n'est
        intercepte.
        """
        try:
            from sipa_core.knowledge import (discover_ssdp, discover_mdns,
                                             netbios_name)
        except Exception as exc:
            self.log(f"[DÉCOUVERTE] Module indisponible : {exc}", tag="warn")
            return {}

        self.log("\n[DÉCOUVERTE] Interrogation des appareils "
                 "(UPnP, mDNS, NetBIOS)...", tag="title")
        announcements = {}

        try:
            for ip, info in discover_ssdp(timeout).items():
                label = info.get("server") or info.get("modele") or ""
                if label:
                    announcements[ip] = f"{announcements.get(ip, '')} {label}".strip()
                    self.log(f"  [UPnP] {ip} : {label[:80]}", tag="info")
        except Exception as exc:
            self.log(f"  [UPnP] Interrogation impossible : {exc}", tag="warn")

        try:
            for ip, services in discover_mdns(timeout).items():
                label = " ".join(services[:8])
                announcements[ip] = f"{announcements.get(ip, '')} {label}".strip()
                self.log(f"  [mDNS] {ip} : {label[:80]}", tag="info")
        except Exception as exc:
            self.log(f"  [mDNS] Interrogation impossible : {exc}", tag="warn")

        # NetBIOS : uniquement sur les hotes deja connus, pour rester rapide.
        if platform.system() == "Windows" and self.nm:
            for host in list(self.nm.all_hosts())[:20]:
                if self.scan_cancel.is_set():
                    break
                name = netbios_name(host)
                if name:
                    announcements[host] = f"{announcements.get(host, '')} {name}".strip()
                    self.log(f"  [NetBIOS] {host} : {name}", tag="info")

        if not announcements:
            self.log("  Aucun appareil ne s'est annoncé "
                     "(multicast filtré ou réseau sans objets connectés).",
                     tag="info")
        else:
            self.log(f"  {len(announcements)} appareil(s) se sont présentés.",
                     tag="success")

        self._announcements = announcements
        return announcements

    @property
    def _announced(self):
        """Ce que chaque appareil a annonce de lui-meme (SSDP/mDNS)."""
        return getattr(self, "_announcements", {}) or {}

    def _host_mac(self, host):
        """Adresse MAC vue par nmap (vide si l'hote n'est pas sur le lien local)."""
        try:
            return (self.nm[host].get("addresses") or {}).get("mac", "") or ""
        except Exception:
            return ""

    def _host_vendor(self, host):
        """Fabricant : celui de nmap, sinon le registre local des OUI.

        nmap ne connait qu'une partie des prefixes ; la base de connaissances
        en contient pres de 39 000, donc elle comble les trous.
        """
        try:
            vendors = self.nm[host].get("vendor") or {}
            if vendors:
                return next(iter(vendors.values()), "")
        except Exception:
            pass
        mac = self._host_mac(host)
        knowledge = getattr(self, "knowledge", None)
        if mac and knowledge is not None:
            try:
                return knowledge.vendor_for(mac)
            except Exception:
                return ""
        return ""

    def _host_type(self, host, vendor, hostname, ports):
        """Nature de l'appareil, deduite de tous les indices disponibles."""
        knowledge = getattr(self, "knowledge", None)
        if knowledge is None:
            return ""
        try:
            kind, score = knowledge.classify(
                vendor=vendor, ports=ports, hostname=hostname,
                announced=(self._announced or {}).get(host, ""))
            # Sous 4 points l'indice est trop faible pour etre affirme.
            return kind if score >= 4 else ""
        except Exception:
            return ""

    def controle_bon_sens(self):
        """Confronte ce qu'est chaque appareil a ce qu'il expose reellement.

        C'est le seul controle de SIPA qui ne se contente pas de lister : il
        rapproche le type d'appareil (deduit du fabricant, des annonces mDNS et
        des ports) de ce qu'on attend d'un appareil de cette nature. Un port
        d'administration a distance sur une imprimante n'est pas anormal en
        soi -- il est inattendu POUR UNE IMPRIMANTE, et c'est cela qu'on dit.

        Renvoie le registre de couverture : ce qui a ete controle, ce qui ne
        l'a pas ete, et pourquoi. Sans cette contrepartie, un silence se lirait
        comme une absence de probleme.
        """
        registre = {"controles": [], "non_controles": []}
        if not self.nm:
            registre["non_controles"].append(
                ("controle du bon sens", "aucun scan nmap disponible"))
            return registre

        for host in self.nm.all_hosts():
            try:
                hostname = self.nm[host].hostname() or ""
            except Exception:
                hostname = ""

            ports = []
            for proto in self.nm[host].all_protocols():
                ports += [p for p, d in self.nm[host][proto].items()
                          if d.get("state") == "open"]

            vendor = self._host_vendor(host)
            type_appareil = self._host_type(host, vendor, hostname, ports)

            if not type_appareil:
                # Dire qu'on n'a pas su, plutot que de se taire : un appareil
                # non classe n'est pas un appareil sans probleme.
                registre["non_controles"].append(
                    (host, "type d'appareil non identifie avec certitude"))
                continue

            if _profils.profil(type_appareil) is None:
                registre["non_controles"].append(
                    (host, f"aucune attente definie pour un appareil de type "
                           f"« {type_appareil} »"))
                continue

            registre["controles"].append((host, type_appareil))
            surprises = _profils.controler(type_appareil, ports)
            if not surprises:
                continue

            for surprise in surprises:
                self.problems_found.append({
                    "type": "INATTENDU POUR CE TYPE",
                    "host": host,
                    "port": surprise["port"],
                    "service": surprise["raison"],
                    "details": (f"Cet appareil est identifie comme « "
                                f"{type_appareil} ». Le port {surprise['port']} "
                                f"expose {surprise['raison']}, ce qui n'est pas "
                                f"attendu pour ce type d'appareil. "
                                f"Regle appliquee : {surprise['regle']}."),
                    "action": ("Verifier que cette exposition est voulue. "
                               "Si elle l'est, marquer le constat comme accepte."),
                    "risk": "MOYEN",
                })

        return registre

    def afficher_bon_sens(self):
        """Journalise le controle du bon sens et sa couverture."""
        registre = self.controle_bon_sens()

        self.log("", tag="info")
        self.log("CONTROLE DU BON SENS", tag="title")
        self.log("Ce que chaque appareil expose est-il coherent avec ce qu'il est ?",
                 tag="info")

        surprises = [p for p in self.problems_found
                     if p.get("type") == "INATTENDU POUR CE TYPE"]

        for host, type_appareil in registre["controles"]:
            propres = [s for s in surprises if s.get("host") == host]
            if propres:
                self.log(f"   {host} — {type_appareil}", tag="accent")
                for s in propres:
                    self.log(f"      port {s['port']} : {s['service']}", tag="warn")
                    self.log("      Inattendu pour ce type. Confirmez que c'est voulu.",
                             tag="info")
            else:
                self.log(f"   {host} — {type_appareil} : rien d'inattendu", tag="ok")

        # La contrepartie, toujours affichee : ce qui n'a PAS ete controle.
        self.log("", tag="info")
        self.log(f"   Controle effectue sur {len(registre['controles'])} appareil(s).",
                 tag="info")
        if registre["non_controles"]:
            self.log(f"   NON controle sur {len(registre['non_controles'])} appareil(s) :",
                     tag="warn")
            for cible, raison in registre["non_controles"]:
                self.log(f"      {cible} : {raison}", tag="warn")
            self.log("   Sur ces appareils, l'absence de constat ne veut pas dire "
                     "l'absence de probleme.", tag="warn")
        return registre

    def _inventory_rows(self):
        """Une ligne par machine détectée.

        Construit a partir du dernier scan nmap et des constats releves. Sans
        dependance a Tkinter, donc testable et reutilisable (export CSV).
        """
        rows = []
        if not self.nm:
            return rows

        seen = self._load_seen()
        for host in self.nm.all_hosts():
            try:
                hostname = self.nm[host].hostname() or ""
            except Exception:
                hostname = ""

            open_ports = 0
            for proto in self.nm[host].all_protocols():
                open_ports += sum(1 for _, data in self.nm[host][proto].items()
                                  if data.get("state") == "open")

            worst = "INFO"
            for problem in self.problems_found:
                if problem.get("host") != host:
                    continue
                severity = self.classify_severity(problem)
                if self.SEVERITIES.index(severity) < self.SEVERITIES.index(worst):
                    worst = severity

            port_numbers = []
            for proto in self.nm[host].all_protocols():
                port_numbers += [p for p, d in self.nm[host][proto].items()
                                 if d.get("state") == "open"]

            history = seen.get(host, {})
            vendor = self._host_vendor(host)
            row = {"ip": host, "nom": hostname or "—",
                   "ports": open_ports, "risque": worst,
                   "os": self._host_os(host) or "—",
                   "fabricant": vendor or "—",
                   "mac": self._host_mac(host) or "—",
                   "type": self._host_type(host, vendor, hostname,
                                           port_numbers) or "—",
                   "vue_le": history.get("premiere", "—"),
                   "vue_dernier": history.get("derniere", "—")}
            rows.append(row)

            # Fiche persistante : le parc se souvient d'un scan a l'autre.
            knowledge = getattr(self, "knowledge", None)
            if knowledge is not None:
                try:
                    knowledge.remember(
                        host, mac=row["mac"] if row["mac"] != "—" else "",
                        fabricant=vendor, type=row["type"] if row["type"] != "—" else "",
                        systeme=row["os"] if row["os"] != "—" else "",
                        nom=hostname, decouvert_par="nmap")
                except Exception:
                    pass

        # Les machines les plus exposees d'abord, puis par adresse.
        rows.sort(key=lambda r: (self.SEVERITIES.index(r["risque"]),
                                 self._ip_sort_key(r["ip"])))
        return rows

    @staticmethod
    def _ip_sort_key(value):
        """Clef de tri numerique pour une IP (10.0.0.9 avant 10.0.0.10)."""
        parts = str(value).split(".")
        if len(parts) == 4 and all(p.isdigit() for p in parts):
            return (0, sum(int(p) << (8 * (3 - i)) for i, p in enumerate(parts)))
        return (1, str(value).lower())

    def show_inventory_window(self, auto=False):
        """Ouvre (ou ramène au premier plan) la fenêtre des appareils détectés.

        `auto=True` signale une ouverture automatique en fin de scan : dans ce
        cas on n'affiche rien s'il n'y a aucun appareil, pour ne pas imposer
        une fenetre vide.
        """
        # En mode console il n'y a pas de racine Tk : ouvrir une Toplevel y
        # levait "'_NullRoot' object has no attribute 'tk'", exception attrapee
        # par run_nmap qui annoncait alors "SCAN FAILURE" alors que le scan
        # s'etait parfaitement deroule.
        if getattr(self, "headless", False):
            rows = self._inventory_rows()
            if rows:
                self.log(f"[INVENTAIRE] {len(rows)} appareil(s) connu(s).", tag="info")
            return

        rows = self._inventory_rows()
        if not rows:
            if auto:
                return
            messagebox.showinfo(
                "Aucun appareil",
                "Lancez d'abord une analyse réseau pour peupler l'inventaire.")
            return

        existing = getattr(self, "inventory_window", None)
        if existing is not None:
            try:
                if existing.winfo_exists():
                    self._fill_inventory(rows)
                    existing.deiconify()
                    existing.lift()
                    return
            except Exception:
                pass

        window = tk.Toplevel(self.root)
        self.inventory_window = window
        window.title(f"Appareils détectés — {self.get_target() or 'réseau local'}")
        window.configure(bg=THEME["bg"])
        scale = max(1.0, window.winfo_fpixels("1i") / 96.0)
        width, height = int(780 * scale), int(520 * scale)
        window.geometry(
            f"{width}x{height}"
            f"+{max(0, (window.winfo_screenwidth() - width) // 2)}"
            f"+{max(0, (window.winfo_screenheight() - height) // 2)}")
        window.minsize(int(520 * scale), int(320 * scale))

        head = tk.Frame(window, bg=THEME["bg"])
        head.pack(fill="x", padx=THEME["padding_lg"], pady=(THEME["padding_lg"], 0))
        tk.Label(head, text="APPAREILS DÉTECTÉS", font=THEME["font_title"],
                 bg=THEME["bg"], fg=THEME["fg"]).pack(side=tk.LEFT)
        self.lbl_inventory_count = tk.Label(
            head, text="", font=THEME["font_label"],
            bg=THEME["bg"], fg=THEME["btn_muted"])
        self.lbl_inventory_count.pack(side=tk.RIGHT)

        body = tk.Frame(window, bg=THEME["bg"])
        body.pack(fill="both", expand=True,
                  padx=THEME["padding_lg"], pady=THEME["padding_std"])

        scroll = ttk.Scrollbar(body, style="Vertical.TScrollbar")
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.inventory = ttk.Treeview(
            body, columns=("ip", "nom", "type", "os", "fabricant", "ports", "risque"),
            show="headings", selectmode="browse", yscrollcommand=scroll.set)
        for key, label, width_px, anchor in (("ip", "Adresse", 120, "w"),
                                             ("nom", "Nom d'hôte", 140, "w"),
                                             ("type", "Type d'appareil", 150, "w"),
                                             ("os", "Système", 150, "w"),
                                             ("fabricant", "Fabricant", 140, "w"),
                                             ("ports", "Ports", 60, "center"),
                                             ("risque", "Risque", 90, "center")):
            self.inventory.heading(key, text=label,
                                   command=lambda c=key: self._sort_inventory(c))
            self.inventory.column(key, width=scaled(width_px), anchor=anchor)
        self.inventory.pack(side=tk.LEFT, fill="both", expand=True)
        scroll.config(command=self.inventory.yview)

        self.inventory.tag_configure("critique", foreground=THEME["btn_critical"])
        self.inventory.tag_configure("eleve", foreground=THEME["btn_action"])
        self.inventory.tag_configure("sain", foreground=THEME["btn_tool"])
        self.inventory.bind("<<TreeviewSelect>>", self._on_inventory_select)
        self._inventory_sort = {"column": None, "reverse": False}

        detail = tk.LabelFrame(window, text=" Détail de l'appareil sélectionné ",
                               font=THEME["font_label"], bg=THEME["bg"],
                               fg=THEME["fg"], bd=1, relief="solid")
        detail.pack(fill="x", padx=THEME["padding_lg"],
                    pady=(0, THEME["padding_std"]))
        self.inventory_detail = tk.Text(
            detail, height=6, bg=THEME["bg_input"], fg=THEME["fg_text"],
            font=THEME["font_mono"], bd=0, highlightthickness=0, wrap="word",
            padx=scaled(10), pady=scaled(8))
        self.inventory_detail.pack(fill="both", expand=True)
        self.inventory_detail.insert("1.0",
                                     "Sélectionnez un appareil pour voir ses constats.")
        self.inventory_detail.config(state="disabled")

        actions = tk.Frame(window, bg=THEME["bg"])
        actions.pack(fill="x", padx=THEME["padding_lg"],
                     pady=(0, THEME["padding_lg"]))
        self.make_button(actions, "FERMER", window.destroy, width=10,
                         color=THEME["btn_muted"]).pack(side=tk.RIGHT)
        self.make_button(actions, "EXPORTER CSV", self.export_inventory_csv,
                         width=15).pack(side=tk.RIGHT, padx=8)
        self.make_button(actions, "ACCEPTER LES CONSTATS",
                         self._accept_selected_host, width=22,
                         color=THEME["btn_tool"]).pack(side=tk.LEFT)

        self._fill_inventory(rows)

    def _fill_inventory(self, rows=None):
        """(Re)remplit le tableau de la fenêtre d'inventaire."""
        if not hasattr(self, "inventory") or self.inventory is None:
            return
        rows = self._inventory_rows() if rows is None else rows
        try:
            for item in self.inventory.get_children(""):
                self.inventory.delete(item)
        except Exception:
            return

        critical = 0
        for row in rows:
            severity = row["risque"]
            if severity == "CRITIQUE":
                critical += 1
            tag = ("critique" if severity == "CRITIQUE"
                   else "eleve" if severity in ("ÉLEVÉ", "MOYEN")
                   else "sain")
            self.inventory.insert(
                "", "end", iid=row["ip"], tags=(tag,),
                values=(row["ip"], row["nom"], row.get("type", "—"),
                        row.get("os", "—"), row.get("fabricant", "—"),
                        row["ports"], "—" if severity == "INFO" else severity))

        if hasattr(self, "lbl_inventory_count"):
            plural = "s" if len(rows) > 1 else ""
            summary = f"{len(rows)} appareil{plural}"
            if critical:
                summary += f" · {critical} critique{'s' if critical > 1 else ''}"
            self.lbl_inventory_count.config(text=summary)

    def _sort_inventory(self, column):
        """Trie sur la colonne cliquée (inverse si déjà triée dessus)."""
        if not hasattr(self, "inventory") or self.inventory is None:
            return
        reverse = (self._inventory_sort["column"] == column
                   and not self._inventory_sort["reverse"])
        rows = [(self.inventory.set(item, column), item)
                for item in self.inventory.get_children("")]

        def key(pair):
            value = pair[0]
            if column == "ip":
                return self._ip_sort_key(value)
            try:
                return (0, float(value))
            except ValueError:
                return (1, value.lower())

        for index, (_, item) in enumerate(sorted(rows, key=key, reverse=reverse)):
            self.inventory.move(item, "", index)
        self._inventory_sort = {"column": column, "reverse": reverse}

    def _on_inventory_select(self, _event=None):
        """Affiche les constats de la machine sélectionnée."""
        if not hasattr(self, "inventory") or self.inventory is None:
            return
        selection = self.inventory.selection()
        if not selection:
            return
        host = self.inventory.set(selection[0], "ip")
        related = [p for p in self.problems_found if p.get("host") == host]

        info = next((r for r in self._inventory_rows() if r["ip"] == host), {})
        lines = [f"{host} — {len(related)} constat(s)"]
        if info:
            lines.append(f"Type : {info.get('type', '—')}   ·   "
                         f"Système : {info.get('os', '—')}")
            lines.append(f"Fabricant : {info.get('fabricant', '—')}   ·   "
                         f"MAC : {info.get('mac', '—')}")
            lines.append(f"Vu pour la première fois : {info.get('vue_le', '—')}"
                         f"   ·   Dernière fois : {info.get('vue_dernier', '—')}")
        lines.append("")
        for problem in related:
            severity = self.classify_severity(problem)
            marker = " (accepté)" if self.is_accepted(problem) else ""
            lines.append(f"[{severity}]{marker} {problem.get('type')} "
                         f"· port {problem.get('port', '—')} "
                         f"· {problem.get('service', '—')}")
            details = (problem.get("details") or "").strip()
            if details:
                lines.append(f"    {details[:220]}")
            action = (problem.get("action") or "").strip()
            if action:
                lines.append(f"    → {action[:180]}")
            lines.append("")
        if not related:
            lines.append("Aucun problème relevé sur cet appareil.")

        if hasattr(self, "inventory_detail"):
            self.inventory_detail.config(state="normal")
            self.inventory_detail.delete("1.0", tk.END)
            self.inventory_detail.insert("1.0", chr(10).join(lines))
            self.inventory_detail.config(state="disabled")

    def _accept_selected_host(self):
        """Declare voulus tous les constats de la machine selectionnee.

        Utile pour un serveur dont on sait que les ports ouverts sont
        legitimes : ils restent listes mais cessent de compter comme risque.
        """
        if not hasattr(self, "inventory") or self.inventory is None:
            return
        selection = self.inventory.selection()
        if not selection:
            messagebox.showinfo("Aucune sélection",
                                "Sélectionnez d'abord un appareil dans la liste.")
            return
        host = self.inventory.set(selection[0], "ip")
        related = [p for p in self.problems_found if p.get("host") == host]
        if not related:
            messagebox.showinfo("Rien à accepter",
                                f"Aucun constat sur {host}.")
            return
        if not messagebox.askyesno(
                "Accepter les constats",
                f"Marquer les {len(related)} constat(s) de {host} comme voulus ?"
                f"{chr(10)}{chr(10)}Ils resteront visibles dans les rapports mais "
                f"ne compteront plus comme risques."):
            return
        for problem in related:
            self.accept_finding(problem, reason=f"parc connu ({host})")
        self._fill_inventory()
        self._on_inventory_select()

    def export_inventory_csv(self):
        """Exporte l'inventaire courant en CSV."""
        rows = self._inventory_rows()
        if not rows:
            self.log("[INVENTAIRE] Aucun appareil à exporter.", tag="warn")
            if not getattr(self, "headless", False):
                messagebox.showinfo("Rien à exporter", "Aucun appareil détecté.")
            return
        filename = f"inventaire_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        try:
            with open(filename, "w", newline="", encoding="utf-8-sig") as handle:
                writer = csv.writer(handle, delimiter=";")
                writer.writerow(["Adresse", "Nom d'hote", "Type", "Systeme",
                                 "Fabricant", "MAC", "Ports ouverts", "Risque",
                                 "Premiere detection", "Derniere detection"])
                for row in rows:
                    writer.writerow([row["ip"], row["nom"], row.get("type", ""),
                                     row.get("os", ""), row.get("fabricant", ""),
                                     row.get("mac", ""), row["ports"],
                                     row["risque"], row.get("vue_le", ""),
                                     row.get("vue_dernier", "")])
            self.log(f"[INVENTAIRE] Export CSV : {os.path.abspath(filename)}",
                     tag="success")
            # Pas de boite modale en mode console ou sous test : elle bloquerait
            # une execution sans interface.
            if not getattr(self, "headless", False):
                messagebox.showinfo("Export terminé",
                                    f"{len(rows)} appareil(s) exporté(s) vers :"
                                    f"{chr(10)}{os.path.abspath(filename)}")
        except Exception as exc:
            self.log(f"[INVENTAIRE] Export impossible : {exc}", tag="error")
            if not getattr(self, "headless", False):
                messagebox.showerror("Export impossible", str(exc))

    def build_logs(self):
        """Crée la zone de journal (pleine largeur).

        L'inventaire des appareils vit dans sa propre fenetre, ouverte a la
        fin d'une analyse : la fenetre principale reste lisible.
        """
        frame = tk.LabelFrame(self.root, text=" Journal d'activité ",
                              font=THEME["font_title"],
                              bg=THEME["bg_secondary"], fg=THEME["fg"],
                              bd=0, relief="flat")
        frame.pack(padx=THEME["padding_lg"], pady=THEME["padding_std"],
                   fill="both", expand=True)

        # Text widget + Scrollbar
        text_frame = tk.Frame(frame, bg=THEME["bg"])
        text_frame.pack(padx=THEME["padding_std"], pady=THEME["padding_std"],
                       fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame, style="Vertical.TScrollbar")
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.text_area = tk.Text(text_frame, width=100, height=12,
                    bg=THEME["bg_input"], fg=THEME["fg_text"],
                    font=THEME["font_mono"],
                    bd=0, highlightthickness=0,
                    insertbackground=THEME["accent"],
                    selectbackground=THEME["accent"],
                    selectforeground=THEME["bg"],
                    relief="flat", wrap=tk.WORD,
                    padx=scaled(12), pady=scaled(12),
                    yscrollcommand=scrollbar.set)
        self.text_area.pack(side=tk.LEFT, fill="both", expand=True)
        scrollbar.config(command=self.text_area.yview)
        
        # Chaque niveau du journal a sa couleur, tiree de la palette active
        # pour suivre la bascule sombre/clair. Auparavant le niveau erreur
        # avait un contraste de 1.87 (illisible), succes etait identique a
        # titre, et les tags `success` et `accent` -- employes 118 fois --
        # n'avaient aucune couleur.
        fond = THEME["bg_input"]
        police = THEME["font_mono"][0]
        niveaux = {
            "title":   (THEME["log_title"], True),
            "accent":  (THEME["log_accent"], True),
            "info":    (THEME["log_info"], False),
            "ok":      (THEME["log_ok"], False),
            "success": (THEME["log_ok"], False),
            "warn":    (THEME["log_warn"], False),
            "error":   (THEME["log_error"], True),
        }
        for nom, (couleur, gras) in niveaux.items():
            self.text_area.tag_config(
                nom, foreground=couleur, background=fond,
                font=(police, 11, "bold") if gras else (police, 11))

        self.text_area.tag_config("matrix", foreground=THEME["log_faint"],
                                  font=(police, 8), background=fond)
        
        # Cache pour optimisation des logs
        self._log_cache = []
        self._max_log_lines = 2000  # Limite pour performance

    def build_progress(self):
        """Barre de progression moderne"""
        self.progress = ttk.Progressbar(self.root, orient=tk.HORIZONTAL, 
                                       mode='indeterminate', style="TProgressbar")
        self.progress.pack(fill="x", padx=0, pady=0)
    
    def log(self, message, tag="info", speed=0.01):
        """Affiche un message dans le journal, depuis n'importe quel thread.

        Le tag demande est respecte tel quel. Auparavant un "effet CRT" le
        remplacait avant affichage : `ok` et `success` devenaient un bordeaux
        quasi noir (contraste 1.04 sur le fond du journal, soit invisible)
        et `warn` comme `error` recevaient le meme rouge, rendant les deux
        niveaux indistinguables.
        """
        # Thread-safe : le message passe par la file, jamais directement au
        # widget -- Tkinter n'est pas utilisable depuis un thread de travail.
        self.msg_queue.put((message, tag))
        
        # Logger aussi dans le fichier structuré
        if tag == "error":
            self.logger.error(message)
        elif tag == "warn":
            self.logger.warning(message)
        elif tag == "success" or tag == "ok":
            self.logger.info(message)
        else:
            self.logger.debug(message)

    def make_button(self, parent, text, command, width=16, color=THEME["fg"], fg=None):
        """Cree un bouton du theme, avec repli si HoverButton echoue.

        Le libelle passe par tr() : c'est le point unique ou l'interface est
        traduite, ce qui evite de dupliquer les cles sur 94 sites d'appel.
        """
        text = self.tr(text)
        # Utiliser le nouveau HoverButton si disponible, sinon créer un bouton standard amélioré
        try:
            return HoverButton(parent, text, command=command, width=width, color=color, fg=fg)
        except:
            # Fallback: bouton classique avec hover effects
            # Texte BLANC par défaut (lisible sur noir)
            fg_color = fg if fg else "#FFFFFF"
            btn = tk.Button(parent, text=text, command=command, 
                           bg=THEME["bg"], fg=fg_color,  # Texte blanc
                           width=width, font=("Courier New", 10, "bold"),
                           activebackground=color,
                           activeforeground="#000000",
                           bd=2, relief="flat", cursor="hand2",
                           padx=THEME["padding_std"], pady=THEME["padding_std"],
                           highlightthickness=1, highlightbackground=color, 
                           highlightcolor=THEME["fg"])
            
            # Effets de survol modernes
            btn.bind("<Enter>", lambda e: self._button_hover_in(btn, color))
            btn.bind("<Leave>", lambda e: self._button_hover_out(btn, color, fg_color))
            btn.bind("<ButtonPress-1>", lambda e: btn.config(relief="raised", bd=3))
            btn.bind("<ButtonRelease-1>", lambda e: btn.config(relief="flat", bd=2))
            
            # Stockage des couleurs
            btn.original_color = color
            btn.original_fg = fg_color
            
            return btn
    
    def _button_hover_in(self, btn, color):
        """Effet de survol entrée - RED ALERT Terminator"""
        btn.config(bg=THEME["fg"], fg="#000000", relief="raised")  # RED + noir
    
    def _button_hover_out(self, btn, color, fg_color):
        """Effet de survol sortie - state normal"""
        btn.config(bg=THEME["bg"], fg=fg_color, relief="flat")  # Retour au blanc

    # ---------- SYSTEM ----------
    def update_time(self):
        now = time.strftime("%H:%M:%S")
        self.lbl_time.config(text=f"SYS_TIME: {now}")
        self.root.after(1000, self.update_time)

    def _change_language(self, event=None):
        """Change la langue de l'interface"""
        selected = self.lang_combo.get()
        for lang_code, lang_name in LANGUAGES.items():
            if lang_name == selected:
                self.current_language = lang_code
                self.update_ui_language()
                break

    def tr(self, texte):
        """Traduit un libelle d'interface vers la langue courante.

        La traduction se fait AU RENDU plutot que dans build_controls : les
        libelles francais restent les cles canoniques du code, et changer de
        langue revient a reconstruire l'interface. Une chaine absente du
        tableau UI_LABELS est renvoyee telle quelle.
        """
        langue = getattr(self, "current_language", "FR")
        if langue == "FR":
            return texte
        return UI_LABELS.get(texte, {}).get(langue, texte)

    def update_ui_language(self):
        """Applique la langue choisie a toute l'interface.

        Cette methode affichait "Buttons will update when available" et ne
        reconfigurait aucun widget : la variable `trans` etait meme calculee
        puis jetee. Les libelles sont desormais reellement retraduits, et le
        choix est enregistre pour le prochain demarrage.
        """
        langue = getattr(self, "current_language", "FR")

        enregistre = False
        config = getattr(self, "config_manager", None)
        if config is not None:
            try:
                config.set("interface.langue", langue)
                enregistre = True
            except Exception as exc:
                self.log(f"[LANGUE] Preference non enregistree : {exc}", tag="warn")

        try:
            self.rebuild_interface()
        except Exception as exc:
            self.log(f"[LANGUE] Reconstruction impossible : {exc}", tag="error")
            self.log("[LANGUE] Redemarrez SIPA pour appliquer la langue.", tag="warn")
            return

        nom = LANGUAGES.get(langue, langue)
        self.log(f"[LANGUE] Interface en {nom}.", tag="ok")
        if langue != "FR":
            traduits = sum(1 for valeurs in UI_LABELS.values() if langue in valeurs)
            self.log(f"[LANGUE] {traduits} libelles traduits. Les messages du "
                     "journal restent en francais.", tag="info")
        if enregistre:
            self.log("[LANGUE] Preference enregistree dans config.json.", tag="info")

    def blink_status(self):
        self.status_alive = not self.status_alive
        color = THEME["warn"] if self.status_alive else THEME["bg_secondary"]
        self.status_led.config(fg=color)
        
        # Effet de pulsation sur le LED
        size = 14 if self.status_alive else 12
        self.status_led.config(font=("Consolas", size))
        
        self.root.after(600, self.blink_status)

    def init_nmap(self):
        """Initialisation robuste de Nmap pour Windows avec retry logic"""
        self.nm = None
        self.nmap_status = "OFFLINE"
        self.nmap_path = None
        
        # Liste des chemins possibles pour nmap.exe
        possible_paths = [
            "nmap",  # Si c'est dans le PATH système
            r"C:\Program Files (x86)\Nmap\nmap.exe",
            r"C:\Program Files\Nmap\nmap.exe",
            os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Programs', 'Common', 'nmap.exe'),
            r"C:\nmap\nmap.exe",
            r"D:\nmap\nmap.exe"
        ]
        
        found_path = ""
        
        # 1. Tester la commande directe (version test)
        try:
            result = subprocess.run(["nmap", "--version"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, 
                                   timeout=5, check=True, text=True)
            if result.returncode == 0:
                found_path = "nmap"
                self.log("NMAP détecté dans le PATH système", tag="ok")
        except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
            self.log(f"NMAP PATH test échoué:{e}", tag="warn")
            
            # 2. Chercher dans les dossiers - DEUXIÈME TENTATIVE
            for path in possible_paths[1:]:
                try:
                    if os.path.exists(path):
                        # Vérifier que c'est vraiment exécutable
                        result = subprocess.run([path, "--version"], stdout=subprocess.PIPE, 
                                              stderr=subprocess.PIPE, timeout=5, check=True, text=True)
                        if result.returncode == 0:
                            found_path = path
                            self.log(f"NMAP trouvé à:{path}", tag="ok")
                            break
                except (subprocess.TimeoutExpired, subprocess.CalledProcessError, OSError) as e:
                    self.log(f"Chemin invalide{path}: {e}", tag="warn")
                    continue

        if found_path:
            try:
                # IMPORT DYNAMIQUE de python-nmap (lourd)
                nmap_module = _load_nmap()
                
                if found_path == "nmap":
                    self.nm = nmap_module.PortScanner()
                else:
                    self.nm = nmap_module.PortScanner(nmap_search_path=[found_path])
                
                self.nmap_path = found_path
                self.nmap_status = "ONLINE"
                self.log("NMAP INITIALIZED - READY FOR SCANNING", tag="ok")
                
                self.log(f"[OK] Nmap opérationnel ({found_path})", tag="success")
                self.enable_buttons()
            except Exception as e:
                self.log(f"[ERREUR] Nmap trouvé mais impossible à charger: {e}", tag="error")
        else:
            self.log("[CRITICAL] NMAP INTROUVABLE. Installez-le depuis nmap.org", tag="error")
            self.log("Astuce: Cochez 'Add to PATH' durant l'installation.", tag="warn")
            messagebox.showerror("Erreur Critique", "Nmap n'est pas installé ou introuvable.\nLe scan ne fonctionnera pas.")

    # ---------- LOG HELPERS ----------
    def typewriter_log(self, text, speed=0.01, tag=None):
        # Écriture directe sur le widget Tk = dangereux hors thread principal
        # (ex. appelé depuis run_nmap sur un thread d'arrière-plan). Dans ce
        # cas on bascule sur le chemin thread-safe (self.log), sans l'effet
        # machine à écrire.
        if threading.current_thread() is not threading.main_thread():
            self.log(text, tag=tag or "info")
            return
        self.text_area.insert(tk.END, "\n> ")
        # Mode performance: instant display (no typewriter effect)
        if self.performance_mode:
            self.text_area.insert(tk.END, text, tag)
            self.text_area.see(tk.END)
            self.root.update_idletasks()
        else:
            for ch in text:
                self.text_area.insert(tk.END, ch, tag)
                self.text_area.see(tk.END)
                self.root.update_idletasks()
                if speed > 0:
                    time.sleep(speed)

    def matrix_processing(self, lines=8, duration=0.8):
        """Effet visuel : defile des octets pendant une attente."""
        # Même précaution que typewriter_log : pas d'accès direct au widget
        # hors thread principal.
        if threading.current_thread() is not threading.main_thread():
            self.log("[Analyse en cours...]", tag="info")
            return

        # Mode performance: skip animations
        if self.performance_mode:
            self.text_area.insert(tk.END, "[Analyse en cours... (mode performance)]\n", "info")
            return

        prefixes = ["CPU", "MEM", "REG", "ALU", "DSP", "VPU"]
        self.text_area.insert(tk.END, "\n> [Analyse en cours...]\n", "info")
        start = time.time()
        line_count = 0

        while time.time() - start < duration and line_count < lines:
            prefix = random.choice(prefixes)
            hex_line = " ".join([f"0x{random.randint(0, 255):02X}" for _ in range(10)])
            self.text_area.insert(tk.END, f"  {prefix}: {hex_line}\n", "matrix")
            self.text_area.see(tk.END)
            self.root.update_idletasks()
            time.sleep(0.06)
            line_count += 1

        self.text_area.insert(tk.END, "> [Analyse terminée]\n", "info")

    # ---------- ACTIONS ----------
    def auto_detect(self):
        self.typewriter_log("TARGETING LOCAL SUBNET...", speed=0.02, tag="title")
        self.matrix_processing(lines=5, duration=0.5)
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
            s.close()
            target = ".".join(local_ip.split('.')[:-1]) + ".0/24"
            self.entry_ip.delete(0, tk.END)
            self.entry_ip.insert(0, target)
            # Le champ contient une vraie cible, plus le texte d'aide.
            self._placeholder_visible = False
            self.entry_ip.config(fg=THEME["accent"])
            self.log(f"TARGET ACQUIRED: {target}", tag="info")
        except Exception as exc:
            self.log(f"SIGNAL LOST: {exc}", tag="warn")

    def run_service_exposure_check(self):
        """
        Audit non intrusif: vérifie l'ouverture des services FTP (21) et SSH (22)
        sans tenter d'authentification ni d'attaque. Utile pour durcissement.
        """
        target = self.get_target().strip()
        if not target:
            messagebox.showwarning("ERR", "NO TARGET")
            return

        self.problems_found = []
        self.disable_buttons()
        # BUG #3: Progression réelle (9 ports à vérifier)
        self.start_progress_tracking(total_steps=9)
        self.text_area.delete(1.0, tk.END)
        self.log("="*60, tag="title")
        self.log(f"AUDIT D'EXPOSITION DES SERVICES SUR{target}", tag="warn")
        self.log("[*] Vérifie uniquement l'ouverture des ports; pas d'authentification.", tag="info")

        thread = threading.Thread(target=self._process_service_exposure_check, args=(target,))
        thread.start()

    def _process_service_exposure_check(self, target):
        try:
            # LISTE COMPLÈTE DES PORTS À TESTER
            ports_to_check = {
                21: ('FTP', 'Restreindre/désactiver si inutile; activer FTPS; limiter IP'),
                22: ('SSH', 'Limiter par IP; activer clés SSH; désactiver mot de passe'),
                23: ('Telnet', 'OBSOLÈTE - Désactiver immédiatement; utiliser SSH'),
                80: ('HTTP', 'Rediriger vers HTTPS; firewall si inutile'),
                443: ('HTTPS', 'Vérifier certificats valides; TLS 1.2+'),
                445: ('SMB', 'CRITIQUE - Accessible seulement en local; appliquer patches'),
                3306: ('MySQL', 'Exposé - Restreindre à localhost ou IP autorisées'),
                3389: ('RDP', 'CRITIQUE - Limiter par IP; activer Network Level Auth'),
                5900: ('VNC', 'Fortement déconseillé - Utiliser SSH tunneling'),
            }
            
            self.log("\n[AUDIT] Test de tous les services critiques...", tag="title")
            open_ports = {}
            
            for port, (service, action) in ports_to_check.items():
                try:
                    # BUG #3: Mise à jour progression réelle
                    self.root.after(0, lambda s=service: self.increment_progress(f"Test port {s}..."))
                    
                    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                        s.settimeout(2)
                        result = s.connect_ex((target, port))

                    if result == 0:
                        open_ports[port] = (service, action)
                        self.log(f"  [OPEN] Port {port:5d} - {service:10s} EXPOSÉ", tag="warn")
                        self.problems_found.append({
                            'type': 'EXPOSED SERVICE',
                            'host': target,
                            'port': port,
                            'details': f'{service} ouvert (port {port})',
                            'action': action
                        })
                    else:
                        self.log(f"  [CLOSED] Port {port:5d} - {service:10s} fermé", tag="info")
                except Exception as e:
                    self.log(f"  [ERROR] Port {port}: {e}", tag="warn")
            
            # [2] DÉTECTION DE BANNIÈRES pour les services ouverts
            self.log("\n[2] ANALYSE DES BANNIÈRES (versions détectées)...", tag="title")
            self._detect_banners(target, open_ports)
            
            # [3] VÉRIFICATION CERTIFICATS SSL/TLS
            if 443 in open_ports:
                self.log("\n[3] VÉRIFICATION SSL/TLS (Port 443)...", tag="title")
                self._check_ssl_certificate(target)
            
            # [4] RÉSOLUTION DNS
            self.log("\n[4] ANALYSE DNS...", tag="title")
            self._analyze_dns(target)
            
            # [5] RAPPORT ET SCORE DE SÉCURITÉ
            self.log("\n[5] RÉSUMÉ AUDIT...", tag="title")
            self._generate_security_score(target, open_ports)
            
            # [6] HISTORIQUE AUDIT
            self._save_audit_history(target, open_ports)
            
        finally:
            self.root.after(0, self.stop_loading)

    def _detect_banners(self, target, open_ports):
        """Détecte les bannières (versions) des services ouverts"""
        banner_ports = {22: 'SSH', 21: 'FTP', 25: 'SMTP', 80: 'HTTP'}
        
        for port in open_ports:
            if port not in banner_ports:
                continue
            
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.settimeout(3)
                    s.connect((target, port))
                    banner = s.recv(1024).decode('utf-8', errors='ignore').strip()

                if banner:
                    self.log(f"  [{port}] {banner[:100]}", tag="info")
                    self.problems_found.append({
                        'type': 'VERSION DETECTED',
                        'host': target,
                        'port': port,
                        'details': f"Bannière: {banner[:80]}",
                        'action': 'Vérifier si version est à jour et patchée'
                    })
            except:
                pass

    def _check_ssl_certificate(self, target):
        """Vérifie la validité du certificat SSL/TLS"""
        import socket as sock_module
        import ssl
        
        try:
            context = ssl.create_default_context()
            context.check_hostname = True
            context.verify_mode = ssl.CERT_REQUIRED
            
            with sock_module.create_connection((target, 443), timeout=5) as sock:
                with context.wrap_socket(sock, server_hostname=target) as ssock:
                    cert = ssock.getpeercert()
                    if cert:
                        self.log(f"  Certificat valide pour{cert.get('subject', 'Unknown')}", tag="ok")
        except ssl.SSLError as e:
            self.log(f"  Erreur SSL:{str(e)[:100]}", tag="warn")
            self.problems_found.append({
                'type': 'SSL CERTIFICATE',
                'host': target,
                'port': 443,
                'details': f'Certificat invalide ou expiré: {str(e)[:80]}',
                'action': 'Renouveler ou corriger le certificat SSL'
            })
        except Exception as e:
            self.log(f"  [-] Impossible vérifier SSL: {e}", tag="warn")

    def _analyze_dns(self, target):
        """Analyse DNS: résolution et enregistrements"""
        try:
            import socket as sock_module
            
            # Résolution inverse (IP vers hostname)
            try:
                hostname = sock_module.gethostbyaddr(target)[0]
                self.log(f"  [Reverse DNS] {target} -> {hostname}", tag="info")
            except:
                self.log(f"  [Reverse DNS] Aucun hostname trouvé pour {target}", tag="info")
            
            # Résolution normale (hostname vers IP)
            if '.' not in target.split('.')[-1] or target.replace('.', '').isdigit():
                # C'est une IP, pas un hostname
                pass
            else:
                try:
                    ip = sock_module.gethostbyname(target)
                    self.log(f"  [Forward DNS] {target} -> {ip}", tag="info")
                except:
                    self.log(f"  [Forward DNS] Impossible résoudre {target}", tag="warn")
        except Exception as e:
            self.log(f"  [-] Erreur DNS: {e}", tag="warn")

    def _generate_security_score(self, target, open_ports):
        """Génère un score de sécurité (0-100)"""
        score = 100
        critical_ports = {445, 3389, 23}
        high_risk_ports = {21, 3306, 5900}
        
        # Pénalités par port
        for port in open_ports:
            if port in critical_ports:
                score -= 25
            elif port in high_risk_ports:
                score -= 15
            elif port in [80, 443]:
                score -= 5  # Normal mais à surveiller
            else:
                score -= 10
        
        # Vérifier si au moins HTTPS est accessible
        if 443 not in open_ports and 80 in open_ports:
            score -= 10
        
        score = max(0, min(100, score))  # Clamper entre 0-100
        
        # Afficher le score
        if score >= 80:
            tag = "ok"
            status = "SÉCURISÉ"
        elif score >= 60:
            tag = "info"
            status = "ACCEPTABLE"
        else:
            tag = "warn"
            status = "RISQUÉ"
        
        self.log(f"\n SCORE DE SÉCURITÉ: {score}/100 - {status}", tag=tag)
        self.log(f"Services exposés: {len(open_ports)}", tag="info")

    def _save_audit_history(self, target, open_ports):
        """Sauvegarde l'audit dans l'historique"""
        import datetime as dt
        
        if not hasattr(self, 'audit_history'):
            self.audit_history = []
        
        audit_record = {
            'timestamp': dt.datetime.now().isoformat(),
            'target': target,
            'open_ports': list(open_ports.keys()),
            'num_exposed': len(open_ports)
        }
        
        self.audit_history.append(audit_record)
        self.log(f"\n Audit sauvegardé dans l'historique", tag="ok")

    def _boutons_actifs(self):
        """Boutons de la grille encore presents a l'ecran."""
        vivants = []
        for bouton in getattr(self, "_boutons_generes", []):
            try:
                if bouton.winfo_exists():
                    vivants.append(bouton)
            except Exception:
                continue
        self._boutons_generes = vivants
        return vivants

    def disable_buttons(self):
        """Grise les boutons pendant un scan.

        Cette methode parcourait 49 attributs (self.btn_fast, self.btn_heatmap,
        ...) qui n'existent plus depuis que les boutons sont generes dans une
        grille : getattr(..., None) renvoyait None pour chacun et la boucle ne
        faisait donc rien. On agit desormais sur les boutons reellement crees.
        """
        for bouton in self._boutons_actifs():
            try:
                bouton.config(state="disabled")
            except Exception:
                continue

    def enable_buttons(self):
        """Reactive les boutons a la fin d'un scan."""
        for bouton in self._boutons_actifs():
            try:
                bouton.config(state="normal")
            except Exception:
                continue

    def set_status(self, message):
        """Affiche un message dans la barre d'etat (sans planter si absente).

        Remplace les anciens self.btn_*.config(...) : les boutons sont
        desormais generes dynamiquement dans une grille et ne sont plus
        accessibles par attribut.
        """
        if hasattr(self, "lbl_status"):
            try:
                self.lbl_status.config(text=message)
            except Exception:
                pass

    def request_cancel(self):
        """Demande l'arret du scan en cours (bouton STOP).

        Les scans nmap deja lances vont a leur terme (python-nmap ne permet
        pas de tuer proprement le sous-processus), mais l'analyse s'arrete
        ensuite : plus aucun hote, port ou cible supplementaire n'est traite.
        """
        if getattr(self, "scan_cancel", None) is None or self.scan_cancel.is_set():
            return
        self.scan_cancel.set()
        self.log("[STOP] Arret demande - fin de l'etape en cours puis arret.", tag="warn")
        if hasattr(self, "lbl_status"):
            self.lbl_status.config(text="Arret demande...")

    def get_target(self):
        """Cible saisie, en ecartant le texte d'aide du champ."""
        value = self.entry_ip.get().strip()
        return "" if value == self.TARGET_PLACEHOLDER else value

    def start_scan(self, scan_type):
        target = self.get_target()
        if not target:
            messagebox.showwarning("ERR", "NO TARGET")
            return

        self.scan_cancel.clear()
        self.problems_found = []
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log(f"EXECUTING {scan_type.upper()} PROTOCOL ON {target}...", speed=0.01, tag="title")
        self.matrix_processing(lines=6, duration=0.6)

        thread = threading.Thread(target=self.run_nmap, args=(target, scan_type))
        thread.start()

    def correlate_cve(self, target):
        """Signale les CVE connues touchant les services detectes.

        Ne tente aucune exploitation : elle met en correspondance produit +
        version avec la base NVD (mise en cache localement) et cree un constat
        par CVE, du plus grave au moins grave. La verification d'exploitabilite
        reelle reste manuelle -- ces constats sont des pistes qualifiees.
        """
        if self.cve_db is None:
            return
        if not self.nm:
            return

        self.log("\n[CVE] Corrélation avec la base de vulnérabilités (NVD)...",
                 tag="title")

        # On regroupe par (produit, version) pour ne pas interroger deux fois
        # la meme chose, et on retient ou chaque service est expose.
        services = {}
        for host in self.nm.all_hosts():
            for proto in self.nm[host].all_protocols():
                for port, data in self.nm[host][proto].items():
                    if data.get("state") != "open":
                        continue
                    product = (data.get("product") or "").strip()
                    version = (data.get("version") or "").strip()
                    if not product or not version:
                        continue
                    services.setdefault((product, version), []).append(
                        (host, port, data.get("name", "")))

        if not services:
            self.log("[CVE] Aucun service versionné détecté "
                     "(utilisez un scan complet pour la détection de version).",
                     tag="info")
            return

        total = 0
        for (product, version), locations in services.items():
            try:
                matches = self.cve_db.correlate(product, version)
            except Exception as exc:
                self.log(f"[CVE] Erreur pour {product} {version} : {exc}",
                         tag="warn")
                continue

            if not matches:
                self.log(f"{product} {version} : aucune CVE connue",
                         tag="ok")
                continue

            self.log(f"{product} {version} : {len(matches)} CVE connue(s)",
                     tag="warn")
            for cve in matches[:15]:  # on borne l'affichage aux plus graves
                for host, port, name in locations:
                    total += 1
                    score = cve.get("score")
                    score_txt = f"CVSS {score}" if score is not None else "score inconnu"
                    self.problems_found.append({
                        "type": "CVE",
                        "host": host,
                        "port": port,
                        "service": f"{product} {version} ({name})".strip(),
                        "severity": cve.get("severity", "INCONNU"),
                        "cve_id": cve.get("cve_id"),
                        "reference": cve.get("url"),
                        "details": (f"{cve.get('cve_id')} ({score_txt}) — "
                                    f"{(cve.get('summary') or '').strip()[:240]}"),
                        "action": ("Vérifier l'applicabilité, puis mettre à jour "
                                   f"{product} ou appliquer le correctif éditeur. "
                                   f"Détails : {cve.get('url')}"),
                    })

        if total:
            self.log(f"[CVE] {total} constat(s) issus de la corrélation.",
                     tag="warn")
        else:
            self.log("[CVE] Aucune CVE connue pour les services détectés.",
                     tag="success")

    def _snapshot_scan(self):
        """Etat des hotes et ports ouverts, sous une forme comparable.

        Structure volontairement simple et serialisable en JSON, pour que deux
        scans puissent etre compares port par port.
        """
        snapshot = {"hosts": {}, "problems": list(self.problems_found)}
        if not self.nm:
            return snapshot

        for host in self.nm.all_hosts():
            ports = {}
            for proto in self.nm[host].all_protocols():
                for port, data in self.nm[host][proto].items():
                    if data.get("state") == "open":
                        name = data.get("name", "")
                        version = f"{data.get('product', '')} {data.get('version', '')}".strip()
                        ports[str(port)] = {"service": name, "version": version}
            snapshot["hosts"][host] = {"ports": ports}
        return snapshot

    # ==========================================================================
    # DÉRIVE RÉSEAU : comparer chaque scan au précédent et alerter
    # ==========================================================================

    #: Fichier conservant, par cible, l'etat du dernier scan.
    BASELINE_FILE = "scan_baselines.json"

    def _baseline_key(self, target):
        """Clef stable et sans caractere problematique pour une cible."""
        import re
        return re.sub(r"[^A-Za-z0-9._/-]", "_", (target or "").strip()) or "inconnue"

    def _load_baseline(self, target):
        """Dernier etat connu de cette cible, ou None si jamais scannee."""
        try:
            with open(self.BASELINE_FILE, "r", encoding="utf-8") as handle:
                return json.load(handle).get(self._baseline_key(target))
        except FileNotFoundError:
            return None
        except Exception as exc:
            self.log(f"[DÉRIVE] Référence illisible ({exc}) : "
                     f"comparaison impossible cette fois.", tag="warn")
            return None

    def _store_baseline(self, target, snapshot):
        """Conserve l'etat courant pour servir de reference au prochain scan."""
        try:
            try:
                with open(self.BASELINE_FILE, "r", encoding="utf-8") as handle:
                    baselines = json.load(handle)
            except (FileNotFoundError, ValueError):
                baselines = {}
            baselines[self._baseline_key(target)] = snapshot
            with open(self.BASELINE_FILE, "w", encoding="utf-8") as handle:
                json.dump(baselines, handle, indent=2)
        except Exception as exc:
            self.log(f"[DÉRIVE] Référence non enregistrée : {exc}", tag="warn")

    def detect_drift(self, previous, current):
        """Changements entre deux photographies de scan.

        Retourne une liste de dictionnaires {kind, host, port, service}, du
        plus notable au moins notable. Un ensemble vide signifie "rien n'a
        bouge", ce qui est en soi une information utile.
        """
        old_hosts = (previous or {}).get("hosts", {})
        new_hosts = (current or {}).get("hosts", {})
        changes = []

        for host in sorted(set(new_hosts) - set(old_hosts)):
            changes.append({"kind": "hote_apparu", "host": host,
                            "port": None, "service": None})
        for host in sorted(set(old_hosts) - set(new_hosts)):
            changes.append({"kind": "hote_disparu", "host": host,
                            "port": None, "service": None})

        for host in sorted(set(old_hosts) & set(new_hosts)):
            before = (old_hosts[host].get("ports") or {})
            after = (new_hosts[host].get("ports") or {})
            for port in sorted(set(after) - set(before), key=int):
                changes.append({"kind": "port_ouvert", "host": host, "port": port,
                                "service": after[port].get("service")})
            for port in sorted(set(before) - set(after), key=int):
                changes.append({"kind": "port_ferme", "host": host, "port": port,
                                "service": before[port].get("service")})
        return changes

    @staticmethod
    def format_drift(changes, target):
        """Resume lisible, utilise aussi bien dans le journal que dans l'alerte."""
        labels = {
            "hote_apparu": "Nouvel hôte",
            "hote_disparu": "Hôte disparu",
            "port_ouvert": "Port ouvert",
            "port_ferme": "Port fermé",
        }
        lines = [f"Changements détectés sur {target} :", ""]
        for change in changes:
            label = labels.get(change["kind"], change["kind"])
            if change["port"]:
                service = change["service"] or "service inconnu"
                lines.append(f"  • {label} — {change['host']}:{change['port']} ({service})")
            else:
                lines.append(f"  • {label} — {change['host']}")
        return chr(10).join(lines)

    def check_drift_and_alert(self, target):
        """Compare le scan qui vient de finir au precedent, et alerte si besoin.

        Un scanner ponctuel dit ce qui est ouvert ; savoir ce qui a CHANGE
        depuis la derniere fois est ce qui rend un audit exploitable dans la
        duree. Les notificateurs email/Slack/Discord existaient mais n'etaient
        appeles nulle part : aucune alerte n'a jamais pu partir.
        """
        current = (self.scan_history[-1].get("results")
                   if self.scan_history else self._snapshot_scan())

        # Reference : le scan precedent de la session s'il existe, sinon celui
        # conserve sur disque. Sans cette persistance, un scan planifie (un
        # processus par execution) ne detecterait jamais aucune derive.
        if len(self.scan_history) >= 2:
            reference = self.scan_history[-2].get("results")
        else:
            reference = self._load_baseline(target)

        self._store_baseline(target, current)

        if reference is None:
            self.log("[DÉRIVE] Première analyse de cette cible : référence "
                     "enregistrée pour les prochains scans.", tag="info")
            return []

        changes = self.detect_drift(reference, current)
        if not changes:
            self.log("[DÉRIVE] Aucun changement depuis le scan précédent.", tag="ok")
            return []

        summary = self.format_drift(changes, target)
        self.log("", tag="info")
        for line in summary.splitlines():
            self.log(line, tag="warn" if "•" in line else "title")

        # Les apparitions deviennent des constats : elles doivent remonter
        # dans les rapports, pas seulement dans le journal.
        for change in changes:
            if change["kind"] in ("hote_apparu", "port_ouvert"):
                self.problems_found.append({
                    "type": "CHANGEMENT RÉSEAU",
                    "host": change["host"],
                    "port": change["port"] or "—",
                    "service": change["service"] or "Découverte",
                    "details": ("Hôte absent du scan précédent"
                                if change["kind"] == "hote_apparu"
                                else "Port ouvert depuis le scan précédent"),
                    "action": "Confirmer que ce changement est voulu",
                })

        self.send_drift_alert(target, changes, summary)
        return changes

    def send_drift_alert(self, target, changes, summary):
        """Diffuse le resume sur les canaux actives dans la configuration."""
        config = getattr(self, "config_manager", None)
        if config is None:
            return

        alerts = config.get("alerts", {}) or {}
        if not alerts.get("on_drift", False):
            self.log("[DÉRIVE] Alertes désactivées (Paramètres → Alertes).", tag="info")
            return

        title = f"Changements réseau détectés — {target}"
        sent = []

        notifier = getattr(self, "email_notifier", None)
        if notifier and (config.get("email", {}) or {}).get("enabled"):
            try:
                if notifier.send_email(title, summary):
                    sent.append("email")
            except Exception as exc:
                self.log(f"[DÉRIVE] Échec de l'envoi email : {exc}", tag="error")

        hooks = getattr(self, "webhook_notifier", None)
        if hooks:
            severity = "WARNING" if any(
                c["kind"] in ("hote_apparu", "port_ouvert") for c in changes) else "INFO"
            for channel, sender in (("slack", getattr(hooks, "send_slack", None)),
                                    ("discord", getattr(hooks, "send_discord", None))):
                if not sender or not (config.get(channel, {}) or {}).get("enabled"):
                    continue
                try:
                    if sender(title, summary, severity):
                        sent.append(channel)
                except Exception as exc:
                    self.log(f"[DÉRIVE] Échec de l'envoi {channel} : {exc}", tag="error")

        if sent:
            self.log(f"[DÉRIVE] Alerte envoyée via : {', '.join(sent)}", tag="success")
        else:
            self.log("[DÉRIVE] Aucun canal d'alerte configuré "
                     "(Paramètres → Alertes).", tag="info")

    def run_nmap(self, target, scan_type, retry_count=0, max_retries=2):
        """Exécute NMAP avec retry logic en cas d'échec"""
        
        # Déterminer les arguments selon le type de scan
        if scan_type == "fast":
            args = "-F -T4"
        elif scan_type == "full":
            # -O : detection du systeme d'exploitation (necessite les droits
            # administrateur ; nmap l'ignore proprement sinon).
            args = "-sV -O -T4"
        elif scan_type == "vuln":
            args = "-sV --script vuln -T4"
        elif scan_type == "backdoor":
            ports = ",".join(str(p) for p in self.backdoor_ports.keys())
            args = f"-p {ports},1-1000 -sV -T4 --script banner"
        else:
            args = "-sV -T4"

        # Mode furtif : le drapeau etait positionne par le bouton mais aucun
        # scan ne le lisait. Les options sont desormais vraiment ajoutees.
        if getattr(self, "stealth_mode", False):
            args = self._options_furtives(args)

        try:
            # VÉRIFIER QUE NMAP EST DISPONIBLE
            if self.nm is None:
                self.log(f"NMAP NOT INITIALIZED - Tentative de ré-initialisation...", tag="warn")
                self.init_nmap()
                
                if self.nm is None:
                    raise Exception("NMAP initialization failed after retry")
            
            self.log(f"Scan en cours ({scan_type}) sur {target}... (Tentative {retry_count + 1}/{max_retries + 1})", tag="info")
            
            # EXÉCUTER LE SCAN
            try:
                self.nm.scan(target, arguments=args)
            except Exception as scan_error:
                self.log(f"Erreur scan NMAP:{scan_error}", tag="warn")
                
                # DEUXIÈME TENTATIVE avec timeout accru
                if retry_count < max_retries:
                    self.log(f"Deuxième tentative de scan en cours...", tag="info")
                    import time
                    time.sleep(2)  # Attendre 2s avant retry
                    
                    # Réinitialiser et retry
                    self.init_nmap()
                    return self.run_nmap(target, scan_type, retry_count=retry_count + 1, max_retries=max_retries)
                else:
                    raise Exception(f"NMAP scan failed after {max_retries + 1} attempts: {scan_error}")

            # TRAITER LES RÉSULTATS
            hosts_scanned = 0
            for host in self.nm.all_hosts():
                if getattr(self, "scan_cancel", None) and self.scan_cancel.is_set():
                    self.log("[STOP] Analyse interrompue par l'utilisateur.", tag="warn")
                    break
                hosts_scanned += 1
                self.matrix_processing(lines=4, duration=0.4)
                self.log("-" * 50)
                self.log(f"MACHINE: {host} ({self.nm[host].hostname()})", tag="info")

                open_ports = 0
                for proto in self.nm[host].all_protocols():
                    ports = list(self.nm[host][proto].items())
                    open_ports += sum(1 for _, data in ports
                                      if data.get("state") == "open")
                    for idx, (port, service) in enumerate(ports):
                        if getattr(self, "scan_cancel", None) and self.scan_cancel.is_set():
                            break
                        # Animation de progression (routee via la file thread-safe :
                        # run_nmap tourne sur un thread d'arrière-plan)
                        if idx % 5 == 0:
                            self.log(f"  SCAN: {random.choice(['█','▓','▒','░'])*3} [{idx}/{len(ports)}]", tag="matrix")
                        
                        name = service.get('name', '')
                        version = service.get('version', '')
                        state = service.get('state', '')

                        if scan_type == "backdoor" and port in self.backdoor_ports and state == 'open':
                            self.problems_found.append({
                                'type': 'BACKDOOR', 'host': host, 'port': port,
                                'service': name, 'details': self.backdoor_ports[port],
                                'action': "Isoler immédiatement la machine"
                            })
                            self.log(f"CRITICAL: {self.backdoor_ports[port]} ON PORT {port}", tag="warn")
                        elif self.is_risky_service(port, name, version):
                            detail = self.get_risk_details(port, name, version)
                            self.problems_found.append({
                                'type': 'SERVICE RISQUE', 'host': host, 'port': port,
                                'service': f"{name} {version}", 'details': detail,
                                'action': "Vérifier la nécessité et sécuriser"
                            })
                            self.log(f"RISK [{port}/{proto}] {name} {version} -> {detail}", tag="warn")
                        else:
                            self.log(f"[{port}/{proto}] {state.upper()} - {name} {version}")

                        if 'script' in service:
                            for script, output in service['script'].items():
                                self.problems_found.append({
                                    'type': 'VULNERABILITE', 'host': host, 'port': port,
                                    'service': f"{name} {version}",
                                    'details': f"{script}: {output[:100]}...",
                                    'action': "Appliquer les correctifs"
                                })
                                self.log(f"VULN: {script} -> {output.strip()[:120]}...", tag="warn")
                                
                                # Détecter exploits connus
                                if any(exploit in output.lower() for exploit in ['eternalblue', 'ms17-010', 'bluekeep', 'zerologon', 'printnightmare']):
                                    self.log(f"CRITICAL EXPLOIT: {script} detected!", tag="warn")
                                    self.problems_found.append({
                                        'type': 'EXPLOIT DÉTECTÉ',
                                        'host': host,
                                        'port': port,
                                        'service': f"{name} {version}",
                                        'details': f'Faille exploitable détectée: {script}',
                                        'action': 'PATCHING URGENT - Système compromettable'
                                    })

            # Les appareils annoncent eux-memes leur modele : on ecoute.
            self.query_announcements()

            # Dates de premiere/derniere detection, pour la fiche appareil.
            self.record_seen(self.nm.all_hosts())

            # Analyser connexions externes après scan
            self.analyze_external_connections(self.nm.all_hosts())
            
            # Test exploits réseau connus
            if hosts_scanned > 0:
                self.check_network_exploits(target)
            
            # Afficher résumé final avec statut
            self.log(f"\n SCAN COMPLÉTÉ avec succès ({hosts_scanned} hôte(s) trouvé(s))", tag="ok")
            # Controle du bon sens : ce que chaque appareil expose est-il
            # coherent avec ce qu'il est ? Tourne AVANT le resume, pour que ses
            # constats y figurent.
            if hosts_scanned > 0:
                try:
                    self.afficher_bon_sens()
                except Exception as exc:
                    self.log(f"[BON SENS] Controle impossible : {exc}", tag="warn")

            self.display_problem_summary(hosts_scanned, target)

            # Enregistrement du scan. auto_save_scan et save_scan_result_real
            # existaient depuis le debut mais n'etaient appeles nulle part :
            # l'historique, la comparaison et le rejeu ne pouvaient donc jamais
            # rien afficher, faute de donnees.
            snapshot = self._snapshot_scan()
            self.auto_save_scan(target, scan_type, snapshot)
            self.save_scan_result_real(target, scan_type, self.problems_found)

            # Correlation avec la base de vulnerabilites (NVD) : quelles CVE
            # connues touchent les versions detectees. Pas d'exploitation, du
            # signalement -- comme Nessus/OpenVAS.
            if scan_type != "fast":
                self.correlate_cve(target)

            # Comparaison au scan precedent et alerte eventuelle.
            self.check_drift_and_alert(target)

            # Inventaire dans sa propre fenetre : on n'encombre pas l'ecran
            # principal pendant le scan, et le resultat reste consultable.
            self.root.after(0, lambda: self.show_inventory_window(auto=True))
            
        except Exception as exc:
            error_msg = str(exc)
            self.log(f"SCAN FAILURE:{error_msg}", tag="warn")
            
            # Si c'est un timeout ou une erreur de connection, suggérer une vérification
            if "timeout" in error_msg.lower() or "connection" in error_msg.lower():
                self.log(f"\n Conseil: Vérifier que:", tag="warn")
                self.log(f"  • La cible {target} est accessible", tag="warn")
                self.log(f"  • Le firewall n'interfère pas", tag="warn")
                self.log(f"  • NMAP est correctement installé (commande: nmap --version)", tag="warn")
            
        finally:
            self.root.after(0, self.stop_loading)

    def is_risky_service(self, port, name, version):
        risky = {21, 23, 139, 445, 3389, 5900, 8080}
        return port in risky

    def _reverse_dns(self, ip, timeout=1.0):
        """Nom d'hote associe a une IP, ou None.

        Donne a l'utilisateur de quoi reconnaitre une destination sans que le
        logiciel ait a decider a sa place si elle est legitime.
        """
        import socket as _socket
        previous = _socket.getdefaulttimeout()
        try:
            _socket.setdefaulttimeout(timeout)
            return _socket.gethostbyaddr(ip)[0]
        except Exception:
            return None
        finally:
            _socket.setdefaulttimeout(previous)

    def analyze_external_connections(self, scanned_hosts):
        """Analyse les connexions sortantes vers l'extérieur"""
        self.log("\n" + "="*50)
        self.log("[ANALYZING EXTERNAL ACCESS...]", tag="info")
        self.matrix_processing(lines=5, duration=0.5)
        
        try:
            if platform.system() == "Windows":
                result = subprocess.run(['netstat', '-ano'], capture_output=True, text=True, encoding='utf-8', errors='ignore', timeout=10)
                external_connections = []
                local_ips = set(['127.0.0.1', '0.0.0.0'])
                
                # Récupérer IP locale
                try:
                    with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
                        s.connect(("8.8.8.8", 80))
                        local_ip = s.getsockname()[0]
                    local_network = '.'.join(local_ip.split('.')[:-1])
                    local_ips.add(local_ip)
                except Exception as e:
                    local_network = '192.168.1'
                    self.log(f"[!] Détection du réseau local impossible ({e}), fallback sur 192.168.1.0/24", tag="warn")
                
                # Parser connexions
                if result and result.stdout:
                    for line in result.stdout.split('\n'):
                        if 'ESTABLISHED' in line:
                            parts = line.split()
                            if len(parts) >= 3:
                                foreign_addr = parts[2]
                                if ':' in foreign_addr:
                                    ip = foreign_addr.split(':')[0]
                                    port = foreign_addr.split(':')[1]
                                    
                                    # Ignorer localhost et réseau local
                                    if ip not in local_ips and not ip.startswith(local_network) and not ip.startswith('192.168') and not ip.startswith('10.'):
                                        external_connections.append((ip, port))
                
                # Analyser connexions externes
                if external_connections:
                    unique_ips = list(set([conn[0] for conn in external_connections]))
                    self.log(f"EXTERNAL CONNECTIONS: {len(unique_ips)} unique IPs", tag="warn")
                    
                    suspicious_ports = {'4444': 'Metasploit', '5555': 'Android Debug', '6666': 'IRC Bot', 
                                       '6667': 'IRC', '1337': 'Elite', '31337': 'BackOrifice'}
                    
                    for ip, port in external_connections[:15]:
                        conn_info = f"{ip}:{port}"
                        is_suspicious = False
                        details = f"Connexion active vers {conn_info}"
                        
                        # Vérifier ports suspects
                        if port in suspicious_ports:
                            is_suspicious = True
                            details = f"Port suspect {port} ({suspicious_ports[port]}) vers {ip}"
                            self.log(f"SUSPICIOUS: {details}", tag="warn")
                        
                        # Vérifier si IP fait partie des cibles scannées
                        elif ip in scanned_hosts:
                            details = f"Communication avec machine scannée {ip}:{port}"
                            self.log(f"LATERAL: {details}", tag="warn")
                            is_suspicious = True
                        
                        if is_suspicious:
                            self.problems_found.append({
                                'type': 'CONNEXION EXTERNE',
                                'host': 'LOCAL',
                                'port': port,
                                'service': 'Network',
                                'details': details,
                                'action': 'Analyser le trafic et bloquer si malveillant'
                            })
                    
                    # Détecter beacon/C2
                    ip_counts = {}
                    for ip, _ in external_connections:
                        ip_counts[ip] = ip_counts.get(ip, 0) + 1
                    
                    # Beaucoup de connexions vers une meme adresse est le
                    # comportement d'un canal de commande (C2)... mais aussi
                    # celui d'un navigateur, d'un client de synchronisation ou
                    # d'une mise a jour. On rapporte le fait observe, sans
                    # conclure : l'ancienne version annoncait "possible Command
                    # & Control / URGENT" pour du trafic parfaitement normal.
                    #
                    # Volontairement AUCUNE liste blanche : declarer une
                    # destination "sure" serait un mensonge commode. Un canal
                    # de commande passe tres souvent par un service cloud
                    # legitime, qui figurerait justement sur une telle liste.
                    for ip, count in sorted(ip_counts.items(),
                                            key=lambda item: item[1], reverse=True):
                        if count > 5:
                            hostname = self._reverse_dns(ip)
                            label = f"{ip} ({hostname})" if hostname else ip
                            self.log(f"[i] {count} connexions vers {label} "
                                     f"— a verifier", tag="info")
                            self.problems_found.append({
                                'type': 'CONNEXIONS RÉPÉTÉES',
                                'host': 'LOCAL',
                                'port': 'Multiple',
                                'service': hostname or 'Destination externe',
                                'details': (f'{count} connexions vers {label}. '
                                            f'Frequence typique d\'un canal de commande '
                                            f'comme d\'un usage normal (navigateur, '
                                            f'synchronisation, mise a jour).'),
                                'action': ("Identifier le programme concerné "
                                           "(onglet Processus) avant toute action")
                            })
                else:
                    self.log("No suspicious external connections", tag="ok")
                    
        except Exception as exc:
            self.log(f"External analysis error: {exc}", tag="warn")
        
        self.log("="*50)

    def check_network_exploits(self, target):
        """Test des exploits réseau connus"""
        self.log("\n[TESTING KNOWN EXPLOITS...]", tag="info")
        self.matrix_processing(lines=6, duration=0.6)
        
        exploit_tests = [
            {'name': 'SMBv1 (EternalBlue)', 'port': 445, 'risk': 'CRITICAL'},
            {'name': 'RDP BlueKeep', 'port': 3389, 'risk': 'CRITICAL'},
            {'name': 'SMB Signing', 'port': 445, 'risk': 'HIGH'},
            {'name': 'Anonymous FTP', 'port': 21, 'risk': 'MEDIUM'},
            {'name': 'Telnet Clear Text', 'port': 23, 'risk': 'HIGH'},
        ]
        
        for test in exploit_tests:
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(2)
                result = sock.connect_ex((target if not target.endswith('/24') else target.split('/')[0].rsplit('.', 1)[0] + '.1', test['port']))
                sock.close()
                
                if result == 0:
                    self.log(f"[{test['risk']}] {test['name']} - Port {test['port']} OPEN", tag="warn")
                    self.problems_found.append({
                        'type': 'FAILLE EXPLOITABLE',
                        'host': target,
                        'port': test['port'],
                        'service': test['name'],
                        'details': (f"Port {test['port']} ouvert ({test['name']}) — "
                                    f"exposition de risque {test['risk']}. "
                                    f"Vulnérabilité NON vérifiée : seule l'ouverture "
                                    f"du port a été testée."),
                        'action': 'Désactiver le service ou appliquer correctifs'
                    })
            except:
                pass

    def get_risk_details(self, port, name, version):
        risks = {
            21: "FTP transmet en clair",
            23: "Telnet obsolète, préférer SSH",
            139: "NetBIOS fuit des infos",
            445: "SMB vulnérable (EternalBlue)",
            3389: "RDP cible des attaques bruteforce",
            5900: "VNC mots de passe faibles",
            8080: "Proxy web potentiellement ouvert"
        }
        return risks.get(port, "Service à vérifier")
    
    def get_service_name(self, port):
        """Retourne le nom du service courant pour un port"""
        services = {
            21: 'FTP', 22: 'SSH', 23: 'Telnet', 25: 'SMTP',
            80: 'HTTP', 110: 'POP3', 139: 'NetBIOS', 143: 'IMAP',
            443: 'HTTPS', 445: 'SMB', 3389: 'RDP', 5900: 'VNC',
            8080: 'HTTP-Proxy', 3306: 'MySQL', 5432: 'PostgreSQL'
        }
        return services.get(port, f'Port-{port}')

    def display_problem_summary(self, hosts_scanned, target):
        self.matrix_processing(lines=3, duration=0.3)
        self.log("\n" + "=" * 60)
        self.log("BILAN DE L'ANALYSE", tag="title")
        self.log(f"Cible : {target}")
        self.log(f"Hôtes analysés : {hosts_scanned}")
        self.log(f"Constats : {len(self.problems_found)}")

        if not self.problems_found:
            self.log("Aucun constat sur les contrôles exécutés.", tag="ok")
        else:
            for i, p in enumerate(self.problems_found, 1):
                self.log(f"{i}. {p['type']} @ {p['host']}:{p['port']} ({p['service']})", tag="warn")
                self.log(f"   DETAIL: {p['details']}")
                self.log(f"   ACTION: {p['action']}")
        self.log("=" * 60)

    def check_network_health(self):
        self.problems_found = []
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log("ANALYZING NETWORK INTEGRITY...", speed=0.02, tag="title")
        self.matrix_processing(lines=7, duration=0.7)

        thread = threading.Thread(target=self.run_network_health)
        thread.start()

    def run_network_health(self):
        self.log(f"OS: {platform.system()} {platform.release()}")
        self.log("SCANNING ACTIVE CONNECTIONS...", tag="info")
        try:
            if platform.system() == "Windows":
                # UTILISATION DE LA COMMANDE SÉCURISÉE
                output = self.run_safe_command(['netstat', '-ano'])
                
                for line in output.split('\n'):
                    if 'ESTABLISHED' in line:
                        parts = line.split()
                        if len(parts) >= 3:
                            foreign = parts[2]
                            for port in self.backdoor_ports:
                                if f":{port}" in foreign:
                                    self.problems_found.append({
                                        'type': 'CONNEXION BACKDOOR', 'host': 'LOCAL', 'port': port,
                                        'service': self.backdoor_ports[port],
                                        'details': f"Connexion vers {foreign}",
                                        'action': "Couper la connexion et scanner"
                                    })
                                    self.log(f"SUSPICIOUS LINK -> {foreign}", tag="warn")
        except Exception as exc:
            self.log(f"ACCESS DENIED: {exc}", tag="warn")

        self.log("LISTENING PORTS...", tag="info")
        try:
            if platform.system() == "Windows":
                # UTILISATION DE LA COMMANDE SÉCURISÉE
                output = self.run_safe_command(['netstat', '-an'])
                
                exposed_services = []
                for line in output.split('\n'):
                    if 'LISTENING' in line or 'ECOUTE' in line:  # Correction 'ECOUTE' sans accent
                        parts = line.split()
                        if len(parts) >= 2:
                            if '0.0.0.0' in parts[1] or '::' in parts[1]:
                                try:
                                    port = int(parts[1].split(':')[-1])
                                except:
                                    continue
                                
                                if port in [21, 22, 23, 80, 139, 443, 445, 3389, 5900, 8080]:
                                    exposed_services.append(port)
                                    self.log(f"EXPOSED TO INTERNET: Port {port} on 0.0.0.0", tag="warn")
                                    self.problems_found.append({
                                        'type': 'SERVICE EXPOSÉ', 'host': 'LOCAL', 'port': port,
                                        'service': self.get_service_name(port),
                                        'details': 'Accessible depuis Internet',
                                        'action': 'Restreindre accès au réseau local'
                                    })
                                
                                if port in self.backdoor_ports:
                                    self.problems_found.append({
                                        'type': 'PORT BACKDOOR', 'host': 'LOCAL', 'port': port,
                                        'service': self.backdoor_ports[port],
                                        'details': "Service backdoor en écoute",
                                        'action': "Identifier et stopper le processus"
                                    })
                                    self.log(f"BACKDOOR LISTENING: {port}", tag="warn")
                            else:
                                try:
                                    port = int(parts[1].split(':')[-1])
                                    if port in self.backdoor_ports:
                                        self.log(f"LOCAL BACKDOOR: {port}", tag="warn")
                                except: pass
        except Exception as exc:
            self.log(f"ACCESS DENIED: {exc}", tag="warn")

        self.matrix_processing(lines=4, duration=0.4)
        self.log(f"THREATS DETECTED: {len(self.problems_found)}", tag="title")
        self.root.after(0, self.stop_loading)


    def start_loading(self):
        """Démarre la barre de progression et désactive les boutons"""
        self.progress.start()
        self.disable_buttons()

    def stop_loading(self):
        self.progress.stop()
        self.enable_buttons()
        messagebox.showinfo("Fini", "Analyse terminée !")

    # =========================================================================
    # PARTIE 3: AUDIT SYSTÈME INTERNALS - WMI WINDOWS (VRAIE DONNÉES)
    # =========================================================================
    def audit_system_internals(self):
        """Audit profond: hotfixes, utilisateurs, logiciels obsolètes (WMI)."""
        self.disable_buttons()
        self.text_area.delete(1.0, tk.END)
        self.log("[AUDIT] SCAN INTÉGRITÉ SYSTÈME (WMI)...", tag="title")
        
        def _run_wmi_audit():
            # QUERY 1: Hotfixes (sécurité patches)
            try:
                result = subprocess.run(
                    ["powershell", "-Command", "Get-HotFix | Select-Object -Property Description,HotFixID | ConvertTo-Json"],
                    capture_output=True, text=True, timeout=10
                )
                if result.stdout:
                    hotfixes = json.loads(result.stdout)
                    if not isinstance(hotfixes, list):
                        hotfixes = [hotfixes]
                    self.log(f"[+] {len(hotfixes)} patches de sécurité détectés.", tag="info")
                    for hf in hotfixes[-5:]:  # Afficher les 5 derniers
                        self.log(f"    • {hf.get('HotFixID', 'N/A')} - {hf.get('Description', '')}", tag="ok")
                    # Détection: patches manquants = mauvais signe
                    if len(hotfixes) < 10:
                        self.log(f"ALERTE: Seulement{len(hotfixes)} patches! Trop peu pour un système en production.", tag="error")
                        self.problems_found.append({
                            'type': 'LOW_PATCH_COUNT',
                            'details': f'{len(hotfixes)} patches détectés (< 10 attendus)',
                            'action': 'Exécuter Windows Update dès que possible'
                        })
            except Exception as e:
                self.log(f"[-] Erreur WMI Hotfixes: {e}", tag="error")
            
            # QUERY 2: Utilisateurs locaux (guest actifs = danger)
            try:
                result = subprocess.run(
                    ["powershell", "-Command", "Get-LocalUser | Select-Object -Property Name,Enabled | ConvertTo-Json"],
                    capture_output=True, text=True, timeout=10
                )
                if result.stdout:
                    users = json.loads(result.stdout)
                    if not isinstance(users, list):
                        users = [users]
                    self.log(f"[+] {len(users)} utilisateurs locaux détectés.", tag="info")
                    for user in users:
                        status = "ACTIF" if user['Enabled'] else "DÉSACTIVÉ"
                        tag = "error" if (user['Enabled'] and 'Guest' in user['Name']) else "ok"
                        self.log(f"    • {user['Name']} ({status})", tag=tag)
                    # Alerte: Guest account activé = VRAIMENT mauvais
                    guest_active = any(u['Name'] == 'Guest' and u['Enabled'] for u in users)
                    if guest_active:
                        self.log("CRITIQUE: Compte Guest activé! C'est une faille béante.", tag="error")
                        self.problems_found.append({
                            'type': 'GUEST_ACCOUNT_ACTIVE',
                            'details': 'Compte Guest Windows activé',
                            'action': 'Désactiver immédiatement (net user Guest /active:no)'
                        })
            except Exception as e:
                self.log(f"[-] Erreur WMI Utilisateurs: {e}", tag="error")
            
            # QUERY 3: Inventaire logiciels (obsolètes = dangereux)
            try:
                result = subprocess.run(
                    ["powershell", "-Command", "Get-ItemProperty 'HKLM:\\Software\\Microsoft\\Windows\\CurrentVersion\\Uninstall\\*' | Select-Object -Property DisplayName,DisplayVersion | ConvertTo-Json"],
                    capture_output=True, text=True, timeout=15
                )
                if result.stdout:
                    software = json.loads(result.stdout)
                    if not isinstance(software, list):
                        software = [software]
                    self.log(f"[+] {len(software)} logiciels installés trouvés.", tag="info")
                    
                    # Logiciels OBSOLÈTES = danger critique
                    obsolete_keywords = ['Java 7', 'Java 6', 'Flash', 'Python 2.7', 'Internet Explorer', 'QuickTime']
                    suspicious = []
                    for sw in software:
                        name = sw.get('DisplayName', '')
                        for obs in obsolete_keywords:
                            if obs.lower() in name.lower():
                                suspicious.append(sw)
                                self.log(f"    DANGER:{name} ({sw.get('DisplayVersion', '?')})", tag="error")
                    
                    if suspicious:
                        self.log(f"ALERTE:{len(suspicious)} logiciel(s) obsolète(s) détecté(s)!", tag="error")
                        for sw in suspicious:
                            self.problems_found.append({
                                'type': 'OBSOLETE_SOFTWARE',
                                'details': f"{sw.get('DisplayName', 'Unknown')} ({sw.get('DisplayVersion', 'N/A')})",
                                'action': 'Désinstaller ce logiciel ou le mettre à jour'
                            })
                    else:
                        self.log("Aucun logiciel obsolète détecté.", tag="success")
            except Exception as e:
                self.log(f"[-] Erreur WMI Logiciels: {e}", tag="error")
            
            self.log("\n[OK] Audit système WMI terminé.", tag="success")
            self.root.after(0, self.stop_loading)

        threading.Thread(target=_run_wmi_audit, daemon=True).start()

    # =========================================================================
    # PARTIE 4: INTERACTIVE HONEYPOT - CAPTURE DE CREDENTIALS RÉELLE
    # =========================================================================
    def deploy_honeypot_plus(self):
        """Déploie un honeypot interactif qui simule un login Windows."""
        self.disable_buttons()
        self.text_area.delete(1.0, tk.END)
        self.log("[HONEYPOT] DÉPLOIEMENT D'UNE FAUSSE INTERFACE DE LOGIN...", tag="title")
        
        def start_trap():
            """Démarre un vrai serveur socket qui capture les credentials."""
            honeypot_socket = None
            client_socket = None
            
            try:
                honeypot_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                honeypot_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                honeypot_socket.bind(('0.0.0.0', 8888))
                honeypot_socket.listen(5)
                honeypot_socket.settimeout(30)
                
                self.log("[+] Honeypot activé sur 0.0.0.0:8888", tag="success")
                self.log("    En attente de connexion (timeout 30 sec)...", tag="info")
                
                connections = 0
                while connections < 5:
                    try:
                        client_socket, addr = honeypot_socket.accept()
                        connections += 1
                        self.log(f"[!] Connexion détectée de {addr[0]}:{addr[1]}", tag="warn")
                        
                        # SÉCURITÉ: Valider l'adresse reçue
                        if not addr or not addr[0]:
                            self.log("    └─ Adresse invalide détectée, connexion ignorée", tag="warn")
                            continue
                        
                        # Envoyer faux prompt Windows
                        prompt = b"Microsoft Windows [Version 10.0.19041]\n"
                        prompt += b"(c) Microsoft Corporation. Tous droits reserv\xe9s.\n\n"
                        prompt += b"C:\\Users\\Administrator> "
                        
                        # SÉCURITÉ: Gérer les erreurs d'envoi
                        try:
                            client_socket.send(prompt)
                        except (ConnectionResetError, BrokenPipeError, OSError) as send_err:
                            self.log(f"    └─ Erreur d'envoi:{send_err}", tag="warn")
                            continue
                        
                        # Recevoir la commande (ou credential) de l'attaquant
                        try:
                            # SÉCURITÉ: Limiter la taille reçue à 1024 bytes max
                            client_socket.settimeout(5)  # Timeout de 5 sec pour recv
                            data = client_socket.recv(1024)
                            
                            if data:
                                # SÉCURITÉ: Validation et sanitization
                                try:
                                    command = data.decode('utf-8', errors='ignore').strip()
                                    # Limiter la longueur affichée
                                    command_display = command[:200] if len(command) > 200 else command
                                    self.log(f"    └─ Entrée capturée: {command_display}", tag="error")
                                    
                                    # Déterminer si c'est un attempt de login
                                    if 'username' in command.lower() or 'password' in command.lower():
                                        self.problems_found.append({
                                            'type': 'HONEYPOT_HIT',
                                            'details': f'Credential attempt from {addr[0]}: {command_display}',
                                            'action': 'Ajouter IP à liste noire du firewall'
                                        })
                                        self.log(f"    └─ CREDENTIAL ATTEMPT CAPTURÉ!", tag="error")
                                    
                                    # Répondre avec erreur (comme Windows)
                                    response = b"Access denied. Your attempt has been logged.\n"
                                    try:
                                        client_socket.send(response)
                                    except (ConnectionResetError, BrokenPipeError, OSError):
                                        pass  # Client déconnecté
                                        
                                except (UnicodeDecodeError, ValueError) as decode_err:
                                    self.log(f"    └─ Données corrompues reçues:{decode_err}", tag="warn")
                                    
                        except socket.timeout:
                            self.log(f"    └─ Timeout de réception pour {addr[0]}", tag="info")
                        except (ConnectionResetError, BrokenPipeError, OSError) as recv_err:
                            self.log(f"    └─ Connexion perdue:{recv_err}", tag="warn")
                            
                    except socket.timeout:
                        # Timeout global du honeypot
                        break
                    except Exception as e:
                        self.log(f"[-] Erreur lors de la connexion honeypot: {type(e).__name__}: {e}", tag="error")
                    finally:
                        # SÉCURITÉ: Toujours fermer proprement le socket client
                        if client_socket:
                            try:
                                client_socket.close()
                            except:
                                pass
                            client_socket = None
                
                self.log(f"\n[OK] Honeypot session terminée ({connections} connexions).", tag="success")
                
            except PermissionError:
                self.log("[!] Port 8888 nécessite les droits admin.", tag="error")
                self.log("    Tentative sur port 18888 (non-privilegié)...", tag="warn")
                
                # Fermer le socket précédent si ouvert
                if honeypot_socket:
                    try:
                        honeypot_socket.close()
                    except:
                        pass
                
                try:
                    honeypot_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    honeypot_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                    honeypot_socket.bind(('0.0.0.0', 18888))
                    honeypot_socket.listen(3)
                    honeypot_socket.settimeout(10)
                    self.log("[+] Honeypot fallback activé sur 0.0.0.0:18888", tag="success")
                    
                    try:
                        client_socket, addr = honeypot_socket.accept()
                        self.log(f"[!] Connexion piégée de {addr[0]}!", tag="warn")
                        self.problems_found.append({
                            'type': 'HONEYPOT_HIT',
                            'details': f'Attaquant détecté via honeypot: {addr[0]}',
                            'action': 'Ajouter {addr[0]} à la liste noire'
                        })
                        client_socket.close()
                    except socket.timeout:
                        self.log("[OK] Pas de connexion honeypot après 10 sec.", tag="info")
                    finally:
                        honeypot_socket.close()
                except Exception as e2:
                    self.log(f"[-] Erreur honeypot fallback: {e2}", tag="error")
            except Exception as e:
                self.log(f"[-] Erreur honeypot déploiement: {e}", tag="error")
            finally:
                # Le chemin nominal (boucle terminée normalement ou timeout
                # global) ne fermait jamais honeypot_socket -- fait ici, quel
                # que soit le chemin de sortie (close() sur un socket déjà
                # fermé est un no-op, donc pas de risque de double-close).
                if honeypot_socket:
                    try:
                        honeypot_socket.close()
                    except Exception:
                        pass
                self.root.after(0, self.stop_loading)

        threading.Thread(target=start_trap, daemon=True).start()

    # =========================================================================
    # PARTIE 5: MODE PERFORMANCE - DESACTIVER LES ANIMATIONS
    # =========================================================================
    def toggle_performance(self):
        """Active/Désactive le mode performance (animations et effets)."""
        self.performance_mode = not self.performance_mode
        status = "ACTIVÉ" if self.performance_mode else "DÉSACTIVÉ"
        self.log(f"\n MODE PERFORMANCE: {status}", tag="warn" if self.performance_mode else "info")
        if self.performance_mode:
            self.log("   → Animations décoratives désactivées", tag="info")
            self.log("   → Affichage Typewriter instantané", tag="info")
            self.log("   → CPU efficacité maximale", tag="success")
        else:
            self.log("   → Animations décoratives réactivées", tag="info")
            self.log("   → Affichage Typewriter standard", tag="info")


    # ========== NOUVELLES FONCTIONNALITES AVANCEES ==========

    def detect_advanced_exploits(self):
        """Détecte les exploits avancés (EternalBlue, Mimikatz, ZeroLogon, etc.)"""
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        
        # ANIMATIONS!
        self.show_scan_radar()
        self.animate_log_fadein("SCANNING FOR ADVANCED EXPLOITS...", tag="title")
        self.matrix_processing(lines=5, duration=0.5)
        
        thread = threading.Thread(target=self._run_advanced_exploit_scan)
        thread.start()

    def _run_advanced_exploit_scan(self):
        """VRAI scan de vulnérabilités critiques via Nmap Scripts"""
        target = self.get_target()
        if not self.nm:
            self.log("Erreur: Scanner non initialisé.", tag="error")
            return

        try:
            self.log("="*60)
            self.log(f"[REAL] ANALYSE DES VULNÉRABILITÉS CRITIQUES SUR {target}", tag="title")
            self.log("Ceci peut prendre 1 à 5 minutes. Patience...", tag="info")
            self.matrix_processing(lines=3, duration=0.3)

            # On utilise les scripts NSE officiels de Nmap pour détecter les failles
            # smb-vuln* cherche les failles Windows SMB (EternalBlue, etc.)
            # http-vuln* cherche les failles Web
            arguments = '-sV --script "smb-vuln-ms17-010,smb-vuln-cve2009-3103,http-vuln-cve2017-5638" -T4'
            
            self.nm.scan(target, arguments=arguments)
            
            vuln_found = False
            
            for host in self.nm.all_hosts():
                self.log(f"\nANALYSE HÔTE: {host}", tag="info")
                
                # Vérifier les résultats des scripts
                if 'host_script' in self.nm[host]:
                    for script in self.nm[host]['host_script']:
                        output = script.get('output', '')
                        if "VULNERABLE" in output or "Risk factor: HIGH" in output:
                            vuln_found = True
                            exploit_name = script.get('id', 'Unknown Exploit')
                            self.log(f"[DANGER] FAILLE DÉTECTÉE:{exploit_name}", tag="error")
                            self.log(f"Détails:\n{output}", tag="warn")
                            
                            self.problems_found.append({
                                'type': 'FAILLE CRITIQUE',
                                'host': host,
                                'service': exploit_name,
                                'details': "Vulnérabilité confirmée par Nmap Script",
                                'action': 'PATCHER IMMÉDIATEMENT'
                            })
                        else:
                            self.log(f"Check{script.get('id')}: SÉCURISÉ", tag="success")
                else:
                    self.log("Aucun script n'a retourné de résultat (Probablement sécurisé ou bloqué).", tag="info")

            if not vuln_found:
                self.log("\n[RÉSULTAT] Aucune faille critique (SMB/HTTP) détectée avec ces scripts.", tag="success")

        except Exception as exc:
            self.log(f"ERREUR SCAN: {exc}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    def visualize_network(self):
        """Crée une visualisation graphique du réseau"""
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log("BUILDING NETWORK TOPOLOGY...", speed=0.02, tag="title")
        self.matrix_processing(lines=4, duration=0.4)
        
        thread = threading.Thread(target=self._build_network_graph)
        thread.start()

    def _build_network_graph(self):
        """Génère une topologie réseau basée sur les VRAIS hôtes détectés"""
        try:
            self.log("[INIT] CARTOGRAPHIE DU RÉSEAU...", tag="title")
            
            # On vérifie si un scan a déjà été fait
            if not self.nm or not self.nm.all_hosts():
                self.log("Aucune donnée. Lancement d'un scan rapide de découverte...", tag="warn")
                # On lance un ping scan rapide (nécessite Nmap)
                target = self.get_target()
                self.nm.scan(target, arguments='-sn -T4') # Ping scan seulement
            
            hosts = self.nm.all_hosts()
            total_hosts = len(hosts)
            
            if total_hosts == 0:
                self.log("Aucun hôte trouvé. Vérifiez la cible ou le pare-feu.", tag="error")
                return

            self.log(f"Construction du graphe pour {total_hosts} appareils détectés...", tag="info")
            self.matrix_processing(lines=3, duration=0.3)

            # Détection de la passerelle (Gateway) - Souvent l'IP finissant par .1 ou .254
            gateway = next((h for h in hosts if h.endswith('.1') or h.endswith('.254')), "Inconnue")

            # Construction du dessin ASCII dynamique
            graph_lines = []
            graph_lines.append("╔════════════════════ NETWORK TOPOLOGY ════════════════════╗")
            graph_lines.append("║                                                          ║")
            graph_lines.append("║      [ INTERNET ]                                        ║")
            graph_lines.append("║           │                                              ║")
            graph_lines.append(f"║      [ GATEWAY ] ─> {gateway:<15}                      ║")
            graph_lines.append("║           │                                              ║")
            
            # Ajouter les hôtes trouvés (limité à 10 pour l'affichage ASCII pour ne pas casser l'écran)
            for i, host in enumerate(hosts[:10]):
                status = self.nm[host].state()
                try:
                    hostname = self.nm[host].hostname() or "Unknown Device"
                except: hostname = "Unknown"
                
                # Arbre visuel
                connector = "├──" if i < len(hosts[:10]) - 1 else "└──"
                line = f"║           {connector} [ {host:<15} ] ({hostname[:15]})"
                graph_lines.append(line)

            if len(hosts) > 10:
                graph_lines.append(f"║               ... et {len(hosts)-10} autres appareils")
            
            graph_lines.append("║                                                          ║")
            graph_lines.append("╚══════════════════════════════════════════════════════════╝")

            # Affichage
            for line in graph_lines:
                self.log(line, tag="matrix")

            # Résumé technique
            self.log(f"\n[Rapport Topologie]", tag="title")
            self.log(f"Total Appareils: {total_hosts}", tag="success")
            self.log(f"Cible analysée : {self.get_target()}", tag="info")

        except Exception as exc:
            self.log(f"ERREUR GRAPHIQUE: {exc}", tag="error")
            self.log("Assurez-vous que Nmap est bien installé.", tag="warn")
        finally:
            self.root.after(0, self.stop_loading)

    def shodan_lookup(self):
        """Intégration Shodan pour vérifier les ports exposés sur internet"""
        # Demander la clé API Shodan
        api_key = tk.simpledialog.askstring("Shodan API", "Entrez votre clé API Shodan:\n(Laissez vide pour démo)")
        
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log("QUERYING SHODAN DATABASE...", speed=0.02, tag="title")
        self.matrix_processing(lines=5, duration=0.5)
        
        thread = threading.Thread(target=self._shodan_scan, args=(api_key,))
        thread.start()

    def _shodan_scan(self, api_key):
        """
        Recherche Shodan RÉELLE.
        """
        if not api_key or api_key.strip() == "":
            self.log("ERREUR : Aucune clé API Shodan fournie.", tag="error")
            self.log("ℹ Shodan est un service payant/restreint. Sans clé, impossible de scanner légalement.", tag="info")
            self.log("Obtenez une clé sur : https://account.shodan.io", tag="accent")
            self.root.after(0, self.stop_loading)
            return

        try:
            import requests
            target_ip = self.get_target()
            # Nettoyage si format CIDR
            if "/" in target_ip: target_ip = target_ip.split("/")[0]

            self.log(f"[*] Interrogation de l'API Shodan pour {target_ip}...", tag="title")
            
            url = f"https://api.shodan.io/shodan/host/{target_ip}?key={api_key}"
            response = requests.get(url, timeout=10)

            if response.status_code == 200:
                data = response.json()
                self.log(f"\n[+] Organisation : {data.get('org', 'N/A')}", tag="ok")
                self.log(f"[+] OS : {data.get('os', 'N/A')}", tag="ok")
                self.log(f"[+] Ports ouverts (Internet) : {data.get('ports', [])}", tag="warn")
                
                for item in data.get('data', []):
                    port = item.get('port')
                    product = item.get('product', 'Unknown')
                    self.log(f"    - Port {port}: {product}", tag="info")
            elif response.status_code == 404:
                self.log("[-] IP non trouvée dans la base Shodan (Bon signe ?)", tag="success")
            elif response.status_code == 401:
                self.log("[-] Clé API invalide.", tag="error")
            else:
                self.log(f"[-] Erreur API : {response.status_code}", tag="error")

        except Exception as e:
            # requests embarque parfois l'URL complète (donc la clé API) dans le
            # message d'erreur -- on la masque avant de logguer/exporter.
            import re
            safe_msg = re.sub(r'key=[^&\s]+', 'key=***', str(e))
            self.log(f"Erreur connexion : {safe_msg}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    # -------------------------------------------------------------------------
    # [MODULAR TRAFFIC ANALYZER]
    # Trois moteurs au choix selon le besoin : vitesse, precision ou instantane.
    # -------------------------------------------------------------------------


    # ========== 8 NOUVELLES FONCTIONNALITES AVANCEES ==========

    def export_reports(self):
        """Exporte les rapports en PDF, Excel ou JSON"""
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log("GENERATING EXPORT REPORTS...", speed=0.02, tag="title")
        
        thread = threading.Thread(target=self._export_all_formats)
        thread.start()

    def _export_all_formats(self):
        """Exporte dans tous les formats disponibles"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # JSON Export (toujours possible)
            json_file = f"scan_report_{timestamp}.json"
            json_data = {
                'timestamp': datetime.now().isoformat(),
                'target': self.get_target(),
                'problems': self.problems_found,
                'total_issues': len(self.problems_found)
            }
            with open(json_file, 'w') as f:
                json.dump(json_data, f, indent=2)
            self.log(f"JSON Export:{json_file}", tag="ok")
            
            # CSV Export
            csv_file = f"scan_report_{timestamp}.csv"
            try:
                with open(csv_file, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Type', 'Host', 'Port', 'Service', 'Details', 'Action'])
                    for issue in self.problems_found:
                        writer.writerow([issue.get('type', ''), issue.get('host', ''), 
                                        issue.get('port', ''), issue.get('service', ''),
                                        issue.get('details', ''), issue.get('action', '')])
                self.log(f"CSV Export:{csv_file}", tag="ok")
            except Exception as e:
                self.log(f"CSV Export failed:{e}", tag="warn")
            
            # PDF Export (si disponible)
            if PDF_AVAILABLE:
                pdf_file = f"scan_report_{timestamp}.pdf"
                try:
                    c = canvas.Canvas(pdf_file, pagesize=letter)
                    c.drawString(50, 750, "NETWORK AUDIT REPORT")
                    c.drawString(50, 730, f"Target: {self.get_target()}")
                    c.drawString(50, 710, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    c.drawString(50, 690, f"Total Issues: {len(self.problems_found)}")
                    
                    y = 670
                    for issue in self.problems_found[:20]:  # Limiter à 20 pour PDF
                        text = f"[{issue.get('type', '')}] {issue.get('host', '')}:{issue.get('port', '')}"
                        if y < 50:
                            c.showPage()
                            y = 750
                        c.drawString(50, y, text)
                        y -= 15
                    
                    c.save()
                    self.log(f"PDF Export:{pdf_file}", tag="ok")
                except Exception as e:
                    self.log(f"PDF Export failed:{e}", tag="warn")
            
            # Excel Export (si disponible)
            if EXCEL_AVAILABLE:
                xlsx_file = f"scan_report_{timestamp}.xlsx"
                try:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Scan Report"
                    ws['A1'] = 'Type'
                    ws['B1'] = 'Host'
                    ws['C1'] = 'Port'
                    ws['D1'] = 'Service'
                    ws['E1'] = 'Details'
                    ws['F1'] = 'Action'
                    
                    for idx, issue in enumerate(self.problems_found, 2):
                        ws[f'A{idx}'] = issue.get('type', '')
                        ws[f'B{idx}'] = issue.get('host', '')
                        ws[f'C{idx}'] = issue.get('port', '')
                        ws[f'D{idx}'] = issue.get('service', '')
                        ws[f'E{idx}'] = str(issue.get('details', ''))[:50]
                        ws[f'F{idx}'] = issue.get('action', '')
                    
                    wb.save(xlsx_file)
                    self.log(f"Excel Export:{xlsx_file}", tag="ok")
                except Exception as e:
                    self.log(f"Excel Export failed:{e}", tag="warn")
            
            self.log("\n[REPORTS GENERATED]", tag="title")
            self.log(f"Location: {os.getcwd()}")
            self.log(f"Total Issues Exported: {len(self.problems_found)}")
            self.display_problem_summary(1, self.get_target())
        except Exception as exc:
            self.log(f"EXPORT ERROR: {exc}", tag="warn")
        finally:
            self.root.after(0, self.stop_loading)

    # =========================================================================
    # [PARTIE 15] ACTIVE DIRECTORY RECONNAISSANCE
    # Cartographie le pouvoir dans un réseau d'entreprise
    # =========================================================================
    def audit_active_directory(self):
        """Scan basique de l'Active Directory (Utilisateurs VIP, DC)."""
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.log("[AD] DÉMARRAGE DE LA RECONNAISSANCE DU DOMAINE...", tag="title")
        
        if platform.system() != "Windows":
            self.log("[-] AD Recon nécessite Windows.", tag="error")
            self.root.after(0, self.stop_loading)
            return

        threading.Thread(target=self._run_ad_audit).start()

    def _run_ad_audit(self):
        try:
            # 1. Identifier le Domaine et le Contrôleur (DC)
            self.log("\n[1] INFO DOMAINE & DC", tag="accent")
            # nltest est présent sur tous les windows pros
            output = self.run_safe_command(["nltest", "/dsgetdc:"])
            
            if "Echec" in output or "Failed" in output:
                self.log("Cette machine n'est pas connectée à un domaine.", tag="warn")
                self.root.after(0, self.stop_loading)
                return
            
            self.log(output.strip(), tag="info")

            # 2. Lister les Administrateurs du Domaine (Les Cibles Prioritaires)
            self.log("\n[2] CHASSE AUX DOMAIN ADMINS", tag="accent")
            # La commande net group est redoutable
            cmd_admins = 'net group "Domain Admins" /domain'
            output_admins = self.run_safe_command(cmd_admins, timeout=10)
            
            if "N'existe pas" not in output_admins and "Error" not in output_admins:
                # Parsing simple pour extraire les noms
                lines = output_admins.split('\n')
                admins = []
                for line in lines:
                    if not line.startswith('La demande') and not line.startswith('Groupe') and not line.startswith('Commentaire') and not line.startswith('Membres') and line.strip():
                        # Nettoyer les séparateurs
                        parts = line.strip().split()
                        admins.extend(parts)
                
                if admins:
                    self.log(f"{len(admins)} ADMINISTRATEURS DOMAINE DÉTECTÉS !", tag="error")
                    for admin in admins:
                        self.log(f"  CIBLE VIP :{admin}", tag="warn")
                        self.problems_found.append({
                            'type': 'HIGH VALUE TARGET',
                            'host': 'AD CONTROLLER',
                            'details': f"Compte Admin Domaine identifié : {admin}",
                            'action': 'Surveiller les connexions de ce compte'
                        })
            else:
                self.log("[-] Impossible de lister les admins (Manque de droits ?)", tag="info")

            # 3. Politique de mots de passe (Faiblesse classique)
            self.log("\n[3] POLITIQUE DE MOT DE PASSE DU DOMAINE", tag="accent")
            cmd_pass = "net accounts /domain"
            output_pass = self.run_safe_command(cmd_pass)
            
            # Vérifier la longueur min
            for line in output_pass.split('\n'):
                if "ueur minimale" in line or "Minimum password length" in line:
                    self.log(f"  {line.strip()}", tag="info")
                    try:
                        length = int(''.join(filter(str.isdigit, line)))
                        if length < 10:
                            self.log("  Politique faible (<10 caractères) !", tag="warn")
                    except: pass
                elif "verrouillage" in line or "Lockout" in line:
                    self.log(f"  {line.strip()}", tag="info")

        except Exception as e:
            self.log(f"Erreur AD: {e}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    # =========================================================================
    # [PARTIE 16] HTML REPORT GENERATOR
    # Crée un rapport interactif et moderne (Bootstrap-like sans dépendance)
    # =========================================================================
    def generate_html_report(self):
        """Genere un rapport HTML autonome."""
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.log("[HTML] GÉNÉRATION DU RAPPORT INTERACTIF...", tag="title")
        
        threading.Thread(target=self._generate_html).start()

    #: Echelle CVSS standard, du plus grave au moins grave. "BAS" reste
    #: accepte en entree comme synonyme de "FAIBLE" (constats plus anciens).
    SEVERITIES = ("CRITIQUE", "ÉLEVÉ", "MOYEN", "FAIBLE", "INFO")

    #: Failles dont la seule mention suffit a classer un constat en critique.
    KNOWN_CRITICAL_FLAWS = (
        "eternalblue", "ms17-010", "bluekeep", "zerologon",
        "printnightmare", "log4shell", "heartbleed",
    )

    #: Constats acceptes par l'operateur (fichier local, une empreinte par ligne).
    ACCEPTED_FILE = "constats_acceptes.json"

    #: Identifiants de vulnerabilite reconnus dans le detail d'un constat.
    _MOTIF_CVE = re.compile(r"CVE-\d{4}-\d{4,7}", re.IGNORECASE)

    @staticmethod
    def finding_key(problem):
        """Empreinte stable d'un constat : type + hote + port + CVE.

        Volontairement independante du texte de detail, qui peut varier d'un
        scan a l'autre (numero de version, compteur), sans quoi une exception
        cesserait de s'appliquer au scan suivant.

        Une exception pres : l'identifiant CVE, quand le detail en contient un.
        Sans lui, accepter UNE vulnerabilite sur un port les acceptait TOUTES
        sur ce port -- y compris celles publiees apres coup. L'operateur
        croyait accorder une exception ciblee et eteignait durablement toute
        une classe d'alertes.
        """
        cve = AuditIA_Ultimate._MOTIF_CVE.search(str(problem.get("details") or ""))
        return "|".join((
            str(problem.get("type") or "").strip().upper(),
            str(problem.get("host") or "").strip(),
            str(problem.get("port") or "").strip(),
            cve.group(0).upper() if cve else "",
        ))

    def load_accepted(self):
        """Empreintes acceptees, chargees depuis le disque (et mises en cache)."""
        if getattr(self, "_accepted", None) is not None:
            return self._accepted
        try:
            with open(self.ACCEPTED_FILE, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            self._accepted = {k: v for k, v in data.items()} if isinstance(data, dict) else {}
        except FileNotFoundError:
            self._accepted = {}
        except Exception as exc:
            self.log(f"[ACCEPTÉS] Fichier illisible ({exc}) : liste ignorée.", tag="warn")
            self._accepted = {}
        return self._accepted

    def is_accepted(self, problem):
        """Vrai si l'operateur a declare ce constat comme voulu."""
        return self.finding_key(problem) in self.load_accepted()

    def accept_finding(self, problem, reason=""):
        """Marque un constat comme accepte : il cesse de remonter en gravite."""
        accepted = self.load_accepted()
        key = self.finding_key(problem)
        accepted[key] = {
            "raison": reason or "accepté par l'opérateur",
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "type": problem.get("type"),
            "host": problem.get("host"),
            "port": problem.get("port"),
        }
        self._save_accepted()
        self.log(f"[ACCEPTÉ] {key} — ne sera plus signalé comme risque.", tag="ok")

    def unaccept_finding(self, problem):
        """Retire un constat de la liste des exceptions."""
        accepted = self.load_accepted()
        if accepted.pop(self.finding_key(problem), None) is not None:
            self._save_accepted()
            self.log(f"[ACCEPTÉ] Exception retirée pour "
                     f"{self.finding_key(problem)}.", tag="info")

    def _save_accepted(self):
        try:
            with open(self.ACCEPTED_FILE, "w", encoding="utf-8") as handle:
                json.dump(self._accepted, handle, indent=2, ensure_ascii=False)
        except Exception as exc:
            self.log(f"[ACCEPTÉS] Enregistrement impossible : {exc}", tag="error")

    def classify_severity(self, problem):
        """Severite d'un constat, du signal le plus precis au plus general.

        Le rapport HTML utilisait deux logiques differentes : le compteur
        comparait le type a une liste exacte, le tableau cherchait des
        sous-chaines. "FAILLE EXPLOITABLE" contient "EXPLOIT" et sortait donc
        en critique dans le tableau sans etre compte dans le resume -- d'ou des
        rapports affichant "Critiques: 0" au-dessus de lignes critiques.

        L'ordre compte : un niveau de risque explicite dans les details prime
        sur le libelle du type, sinon un port Telnet note "Risque: HIGH"
        remonterait en critique juste parce que son type contient "EXPLOIT".
        """
        # -1. Constat declare voulu par l'operateur : il reste visible dans
        #     les rapports mais ne pollue plus le decompte des risques.
        if self.is_accepted(problem):
            return "INFO"

        kind = (problem.get("type") or "").upper()
        details = (problem.get("details") or "").lower()

        # 0. Severite explicite portee par le constat (ex. corrélation CVE, qui
        #    connait le score CVSS exact). Elle prime sur toute heuristique.
        explicit = (problem.get("severity") or "").upper().strip()
        aliases = {"BAS": "FAIBLE", "LOW": "FAIBLE", "HIGH": "ÉLEVÉ",
                   "ELEVE": "ÉLEVÉ", "CRITICAL": "CRITIQUE", "MEDIUM": "MOYEN"}
        explicit = aliases.get(explicit, explicit)
        if explicit in self.SEVERITIES:
            return explicit

        # 1. Niveau de risque explicitement porte par les details.
        for marker, level in (("risque: critical", "CRITIQUE"),
                              ("risque: high", "ÉLEVÉ"),
                              ("risque: medium", "MOYEN"),
                              ("risque: low", "FAIBLE")):
            if marker in details:
                return level

        # 2. Faille nommement connue et sans ambiguite.
        if any(flaw in details for flaw in self.KNOWN_CRITICAL_FLAWS):
            return "CRITIQUE"

        # 3. Nature du constat.
        # "CONNEXIONS RÉPÉTÉES" n'est volontairement pas critique : c'est une
        # observation a verifier, pas une compromission constatee.
        if any(word in kind for word in ("CRITIQUE", "BACKDOOR", "MALVEILLANTE",
                                         "ROOTKIT", "HONEYPOT_HIT",
                                         "HIGH VALUE", "EXPLOIT DÉTECTÉ")):
            return "CRITIQUE"
        if any(word in kind for word in ("VULNERABILITE", "VULNÉRABILITÉ",
                                         "CVE", "FAILLE")):
            return "ÉLEVÉ"
        if any(word in kind for word in ("RISQUE", "EXPOSED", "WARNING", "WARN",
                                         "BRUTE", "PERSISTENCE", "ANOMALIE",
                                         "RÉPÉTÉES", "CONNEXION")):
            return "MOYEN"
        return "FAIBLE"

    def count_by_severity(self):
        """Repartition des constats par severite."""
        counts = {level: 0 for level in self.SEVERITIES}
        for problem in self.problems_found:
            counts[self.classify_severity(problem)] += 1
        return counts

    def _generate_html(self):
        try:
            filename = f"RAPPORT_AUDIT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            
            # Statistiques
            total = len(self.problems_found)
            counts = self.count_by_severity()
            high = counts["CRITIQUE"]
            
            # HTML Template (Minimaliste et Sombre)
            html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Rapport d'audit {APP_NAME}</title>
    <style>
        body {{ background-color: #050505; color: #FFFFFF; font-family: 'Segoe UI', sans-serif; padding: 20px; margin: 0; }}
        h1 {{ color: #FF0000; border-bottom: 2px solid #FF0000; padding-bottom: 10px; margin-bottom: 30px; }}
        h2 {{ color: #FFFFFF; margin-top: 30px; padding-top: 20px; border-top: 1px solid #333; }}
        .stat-box {{ background: #0a0a0a; padding: 15px; margin: 10px 0; border-left: 5px solid #FF0000; border-radius: 5px; }}
        .danger {{ border-left-color: #FF0000; background-color: #330000; }}
        .warning {{ border-left-color: #FF0000; background-color: #330000; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; background: #0f0f0f; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #333; }}
        th {{ background-color: #1a1a1a; color: #FF0000; font-weight: bold; }}
        tr:hover {{ background-color: #1a1a1a; }}
        .tag {{ padding: 5px 10px; border-radius: 4px; font-size: 0.85em; font-weight: bold; display: inline-block; }}
        .tag-crit {{ background: #FF0000; color: white; }}
        .tag-warn {{ background: #FF0000; color: white; }}
        .tag-info {{ background: #CCCCCC; color: black; }}
        .tag-ok {{ background: #FFFFFF; color: black; }}
        .footer {{ margin-top: 50px; padding-top: 20px; border-top: 1px solid #333; font-size: 0.8em; color: #666; text-align: center; }}
        .header {{ background: linear-gradient(135deg, #050505 0%, #0a0a0a 100%); padding: 30px; border-radius: 10px; margin-bottom: 30px; }}
        .status {{ font-size: 2em; font-weight: bold; margin: 10px 0; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>{APP_NAME} — Rapport d'audit</h1>
        <p style="margin: 5px 0;"><strong>Cible:</strong> {self.get_target()}</p>
        <p style="margin: 5px 0;"><strong>Date:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
    </div>

    <div class="stat-box danger">
        <h2 style="margin-top: 0; color: #ff0000;"> RÉSUMÉ DES MENACES</h2>
        <div class="status" style="color: #ff0000;">{total} MENACES DÉTECTÉES</div>
        <p style="font-size: 1.1em; margin: 5px 0;">
            Critiques : <strong style="color:#ff5252;">{counts['CRITIQUE']}</strong> &nbsp;·&nbsp;
            Élevées : <strong>{counts['ÉLEVÉ']}</strong> &nbsp;·&nbsp;
            Moyennes : <strong>{counts['MOYEN']}</strong> &nbsp;·&nbsp;
            Faibles : <strong>{counts['FAIBLE']}</strong> &nbsp;·&nbsp;
            Info : <strong>{counts['INFO']}</strong>
        </p>
    </div>

    <h2> DÉTAILS DES MENACES</h2>
    <table>
        <thead>
            <tr>
                <th>TYPE</th>
                <th>HÔTE/CIBLE</th>
                <th>PORT</th>
                <th>SERVICE</th>
                <th>DÉTAILS</th>
                <th>ACTION RECOMMANDÉE</th>
                <th>SÉVÉRITÉ</th>
            </tr>
        </thead>
        <tbody>
"""
            
            badges = {"CRITIQUE": ("tag-crit", "CRITIQUE"),
                      "ÉLEVÉ": ("tag-crit", "ÉLEVÉ"),
                      "MOYEN": ("tag-warn", "MOYEN"),
                      "FAIBLE": ("tag-info", "FAIBLE"),
                      "INFO": ("tag-info", "INFO"),
                      "BAS": ("tag-info", "FAIBLE")}

            # Les constats les plus graves d'abord : un rapport se lit du haut.
            ordered = sorted(self.problems_found,
                             key=lambda item: self.SEVERITIES.index(
                                 self.classify_severity(item)))

            for p in ordered:
                style, severity_color = badges[self.classify_severity(p)]
                
                html_content += f"""
            <tr>
                <td><span class="tag {style}">{p.get('type', 'UNKNOWN')}</span></td>
                <td style="font-family: monospace;">{p.get('host', 'N/A')}</td>
                <td style="font-family: monospace;">{p.get('port', '—')}</td>
                <td>{p.get('service', '—')}</td>
                <td>{p.get('details', 'Pas de détails')}</td>
                <td><em>{p.get('action', 'Investiguer')}</em></td>
                <td>{severity_color}</td>
            </tr>
"""
            
            html_content += f"""
        </tbody>
    </table>

    <div class="stat-box warning">
        <h2 style="color: #ffcc00; margin-top: 0;"> RECOMMANDATIONS</h2>
        <ul>
            <li><strong>Immédiat:</strong> Isoler tout système avec des menaces CRITIQUES</li>
            <li><strong>Court terme:</strong> Investiguer et corriger toutes les failles détectées</li>
            <li><strong>Moyen terme:</strong> Renforcer les politiques de sécurité</li>
            <li><strong>Long terme:</strong> Audit régulier de conformité et mise à jour</li>
        </ul>
    </div>

    <div class="footer">
        <p>Genere par{APP_NAME} {APP_VERSION}</p>
        <p>Ce rapport est confidentiel et destiné aux autorités de sécurité informatique autorisées.</p>
    </div>
</body>
</html>
"""
            
            with open(filename, "w", encoding="utf-8") as f:
                f.write(html_content)
            
            self.log(f"Rapport HTML généré :{filename}", tag="ok")
            self.log(f"Chemin:{os.path.abspath(filename)}", tag="info")
            
            # Ouverture du navigateur : jamais en mode console, sinon une tache
            # planifiee ouvrirait une fenetre sur une session non surveillee.
            try:
                if getattr(self, "headless", False):
                    raise RuntimeError("mode console : pas d'ouverture automatique")
                import webbrowser
                webbrowser.open(os.path.abspath(filename))
                self.log("Ouverture dans le navigateur...", tag="info")
            except:
                self.log("(Ouverture manuelle requise)", tag="info")
                
        except Exception as e:
            self.log(f"Erreur HTML:{e}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    def detect_brute_force(self):
        """Détecte les tentatives de brute force SSH/RDP"""
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        
        # ANIMATIONS!
        self.show_loading_spinner(duration=1.5)
        self.animate_log_fadein("ANALYZING BRUTE FORCE ATTEMPTS...", tag="title")
        
        thread = threading.Thread(target=self._brute_force_analysis)
        thread.start()

    def _brute_force_analysis(self):
        """Analyse RÉELLE et complète des tentatives d'intrusion"""
        try:
            self.log("[BRUTE FORCE DETECTION] ANALYSE COMPLÈTE DES LOGS SÉCURITÉ", tag="title")
            self.log("="*70, tag="info")
            
            issues_found = 0
            
            if platform.system() != "Windows":
                self.log("[-] Module adapté Windows. Analysant alternative...", tag="warn")
                # Sur Linux, on peut analyser /var/log/auth.log
                try:
                    result = subprocess.run(['tail', '-n', '50', '/var/log/auth.log'], 
                                          capture_output=True, text=True)
                    failed_logins = result.stdout.count('Failed password')
                    self.log(f"[+] Failed SSH attempts: {failed_logins} in last 50 lines", tag="info")
                except:
                    self.log("[-] Cannot access Linux auth logs", tag="warn")
                return
            
            # ========== JOURNAUX SÉCURITÉ WINDOWS ==========
            # Section simplifiée (l'ancienne version était corrompue)
            self.log("\n[1] ANALYSE JOURNAUX SÉCURITÉ WINDOWS", tag="accent")
            self.log("   Journal Security non parsé ici (placeholder)", tag="info")
            
            # ========== FIREWALL LOG ANALYSIS (si disponible) ==========
            self.log("\n[2] FIREWALL BLOCKED CONNECTIONS ANALYSIS", tag="accent")
            
            # Chercher les connections bloquées par le Firewall Windows
            ps_firewall = """
            $blocked = Get-EventLog -LogName Security -InstanceId 5157 -Newest 20 -ErrorAction SilentlyContinue | 
                       Where-Object {$_.Message -like '*DROP*' -or $_.Message -like '*Denied*'}
            if ($blocked) {
                Write-Host "FOUND: $($blocked.Count)"
                foreach ($evt in $blocked) { Write-Host "$($evt.TimeGenerated)|$($evt.Message)" }
            } else {
                Write-Host "NONE"
            }
            """
            
            try:
                result = subprocess.run(['powershell', '-Command', ps_firewall], 
                                      capture_output=True, text=True, timeout=5)
                
                if "NONE" not in result.stdout and "FOUND:" in result.stdout:
                    blocked_count = result.stdout.count("|")
                    self.log(f"{blocked_count} connexions bloquées par Firewall", tag="warn")
                    issues_found += blocked_count
            except:
                self.log("   ℹ Firewall logs not accessible", tag="info")
            
            # ========== PORT SCANNING DETECTION ==========
            self.log("\n[3] PORT SCANNING DETECTION", tag="accent")
            
            # Chercher les activités anormales de scan de ports
            ps_scan = """
            $scan = Get-NetTCPConnection -State SynSent -ErrorAction SilentlyContinue | 
                    Where-Object {$_.RemoteAddress -notmatch '^127\\.|^192\\.168\\.|^10\\.|^172\\.'} | 
                    Measure-Object | Select-Object Count
            if ($scan.Count -gt 5) {
                Write-Host "SCANNING_DETECTED: $($scan.Count)"
            } else {
                Write-Host "NORMAL"
            }
            """
            
            try:
                result = subprocess.run(['powershell', '-Command', ps_scan], 
                                      capture_output=True, text=True, timeout=5)
                
                if "SCANNING_DETECTED" in result.stdout:
                    scan_count = int(result.stdout.split(":")[-1].strip())
                    self.log(f"   POSSIBLE PORT SCAN DETECTED:{scan_count} SYN connections pending!", tag="error")
                    issues_found += 1
                    self.problems_found.append({
                        'type': 'NETWORK SCAN ACTIVITY',
                        'details': f"{scan_count} SYN-SENT connections to external IPs",
                        'action': 'Verify if legitimate network activity'
                    })
                else:
                    self.log(f"   No suspicious port scanning detected", tag="success")
            except:
                self.log("   ℹ Port scan detection skipped", tag="info")
            
            # ========== RÉSUMÉ ==========
            self.log("\n" + "="*70, tag="info")
            if issues_found == 0:
                self.log("[RÉSULTAT] Aucune attaque brute force détectée", tag="success")
            else:
                self.log(f"[RÉSULTAT]{issues_found} problème(s) de sécurité détecté(s)!", tag="error")
            
            self.display_problem_summary(len(self.problems_found), self.get_target())
            
        except Exception as exc:
            self.log(f"ERREUR ANALYSE: {exc}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    def geolocate_ips(self):
        """Géolocalise les IPs externes détectées"""
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log("GEOLOCATING EXTERNAL IPs...", speed=0.02, tag="title")
        
        thread = threading.Thread(target=self._geolocation_scan)
        thread.start()

    def _geolocation_scan(self):
        """Géolocalisation RÉELLE des connexions actives via API"""
        try:
            import requests
        except ImportError:
            self.log("Erreur: Module 'requests' manquant. (pip install requests)", tag="error")
            return

        try:
            self.log("[GÉOLOCALISATION] ANALYSE DES CONNEXIONS ACTIVES...", tag="title")
            self.matrix_processing(lines=3, duration=0.3)

            # 1. Récupérer les IPs connectées via netstat
            active_ips = set()
            if platform.system() == "Windows":
                # Commande pour voir les connexions établies
                output = self.run_safe_command(['netstat', '-an'])
                for line in output.split('\n'):
                    if 'ESTABLISHED' in line:
                        parts = line.split()
                        if len(parts) >= 3:
                            # Récupérer l'IP distante (avant le :)
                            ip_port = parts[2]
                            if ':' in ip_port:
                                ip = ip_port.split(':')[0]
                                # Ignorer IPs locales
                                if not ip.startswith('127.') and not ip.startswith('192.168.') and not ip.startswith('10.'):
                                    active_ips.add(ip)
            
            if not active_ips:
                self.log("Aucune connexion externe active trouvée.", tag="warn")
                return

            self.log(f"{'IP':20} {'PAYS':15} {'VILLE':15} {'F.A.I (ISP)':25}")
            self.log("-" * 80)

            # 2. Interroger l'API pour chaque IP (Max 5 pour pas spammer l'API gratuite)
            for ip in list(active_ips)[:10]: 
                try:
                    response = requests.get(f"http://ip-api.com/json/{ip}", timeout=3)
                    data = response.json()
                    
                    if data['status'] == 'success':
                        country = data.get('country', 'Inconnu')
                        city = data.get('city', 'Inconnu')
                        isp = data.get('isp', 'Inconnu')
                        
                        # Détection pays à risque
                        tag = "warn" if country in ["China", "Russia", "North Korea"] else "info"
                        
                        self.log(f"{ip:20} {country:15} {city:15} {isp:25}", tag=tag)
                    else:
                        self.log(f"{ip:20} [Privée/Introuvable]")
                        
                except Exception:
                    self.log(f"{ip:20} Erreur API")
                
                time.sleep(0.5) # Pause pour respecter l'API

        except Exception as exc:
            self.log(f"ERREUR GEO: {exc}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    def scan_ssl_tls(self):
        """Scanne les certificats SSL/TLS"""
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log("SCANNING SSL/TLS CERTIFICATES...", speed=0.02, tag="title")
        
        thread = threading.Thread(target=self._ssl_tls_scan)
        thread.start()

    def _ssl_tls_scan(self):
        """Analyse RÉELLE du certificat SSL de la cible"""
        import ssl
        import socket
        
        target = self.get_target()
        # Si c'est une plage IP (ex: /24), on prend juste la première ou on demande une IP unique
        if '/' in target:
            target = target.split('/')[0]
            self.log(f"Cible ajustée pour SSL: {target}", tag="info")

        try:
            self.log(f"[SSL] CONNEXION À {target}:443...", tag="title")
            
            context = ssl.create_default_context()
            context.check_hostname = False
            context.verify_mode = ssl.CERT_NONE # Pour accepter même les certs expirés et les analyser
            
            with socket.create_connection((target, 443), timeout=5) as sock:
                with context.wrap_socket(sock, server_hostname=target) as ssock:
                    cert = ssock.getpeercert()
                    cipher = ssock.cipher()
                    version = ssock.version()
                    
                    self.log("\n[RAPPORT DE CHIFFREMENT]", tag="ok")
                    self.log(f"Protocole: {version}", tag="accent")
                    self.log(f"Cipher:    {cipher[0]} ({cipher[2]} bits)", tag="accent")
                    
                    # Analyse basique du certificat (si disponible)
                    if cert:
                        self.log("\n[DÉTAILS CERTIFICAT]", tag="info")
                        # Note: getpeercert() retourne un dict vide si verify_mode=CERT_NONE parfois
                        # Pour une analyse complète, il faudrait des lib tierces, mais on a déjà les infos de connexion
                        self.log("Certificat présenté et accepté par le socket.", tag="info")
                    else:
                        self.log("Certificat présent mais détails bruts non parsés (Mode Raw).", tag="warn")

                    # Vérification faiblesse
                    if "RC4" in cipher[0] or "MD5" in cipher[0] or version in ["TLSv1", "SSLv3"]:
                        self.log("SÉCURITÉ FAIBLE DÉTECTÉE !", tag="warn")
                        self.problems_found.append({
                            'type': 'WEAK SSL',
                            'host': target,
                            'details': f"Protocole obsolète: {version}",
                            'action': 'Mettre à jour le serveur web'
                        })
                    else:
                        self.log("Connexion sécurisée (Standards modernes respectés)", tag="success")

        except ConnectionRefusedError:
            self.log(f"Le port 443 est fermé sur {target}.", tag="error")
        except Exception as e:
            self.log(f"Erreur SSL: {e}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    def virustotal_scan(self):
        """Intégration VirusTotal pour vérifier les IPs/URLs"""
        api_key = tk.simpledialog.askstring("VirusTotal API", "Entrez votre clé VirusTotal API:\n(Laissez vide pour démo)")
        
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log("CHECKING WITH VIRUSTOTAL...", speed=0.02, tag="title")
        
        thread = threading.Thread(target=self._virustotal_check, args=(api_key,))
        thread.start()

    def _virustotal_check(self, api_key):
        """Vérifie la réputation de la cible réelle via l'API VirusTotal v3."""
        try:
            self.log("[VIRUSTOTAL REPUTATION CHECK]", tag="title")

            if not api_key or not api_key.strip():
                self.log("Aucune clé API VirusTotal fournie.", tag="error")
                self.log("Obtenez une clé gratuite sur : https://www.virustotal.com/gui/my-apikey", tag="accent")
                self._virustotal_demo()
                return

            # On interroge la cible réellement saisie, pas des IP d'exemple.
            target = self.get_target().strip()
            if '/' in target:
                target = target.split('/')[0]
            if not target:
                self.log("[-] Aucune cible définie.", tag="error")
                return

            self.log(f"[*] Interrogation de VirusTotal pour {target}...", tag="info")

            url = f"https://www.virustotal.com/api/v3/ip_addresses/{target}"
            response = requests.get(url, headers={"x-apikey": api_key}, timeout=10)

            if response.status_code == 401:
                self.log("[-] Clé API VirusTotal invalide.", tag="error")
                return
            if response.status_code == 429:
                self.log("[-] Quota VirusTotal dépassé (4 requêtes/min en gratuit).", tag="error")
                return
            if response.status_code == 404:
                self.log(f"[-] {target} est inconnu de VirusTotal.", tag="info")
                return
            if response.status_code != 200:
                self.log(f"[-] Erreur API VirusTotal : {response.status_code}", tag="error")
                return

            stats = response.json().get('data', {}).get('attributes', {}).get('last_analysis_stats', {})
            malicious = stats.get('malicious', 0)
            suspicious = stats.get('suspicious', 0)
            harmless = stats.get('harmless', 0)
            undetected = stats.get('undetected', 0)

            self.log(f"\n{'IP':20} {'Malicious':12} {'Suspicious':12} {'Harmless':12} {'Undetected':12}", tag="title")
            self.log("-" * 71)
            verdict_tag = "error" if malicious else ("warn" if suspicious else "ok")
            self.log(f"{target:20} {malicious:<12} {suspicious:<12} {harmless:<12} {undetected:<12}", tag=verdict_tag)

            if malicious or suspicious:
                self.log(f"\n{target} signalé par {malicious} moteur(s) comme malveillant, "
                         f"{suspicious} comme suspect.", tag="error")
                self.problems_found.append({
                    'type': 'VIRUSTOTAL ALERT',
                    'host': target,
                    'port': 'N/A',
                    'service': 'Reputation',
                    'details': f"Détecté par {malicious} moteur(s) antivirus ({suspicious} suspect(s))",
                    'action': 'Bloquer au niveau du firewall'
                })
            else:
                self.log(f"\n {target} n'est signalé par aucun moteur antivirus.", tag="success")

        except Exception as exc:
            self.log(f"VIRUSTOTAL ERROR: {exc}", tag="warn")
        finally:
            self.root.after(0, self.stop_loading)

    def _virustotal_demo(self):
        """Explique le mode démo SANS fabriquer de résultats.

        L'ancienne version inventait des verdicts pour des IP d'exemple et les
        injectait dans problems_found -- ces fausses menaces se retrouvaient
        ensuite dans les rapports PDF/Excel exportés comme si elles étaient
        réelles. On n'affiche plus aucun résultat fictif.
        """
        self.log("\n[MODE DÉMONSTRATION]", tag="warn")
        self.log("Aucune vérification n'a été effectuée : une clé API est requise.", tag="info")
        self.log("Aucun résultat fictif n'est ajouté au rapport.", tag="info")

    def scan_credentials(self):
        """Cherche des mots de passe en clair dans les fichiers locaux de l'utilisateur.

        Ce n'est PAS du test de credentials contre une cible distante : le scan
        parcourt le dossier Documents de la machine locale.
        """
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log("RECHERCHE DE SECRETS DANS VOS FICHIERS LOCAUX...", speed=0.02, tag="title")
        
        thread = threading.Thread(target=self._credential_scan)
        thread.start()

    def _credential_scan(self):
        """Recherche RÉELLE de mots de passe en clair dans les fichiers"""
        import os
        
        self.log("[CREDENTIALS] RECHERCHE DE MOTS DE PASSE EN CLAIR...", tag="title")
        self.log("Scan du dossier 'Documents' pour fichiers .txt, .conf, .ini...", tag="info")
        
        target_dir = os.path.join(os.environ.get('USERPROFILE'), 'Documents')
        keywords = ['password', 'passwd', 'pwd', 'api_key', 'secret', 'token']
        found_count = 0
        
        try:
            if not os.path.exists(target_dir):
                self.log("Dossier Documents introuvable.", tag="warn")
                return

            for root, dirs, files in os.walk(target_dir):
                for file in files:
                    if file.endswith(('.txt', '.ini', '.conf', '.env', '.log')):
                        filepath = os.path.join(root, file)
                        try:
                            with open(filepath, 'r', errors='ignore') as f:
                                content = f.read()
                                for kw in keywords:
                                    if kw in content.lower():
                                        self.log(f"POTENTIEL LEAK: '{kw}' trouvé dans {file}", tag="warn")
                                        self.problems_found.append({
                                            'type': 'CREDENTIAL LEAK',
                                            'details': f"Mot clé '{kw}' trouvé dans {file}",
                                            'action': 'Supprimer ou chiffrer ce fichier'
                                        })
                                        found_count += 1
                                        break # Un seul match par fichier suffit
                        except Exception:
                            pass # Fichier verrouillé ou illisible
                            
            if found_count == 0:
                self.log("Aucun fichier suspect trouvé dans Documents.", tag="success")
            else:
                self.log(f"\n[FIN] {found_count} fichiers contenant des mots-clés sensibles.", tag="warn")

        except Exception as exc:
            self.log(f"ERREUR SCAN FICHIERS: {exc}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    def show_dashboard(self):
        """Affiche un dashboard temps réel avec graphiques"""
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.typewriter_log("LOADING LIVE DASHBOARD...", speed=0.02, tag="title")
        
        thread = threading.Thread(target=self._display_dashboard)
        thread.start()

    def _display_dashboard(self):
        """Affiche un Dashboard avec des VRAIES données système (via psutil)"""
        try:
            import psutil
        except ImportError:
            self.log("Erreur: 'psutil' manquant. Installez avec: pip install psutil", tag="error")
            return

        try:
            self.log("="*60)
            self.log("[SYSTEM DIAGNOSTIC - LIVE FEED]", tag="title")
            self.matrix_processing(lines=3, duration=0.3)

            # Récupération des vraies données
            cpu_usage = psutil.cpu_percent(interval=1)
            memory = psutil.virtual_memory()
            disk = psutil.disk_usage('/')
            
            # Info réseau (octets envoyés/reçus)
            net_io = psutil.net_io_counters()
            sent_mb = round(net_io.bytes_sent / (1024 * 1024), 2)
            recv_mb = round(net_io.bytes_recv / (1024 * 1024), 2)

            # Création des barres de progression visuelles
            def make_bar(percent, length=20):
                filled = int(length * percent / 100)
                bar = "█" * filled + "░" * (length - filled)
                return bar

            # Construction de l'affichage
            dashboard = f"""
╔════════════════════════════════════════════════════════════════════════════╗
║                 SURVEILLANCE SYSTÈME (TEMPS RÉEL)                          ║
╠════════════════════════════════════════════════════════════════════════════╣
║                                                                            ║
║  CPU CORE:    [{make_bar(cpu_usage)}] {cpu_usage}%                         ║
║  MEMORY RAM:  [{make_bar(memory.percent)}] {memory.percent}% ({round(memory.used/1024**3, 1)}GB / {round(memory.total/1024**3, 1)}GB)  ║
║  DISK (C:):   [{make_bar(disk.percent)}] {disk.percent}% Used              ║
║                                                                            ║
║  NETWORK TELEMETRY                                                         ║
║  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━        ║
║  UPLOAD:      {sent_mb} MB total                                           ║
║  DOWNLOAD:    {recv_mb} MB total                                           ║
║                                                                            ║
║  ACTIVE PROCESSES: {len(psutil.pids())} running tasks                      ║
║  SYSTEM UPTIME:    {int(time.time() - psutil.boot_time()) // 3600} Hours   ║
║                                                                            ║
╚════════════════════════════════════════════════════════════════════════════╝
            """
            
            self.log(dashboard, tag="info")
            
            # Alerte si le système est en surcharge
            if cpu_usage > 85:
                self.log("ALERTE: SURCHAUFFE CPU DÉTECTÉE", tag="warn")
            if memory.percent > 90:
                self.log("ALERTE: MÉMOIRE CRITIQUE", tag="warn")
            
            self.log("[OK] SYSTÈME OPÉRATIONNEL", tag="success")

        except Exception as exc:
            self.log(f"ERREUR DASHBOARD: {exc}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    # ========== NOUVELLES FONCTIONNALITÉS V4.0 ==========
    

    def scan_wifi(self):
        """Scan WiFi networks"""
        self.start_loading()
        threading.Thread(target=self._scan_wifi_thread, daemon=True).start()
    
    # =========================================================================
    # PARTIE 9: WIFI CREDENTIAL HARVESTER - EXTRACTION DES CLÉS ENREGISTRÉES
    # =========================================================================
    def _scan_wifi_thread(self):
        """Scan WiFi RÉEL + Extraction des clés enregistrées (Windows)."""
        try:
            self.log("[WIFI] DÉMARRAGE DU SCAN ET DE L'EXTRACTION...", tag="title")
            
            if platform.system() != "Windows":
                self.log("[-] Fonctionnalité exclusive à Windows.", tag="error")
                return

            # 1. Scanner les réseaux environnants (Visibilité)
            self.log("\n[1] RÉSEAUX À PORTÉE (AIR SCAN)", tag="accent")
            cmd_scan = "netsh wlan show networks mode=bssid"
            output = self.run_safe_command(cmd_scan.split())
            
            if "pas en cours" in output:
                self.log("Erreur: Interface WiFi éteinte.", tag="error")
                return

            ssids_seen = set()
            for line in output.split('\n'):
                if "SSID" in line and ":" in line:
                    ssid = line.split(":")[1].strip()
                    if ssid and ssid not in ssids_seen:
                        self.log(f"  VISIBLE:{ssid}", tag="info")
                        ssids_seen.add(ssid)

            # 2. Extraire les profils enregistrés (Extraction de clés)
            self.log("\n[2] EXTRACTION DES CLÉS ENREGISTRÉES (DUMP)", tag="accent")
            
            # Récupérer la liste des profils
            cmd_profiles = "netsh wlan show profiles"
            output_profiles = self.run_safe_command(cmd_profiles.split())
            
            profiles = []
            for line in output_profiles.split('\n'):
                if "Profil Tous les utilisateurs" in line or "All User Profile" in line:
                    parts = line.split(":")
                    if len(parts) > 1:
                        profiles.append(parts[1].strip())
            
            self.log(f"[*] {len(profiles)} profils trouvés. Tentative de décryptage...", tag="warn")
            
            extracted_count = 0
            for profile in profiles:
                # SÉCURITÉ: Utiliser liste d'arguments au lieu de shell=True
                cmd_key = ["netsh", "wlan", "show", "profile", f"name={profile}", "key=clear"]
                
                try:
                    res = subprocess.run(cmd_key, capture_output=True, text=True)
                    content = res.stdout
                    
                    # Chercher la clé
                    key = "NON TROUVÉ"
                    for line in content.split('\n'):
                        if "Contenu de la" in line or "Key Content" in line:
                            if ":" in line:
                                key = line.split(":")[1].strip()
                                break
                    
                    if key != "NON TROUVÉ":
                        self.log(f"  SSID:{profile:<20} | PASS: {key}", tag="ok")
                        self.problems_found.append({
                            'type': 'WIFI CREDENTIAL',
                            'host': 'LOCAL',
                            'service': profile,
                            'details': f"Clé WPA récupérée: {key}",
                            'action': 'Rotation de mot de passe requise'
                        })
                        extracted_count += 1
                    else:
                        self.log(f"  SSID:{profile:<20} | PASS: [Chiffré/Enterprise]", tag="info")
                        
                except Exception:
                    pass

            if extracted_count > 0:
                self.log(f"\n[SUCCÈS] {extracted_count} clés Wifi extraites du système.", tag="success")
            else:
                self.log("\n[INFO] Aucune clé en clair récupérée (Enterprise ou pas de droits).", tag="info")

        except Exception as exc:
            self.log(f"ERREUR WIFI: {exc}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    # =========================================================================
    # PARTIE 10: PANIC BUTTON - EMERGENCY WIPE (OPSEC)
    # =========================================================================
    def emergency_wipe(self):
        """Procédure d'urgence : Suppression de toutes les données et arrêt."""
        if not messagebox.askyesno("CONFIRMATION", " ATTENTION \n\nCela va supprimer :\n- Tous les logs\n- La base de données\n- Les rapports générés\n- L'historique\n\nContinuer ?"):
            return

        self.log("INITIATING SELF-DESTRUCT SEQUENCE...", tag="error")
        
        files_to_destroy = [
            self.logs_file,
            "t800_neural_net.db",
            "network_audit_logs.txt"
        ]
        
        # Ajouter tous les JSON/PDF/CSV du dossier courant
        import glob
        for ext in ["*.json", "*.pdf", "*.csv", "*.xlsx", "*.png"]:
            files_to_destroy.extend(glob.glob(ext))

        deleted_count = 0
        
        # Fermer la DB avant suppression
        try:
            if hasattr(self, 'db_conn'):
                self.db_conn.close()
        except: pass

        for f in files_to_destroy:
            try:
                if os.path.exists(f):
                    os.remove(f)
                    print(f"Deleted: {f}") # Print console car GUI va fermer
                    deleted_count += 1
            except Exception as e:
                print(f"Error deleting {f}: {e}")

        messagebox.showinfo("WIPED", f"Nettoyage terminé.\n{deleted_count} fichiers détruits.\nArrêt du système.")
        self.root.destroy()
        import sys
        sys.exit()

    # =========================================================================
    # PARTIE 12: WEB HUNTER - RECONNAISSANCE WEB & ANALYSE HEADERS
    # =========================================================================
    def scan_web_technologies(self):
        """Scan des vulnérabilités Web basiques et Headers de sécurité."""
        target = self.get_target()
        # Nettoyage IP/URL
        if "/" in target: target = target.split("/")[0]
        
        self.disable_buttons()
        self.text_area.delete(1.0, tk.END)
        self.log(f"[WEB HUNTER] ANALYSE DES SERVICES WEB SUR {target}...", tag="title")
        
        threading.Thread(target=self._run_web_recon, args=(target,), daemon=True).start()

    def _run_web_recon(self, target):
        # Liste des chemins sensibles à tester
        SENSITIVE_PATHS = [
            'admin', 'login', 'dashboard', 'backup', 'config', 'db', 
            '.env', '.git/HEAD', 'phpinfo.php', 'wp-login.php'
        ]
        
        protocols = ['http', 'https']
        found_vulns = 0
        
        try:
            for proto in protocols:
                url = f"{proto}://{target}"
                self.log(f"\n[SCAN] Test de connexion {url}...", tag="info")
                
                try:
                    # 1. Analyse des Headers de Sécurité
                    resp = requests.get(url, timeout=3, verify=False)
                    headers = resp.headers
                    
                    self.log(f"  Serveur : {headers.get('Server', 'Masqué')}", tag="accent")
                    self.log(f"  Techno  : {headers.get('X-Powered-By', 'Inconnue')}", tag="accent")
                    
                    # Vérification des headers manquants
                    security_headers = [
                        ('Strict-Transport-Security', 'HSTS manquant (Risque MITM)'),
                        ('X-Frame-Options', 'Clickjacking possible'),
                        ('X-Content-Type-Options', 'MIME Sniffing possible'),
                        ('Content-Security-Policy', 'XSS protection faible (Pas de CSP)')
                    ]
                    
                    for header, alert in security_headers:
                        if header not in headers:
                            self.log(f"{alert}", tag="warn")
                            found_vulns += 1
                            self.problems_found.append({
                                'type': 'WEAK WEB CONFIG',
                                'host': target,
                                'service': f"{proto.upper()}",
                                'details': f"Header manquant: {header}",
                                'action': 'Configurer le serveur web'
                            })
                    
                    self.log(f"  Code retour racine :{resp.status_code}", tag="ok")

                    # 2. Recherche de dossiers cachés (Dirbusting léger)
                    self.log(f"  [DIR] Recherche de fichiers sensibles ({len(SENSITIVE_PATHS)} tests)...", tag="info")
                    
                    for path in SENSITIVE_PATHS:
                        full_url = f"{url}/{path}"
                        try:
                            sub_resp = requests.get(full_url, timeout=2, verify=False)
                            if sub_resp.status_code == 200:
                                self.log(f"  ALERTE :{full_url} est ACCESSIBLE (200 OK) !", tag="error")
                                found_vulns += 1
                                self.problems_found.append({
                                    'type': 'EXPOSED WEB PATH',
                                    'host': target,
                                    'details': f"Chemin sensible exposé : /{path}",
                                    'action': "Restreindre l'accès ou supprimer"
                                })
                            elif sub_resp.status_code == 403:
                                self.log(f"{path} détecté mais protégé (403 Forbidden)", tag="ok")
                        except: pass

                except requests.exceptions.ConnectionError:
                    self.log(f"  [-] Pas de service web sur {proto}://{target}", tag="info")
                except Exception as e:
                    self.log(f"  [-] Erreur Web: {e}", tag="warn")

            self.log("="*60)
            if found_vulns == 0:
                self.log("[RÉSULTAT] Aucun problème web évident détecté.", tag="success")
            else:
                self.log(f"[RÉSULTAT]{found_vulns} faiblesses web identifiées.", tag="warn")

        finally:
            self.root.after(0, self.stop_loading)

    # =========================================================================
    # PARTIE 13: ARP SENTINEL - DÉTECTION DU SPOOFING ARP
    # =========================================================================
    def check_arp_integrity(self):
        """Analyse la table ARP locale pour détecter les empoisonnements."""
        self.log("[ARP SENTINEL] ANALYSE INTÉGRITÉ RÉSEAU LOCAL...", tag="title")
        
        try:
            # Récupération de la table ARP
            cmd = "arp -a"
            output = self.run_safe_command(cmd.split())
            
            mac_table = {}
            duplicates = []
            
            for line in output.split('\n'):
                parts = line.split()
                # Parsing basique (IP Mac Type)
                if len(parts) >= 3:
                    ip = parts[0]
                    mac = parts[1].replace('-', ':').upper()
                    
                    # On ignore les adresses multicast/broadcast
                    if "224.0.0" in ip or "255.255" in ip or "FF:FF" in mac:
                        continue
                        
                    # Vérification format MAC (simple)
                    if len(mac) == 17 and ":" in mac:
                        if mac in mac_table:
                            duplicates.append((mac, mac_table[mac], ip))
                        else:
                            mac_table[mac] = ip
                            self.log(f"  {ip:<15} -> {mac}", tag="matrix")

            self.log("-" * 60)
            
            if duplicates:
                self.log("ALERTE : ARP SPOOFING DÉTECTÉ !", tag="error")
                for mac, ip1, ip2 in duplicates:
                    self.log(f"  La MAC{mac} usurpe {ip1} ET {ip2}", tag="error")
                    self.problems_found.append({
                        'type': 'ARP POISONING',
                        'host': 'RÉSEAU LOCAL',
                        'details': f"MAC {mac} associée à multiples IPs ({ip1}, {ip2})",
                        'action': 'Isoler la machine physique correspondante'
                    })
                    self.send_notification("ALERTE MITM", f"ARP Spoofing détecté sur {mac}")
            else:
                self.log("Table ARP saine. Pas d'attaque Man-in-the-Middle détectée.", tag="success")

        except Exception as e:
            self.log(f"Erreur ARP: {e}", tag="error")

    # =========================================================================
    # PARTIE 14: MAC VENDOR LOOKUP - IDENTIFICATION DES CONSTRUCTEURS
    # =========================================================================
    def get_mac_vendor(self, mac_address):
        """Récupère le constructeur via API MacVendors (nécessite Internet)."""
        try:
            url = f"https://api.macvendors.com/{mac_address}"
            resp = requests.get(url, timeout=2)
            if resp.status_code == 200:
                return resp.text.strip()
        except:
            pass
        return "Inconnu"
    
    def tool_mac_lookup(self):
        """Outil rapide pour identifier une MAC."""
        mac = tk.simpledialog.askstring("MAC Lookup", "Entrez une adresse MAC :")
        if mac:
            vendor = self.get_mac_vendor(mac)
            messagebox.showinfo("Résultat", f"MAC: {mac}\nConstructeur: {vendor}")
            self.log(f"[LOOKUP] {mac} = {vendor}", tag="info")

    def scan_cve_real(self):
        """
        Remplaçant GRATUIT de Nessus/Qualys.
        Utilise Nmap + Vulners NSE pour trouver les CVE réelles sur la cible.
        """
        target = self.get_target()
        if not target:
            messagebox.showwarning("Erreur", "Aucune cible définie.")
            return

        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        
        self.log("="*60)
        self.log(f"[*] DÉMARRAGE AUDIT VULNÉRABILITÉS (Moteur Open Source)", tag="title")
        self.log("[*] Utilisation de Nmap + Base de données Vulners...", tag="info")
        self.log("[*] Ceci remplace les solutions payantes (Nessus/Qualys).", tag="accent")
        
        thread = threading.Thread(target=self._run_real_vuln_scan, args=(target,))
        thread.start()

    def _run_real_vuln_scan(self, target):
        # Vérification critique: Nmap disponible?
        if not self.nm:
            self.log("[!] ERREUR: Nmap n'est pas disponible", tag="error")
            self.log("[*] Installez Nmap depuis https://nmap.org/download.html", tag="info")
            self.root.after(0, self.stop_loading)
            return
            
        try:
            # -sV : Détecte les versions (indispensable pour trouver les failles)
            # --script vulners : Le script magique qui compare les versions aux CVE connues
            # -T4 : Rapide
            arguments = "-sV --script vulners -T4"
            
            self.log(f"[*] Scan en cours sur {target}...", tag="info")
            self.log("[!] Cette opération peut prendre 2 à 5 minutes.", tag="warn")
            
            # Lancement du vrai scan
            self.nm.scan(target, arguments=arguments)
            
            hosts = self.nm.all_hosts()
            total_cves = 0
            
            if not hosts:
                self.log("[-] Aucun hôte trouvé ou pare-feu bloquant.", tag="error")
                return

            for host in hosts:
                self.log(f"\n[+] RAPPORT POUR : {host}", tag="title")
                
                # Parcourir tous les protocoles (TCP/UDP)
                for proto in self.nm[host].all_protocols():
                    ports = self.nm[host][proto].keys()
                    for port in ports:
                        service = self.nm[host][proto][port]
                        service_name = service.get('name', 'unknown')
                        product = service.get('product', '')
                        version = service.get('version', '')
                        
                        # Vérifier s'il y a des résultats de scripts
                        if 'script' in service and 'vulners' in service['script']:
                            vuln_output = service['script']['vulners']
                            
                            self.log(f"  PORT{port}/{proto} ({product} {version}) : VULNÉRABLE", tag="error")
                            
                            # Parser le texte brut de vulners pour le rendre joli
                            for line in vuln_output.split('\n'):
                                if "CVE-" in line or "SSV:" in line:
                                    # Nettoyer la ligne pour l'affichage
                                    clean_line = line.strip().replace('\t', ' ')
                                    
                                    # Déterminer la gravité (basique)
                                    tag = "warn"
                                    if "9." in line or "10.0" in line: # Score CVSS haut
                                        tag = "error"
                                        clean_line = "CRITIQUE:" + clean_line
                                    
                                    self.log(f"     -> {clean_line}", tag=tag)
                                    
                                    # Ajouter à la liste des problèmes pour l'export PDF/JSON
                                    total_cves += 1
                                    self.problems_found.append({
                                        'type': 'CVE VULNERABILITY',
                                        'host': host,
                                        'port': port,
                                        'service': f"{product} {version}",
                                        'details': clean_line,
                                        'action': 'Mise à jour requise urgemment'
                                    })
                        else:
                            # Service sain ou pas d'info
                            if product:
                                self.log(f"  PORT{port}/{proto} ({product} {version}) : Aucun CVE connu trouvé", tag="success")

            self.log("="*60)
            if total_cves > 0:
                self.log(f"RÉSULTAT :{total_cves} vulnérabilités connues détectées.", tag="error")
                self.log("[conseil] Exportez le rapport pour avoir les liens des correctifs.", tag="info")
            else:
                self.log("RÉSULTAT : Aucune vulnérabilité connue détectée avec cette base.", tag="success")

        except Exception as e:
            self.log(f"ERREUR SCAN VULN: {e}", tag="error")
            self.log("Astuce : Vérifiez que les scripts nmap sont installés (nmap --script-updatedb)", tag="info")
        finally:
            self.root.after(0, self.stop_loading)

    def run_real_searchsploit(self):
        """
        Implémentation native de SearchSploit en Python.
        Télécharge la base Exploit-DB officielle et croise avec les résultats Nmap.
        """
        # Vérifier si on a des résultats Nmap
        if not self.nm or not self.nm.all_hosts():
            messagebox.showwarning("Requis", "Veuillez d'abord lancer un SCAN RAPIDE ou COMPLET pour identifier les versions des services.")
            return

        self.disable_buttons()
        self.text_area.delete(1.0, tk.END)
        self.log("="*60)
        self.log("[*] SEARCHSPLOIT (EXPLOIT-DB) - ANALYSE RÉELLE", tag="title")
        
        # Lancer dans un thread pour ne pas figer l'interface
        threading.Thread(target=self._process_searchsploit).start()

    def _process_searchsploit(self):
        try:
            # 1. Téléchargement de la base de données officielle (fichier CSV)
            self.log("[*] Téléchargement de la base de données Exploit-DB...", tag="info")
            url = "https://gitlab.com/exploit-database/exploitdb/-/raw/main/files_exploits.csv"
            
            try:
                response = requests.get(url, timeout=10)
                response.raise_for_status()
                # On lit le CSV en mémoire
                csv_content = response.content.decode('utf-8')
                reader = csv.DictReader(io.StringIO(csv_content))
                exploits_db = list(reader) # Liste de dictionnaires
                self.log(f"[+] Base de données chargée : {len(exploits_db)} exploits connus.", tag="success")
            except Exception as e:
                self.log(f"[-] Erreur téléchargement DB: {e}", tag="error")
                self.root.after(0, self.stop_loading)
                return

            self.log("[*] Comparaison avec les services détectés...", tag="info")
            found_exploits = 0

            # 2. Croiser les données
            for host in self.nm.all_hosts():
                for proto in self.nm[host].all_protocols():
                    lport = self.nm[host][proto].keys()
                    for port in lport:
                        service = self.nm[host][proto][port]
                        product = service.get('product', '')
                        version = service.get('version', '')
                        
                        if not product: continue

                        # Construction de la requête de recherche (ex: "Apache 2.4")
                        query = f"{product} {version}".strip().lower()
                        self.log(f"\n Recherche pour : {query} (Port {port})", tag="accent")
                        
                        # Recherche dans la base
                        hits = []
                        for row in exploits_db:
                            description = row['description'].lower()
                            # Si le produit et la version sont dans la description
                            if product.lower() in description and (version.lower() in description if version else True):
                                hits.append(row)

                        # Affichage des résultats
                        if hits:
                            found_exploits += len(hits)
                            self.log(f"{len(hits)} Exploits potentiels trouvés !", tag="warn")
                            for exploit in hits[:5]: # On affiche les 5 premiers max
                                self.log(f"    -> ID: {exploit['id']} | {exploit['description'][:60]}...", tag="warn")
                                self.log(f"       Link: https://www.exploit-db.com/exploits/{exploit['id']}", tag="info")
                                
                                # Ajout au rapport final
                                self.problems_found.append({
                                    'type': 'EXPLOIT PUBLIC',
                                    'host': host,
                                    'service': query,
                                    'details': exploit['description'],
                                    'action': f"Vérifier exploit {exploit['id']}"
                                })
                        else:
                            self.log("  Aucun exploit direct trouvé dans la base.", tag="ok")

            self.log("="*60)
            self.log(f"[FIN] Total exploits potentiels identifiés : {found_exploits}", tag="title")

        except Exception as e:
            self.log(f"Erreur critique SearchSploit: {e}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    # -------------------------------------------------------------------------
    # CONNECTEUR OPENVAS WINDOWS (VIA DOCKER/TCP)
    # -------------------------------------------------------------------------
    def run_openvas_windows(self):
        """
        Connecteur Spécial Windows pour OpenVAS.
        Se connecte au conteneur Docker local ou serveur distant.
        """
        if not GVM_AVAILABLE:
            messagebox.showerror("Manque Module", "Installez la librairie : pip install python-gvm")
            return

        # Fenêtre de configuration connexion
        dialog = tk.Toplevel(self.root)
        dialog.title("Connexion OpenVAS (Docker/Remote)")
        dialog.geometry("450x350")
        dialog.configure(bg=THEME["bg"])

        tk.Label(dialog, text="CONNEXION GVM/OPENVAS", 
                 font=("Arial", 12, "bold"), bg=THEME["bg"], fg=THEME["fg"]).pack(pady=10)

        entries = {}
        # Par défaut, Docker expose souvent sur localhost
        fields = [
            ("IP Serveur (ex: 127.0.0.1)", "127.0.0.1"), 
            ("Port GMP (ex: 9390)", "9390"), 
            ("Utilisateur", "admin"), 
            ("Mot de passe", "admin")
        ]
        
        for i, (lbl, default) in enumerate(fields):
            tk.Label(dialog, text=lbl, bg=THEME["bg"], fg=THEME["accent"]).pack(pady=2)
            e = tk.Entry(dialog, bg="#333", fg="white", width=30)
            e.insert(0, default)
            if "Mot de passe" in lbl: e.config(show="*")
            e.pack(pady=2)
            entries[lbl] = e

        def execute_connection():
            host = entries["IP Serveur (ex: 127.0.0.1)"].get()
            port = int(entries["Port GMP (ex: 9390)"].get())
            user = entries["Utilisateur"].get()
            pwd = entries["Mot de passe"].get()
            target_ip = self.get_target()
            
            dialog.destroy()
            self.disable_buttons()
            self.text_area.delete(1.0, tk.END)
            self.log("="*60)
            self.log(f"[*] TENTATIVE DE CONNEXION AU MOTEUR OPENVAS ({host}:{port})...", tag="title")
            
            # Lancement en thread pour ne pas bloquer l'interface
            threading.Thread(target=self._process_openvas_scan, args=(host, port, user, pwd, target_ip)).start()

        btn = tk.Button(dialog, text="LANCER L'AUDIT", command=execute_connection, 
                  bg=THEME["fg"], fg="black", font=("Arial", 10, "bold"))
        btn.pack(pady=20)

    def _process_openvas_scan(self, host, port, user, password, target_ip):
        try:
            # Sur Windows, on utilise TLS (Réseau) et non UnixSocket
            from gvm.connections import TLSConnection
            from gvm.protocols.gmp import Gmp
            from gvm.transforms import EtreeTransform

            # Connexion sécurisée au Docker/Serveur
            connection = TLSConnection(hostname=host, port=port)
            
            # Utilisation du protocole GMP (Greenbone Management Protocol)
            with Gmp(connection, transform=EtreeTransform()) as gmp:
                self.log(f"[+] Canal TLS établi avec {host}.", tag="success")
                
                # 1. Authentification
                gmp.authenticate(user, password)
                self.log("[+] Authentification Admin : OK", tag="success")

                # 2. Création de la cible (Target)
                self.log(f"[*] Configuration de la cible : {target_ip}", tag="info")
                # Création d'une cible temporaire
                response = gmp.create_target(
                    name=f"Audit_T800_{target_ip}_{int(time.time())}", 
                    hosts=[target_ip], 
                    port_list_id="33d0cd82-57c6-11e1-8ed1-406186ea4fc5" # Scan complet et rapide
                )
                target_id = response.get('id')
                self.log(f"[+] Cible enregistrée ID: {target_id}", tag="ok")
                
                # 3. Récupération de la config de scan "Full and Fast"
                # (UUID standard OpenVAS, ne change généralement pas)
                config_id = "daba56c8-73ec-11df-a475-002264764cea" 

                # 4. Création de la tâche (Task)
                self.log("[*] Création de la tâche d'audit...", tag="info")
                task_res = gmp.create_task(
                    name=f"Task_T800_{target_ip}", 
                    config_id=config_id, 
                    target_id=target_id, 
                    scanner_id="08b69003-5fc2-4037-a479-93b440211c73" # OpenVAS Default Scanner
                )
                task_id = task_res.get('id')

                # 5. Démarrage
                self.log(f"DÉMARRAGE DU MOTEUR DE SCAN (ID:{task_id})...", tag="warn")
                gmp.start_task(task_id)
                
                self.log("[+] COMMANDE ENVOYÉE AU MOTEUR !", tag="success")
                self.log(f"[*] Le serveur OpenVAS analyse {target_ip} en arrière-plan.", tag="info")
                self.log("[*] Vous pouvez suivre l'avancement via l'interface Web Greenbone (https://127.0.0.1:9392)", tag="info")
                self.log("[*] Revenez vérifier les rapports plus tard.", tag="accent")

        except ConnectionRefusedError:
            self.log(f"[-] ÉCHEC : Impossible de se connecter à {host}:{port}", tag="error")
            self.log("[-] Vérifiez que Docker tourne et que le port 9390 est ouvert.", tag="warn")
        except Exception as e:
            self.log(f"[-] ERREUR CRITIQUE GVM : {e}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    def run_openvas_bridge(self):
        """
        Mode BRIDGE - Connexion à OpenVAS déjà en cours d'exécution
        Avec paramètres prédéfinis pour Docker local.
        """
        # Valeurs par défaut pour Docker local
        host = "127.0.0.1"
        port = 9390
        user = "admin"
        password = "admin"
        target_ip = self.get_target()
        
        if not target_ip:
            messagebox.showwarning("Erreur", "Aucune IP cible définie.")
            return
        
        self.disable_buttons()
        self.text_area.delete(1.0, tk.END)
        self.log("="*60)
        self.log(f"[*] OPENVAS BRIDGE MODE - Connexion auto à {host}:{port}...", tag="title")
        
        # Lancement en thread
        threading.Thread(target=self._process_openvas_scan, args=(host, port, user, password, target_ip)).start()

    def schedule_scans(self):
        """Configure une planification REELLE de scans (APScheduler)."""
        if self.scan_scheduler is None:
            self._not_implemented(
                "PLANIFICATION DE SCANS",
                "Le service de planification n'a pas pu etre initialise.")
            return

        cfg = dict(self.scan_scheduler.config.get_schedule_config() or {})

        dialog = tk.Toplevel(self.root)
        dialog.title("Planification des scans")
        dialog.configure(bg=THEME["bg"])
        scale = max(1.0, dialog.winfo_fpixels("1i") / 96.0)
        dialog.geometry(f"{int(480 * scale)}x{int(430 * scale)}")
        dialog.resizable(True, True)

        tk.Label(dialog, text="PLANIFICATION AUTOMATIQUE", font=THEME["font_title"],
                 bg=THEME["bg"], fg=THEME["fg"]).pack(pady=(14, 4))
        tk.Label(dialog, text=self.scan_scheduler.describe(), font=THEME["font_label"],
                 bg=THEME["bg"], fg=THEME["btn_tool"], wraplength=int(420 * scale)
                 ).pack(pady=(0, 10))

        body = tk.Frame(dialog, bg=THEME["bg"])
        body.pack(fill="both", expand=True, padx=18)

        enabled_var = tk.BooleanVar(value=bool(cfg.get("enabled")))
        tk.Checkbutton(body, text="Activer les scans planifies", variable=enabled_var,
                       bg=THEME["bg"], fg=THEME["fg_text"], selectcolor=THEME["bg_input"],
                       activebackground=THEME["bg"], activeforeground=THEME["fg"],
                       font=THEME["font_label"]).pack(anchor="w", pady=6)

        row = tk.Frame(body, bg=THEME["bg"])
        row.pack(fill="x", pady=6)
        tk.Label(row, text="Heure (HH:MM) :", bg=THEME["bg"], fg=THEME["fg_text"],
                 font=THEME["font_label"]).pack(side=tk.LEFT)
        time_var = tk.StringVar(value=cfg.get("time", "02:00"))
        tk.Entry(row, textvariable=time_var, width=8, bg=THEME["bg_input"],
                 fg=THEME["accent"], insertbackground=THEME["fg"],
                 font=THEME["font_mono"]).pack(side=tk.LEFT, padx=8)

        row2 = tk.Frame(body, bg=THEME["bg"])
        row2.pack(fill="x", pady=6)
        tk.Label(row2, text="Type de scan :", bg=THEME["bg"], fg=THEME["fg_text"],
                 font=THEME["font_label"]).pack(side=tk.LEFT)
        type_var = tk.StringVar(value=cfg.get("scan_type", "fast"))
        ttk.Combobox(row2, textvariable=type_var, state="readonly", width=12,
                     values=["fast", "full", "vuln", "backdoor"],
                     font=THEME["font_label"]).pack(side=tk.LEFT, padx=8)

        tk.Label(body, text="Jours :", bg=THEME["bg"], fg=THEME["fg_text"],
                 font=THEME["font_label"]).pack(anchor="w", pady=(10, 2))
        days_frame = tk.Frame(body, bg=THEME["bg"])
        days_frame.pack(fill="x")
        labels = ("lun", "mar", "mer", "jeu", "ven", "sam", "dim")
        selected = set(cfg.get("days") or [])
        day_vars = {}
        for index, name in enumerate(ScanScheduler.DAYS):
            var = tk.BooleanVar(value=name in selected)
            day_vars[name] = var
            tk.Checkbutton(days_frame, text=labels[index], variable=var,
                           bg=THEME["bg"], fg=THEME["fg_text"],
                           selectcolor=THEME["bg_input"], activebackground=THEME["bg"],
                           activeforeground=THEME["fg"], font=THEME["font_label"]
                           ).pack(side=tk.LEFT)

        tk.Label(body, text="La cible utilisee sera celle saisie dans l'interface "
                            "au moment du scan.", bg=THEME["bg"], fg=THEME["btn_muted"],
                 font=THEME["font_label"], wraplength=int(420 * scale),
                 justify="left").pack(anchor="w", pady=(12, 0))

        def save():
            raw_time = time_var.get().strip()
            try:
                hour, minute = (int(part) for part in raw_time.split(":"))
                if not (0 <= hour < 24 and 0 <= minute < 60):
                    raise ValueError
            except Exception:
                messagebox.showerror("Heure invalide",
                                     "Indiquez une heure au format HH:MM (ex. 02:30).",
                                     parent=dialog)
                return

            days = [name for name, var in day_vars.items() if var.get()]
            if enabled_var.get() and not days:
                messagebox.showerror("Jours manquants",
                                     "Selectionnez au moins un jour.", parent=dialog)
                return

            manager = self.scan_scheduler.config
            manager.set("scheduling.enabled", enabled_var.get())
            manager.set("scheduling.time", f"{hour:02d}:{minute:02d}")
            manager.set("scheduling.days", days)
            manager.set("scheduling.scan_type", type_var.get())

            if self.scan_scheduler.apply_config():
                self.log(f"[PLANIFICATION] {self.scan_scheduler.describe()}", tag="success")
            elif enabled_var.get():
                self.log("[PLANIFICATION] Activation impossible "
                         "(APScheduler absent ?).", tag="error")
            else:
                self.log("[PLANIFICATION] Scans planifies desactives.", tag="info")
            dialog.destroy()

        buttons = tk.Frame(dialog, bg=THEME["bg"])
        buttons.pack(side=tk.BOTTOM, fill="x", padx=18, pady=14)
        self.make_button(buttons, "ENREGISTRER", save, width=14).pack(side=tk.RIGHT)
        self.make_button(buttons, "ANNULER", dialog.destroy, width=12,
                         color=THEME["btn_muted"]).pack(side=tk.RIGHT, padx=8)

    def _run_scheduled_scan(self):
        """Point d'entree appele par le scheduler a l'heure programmee."""
        scan_type = (self.scan_scheduler.config.get_schedule_config() or {}).get(
            "scan_type", "fast")
        self.log(f"\n[PLANIFICATION] Declenchement du scan programme ({scan_type}).",
                 tag="title")
        # start_scan touche des widgets : on repasse par le thread principal.
        self.root.after(0, lambda: self.start_scan(scan_type))

    def scan_multi_target(self):
        """Scanne plusieurs cibles a la suite (separees par des virgules).

        L'ancienne version annoncait "Will scan N targets simultaneously"
        puis ne lancait aucun scan.
        """
        raw = self.get_target()
        targets = [part.strip() for part in raw.split(",") if part.strip()]

        if not targets:
            messagebox.showwarning(
                "Aucune cible",
                "Indiquez une ou plusieurs cibles separees par des virgules.\n"
                "Exemple : 192.168.1.1, 192.168.1.10, scanme.nmap.org")
            return

        if len(targets) == 1:
            self.log("[i] Une seule cible fournie : lancement d'un scan simple.",
                     tag="info")
            self.start_scan("fast")
            return

        scan_type = tk.simpledialog.askstring(
            "Scan multi-cibles",
            f"{len(targets)} cibles detectees.\n"
            "Type de scan (fast / full / vuln / backdoor) :",
            initialvalue="fast")
        if not scan_type:
            return
        scan_type = scan_type.strip().lower()
        if scan_type not in ("fast", "full", "vuln", "backdoor"):
            messagebox.showerror("Type inconnu",
                                 f"Type de scan non reconnu : {scan_type}")
            return

        self.scan_cancel.clear()
        self.problems_found = []
        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.log(f"[MULTI-CIBLES] {len(targets)} cibles, scan « {scan_type} »",
                 tag="title")

        def run_all():
            try:
                for index, target in enumerate(targets, 1):
                    if self.scan_cancel.is_set():
                        self.log("[STOP] Analyse multi-cibles interrompue.", tag="warn")
                        break
                    self.log(f"\n[{index}/{len(targets)}] Cible : {target}",
                             tag="accent")
                    before = len(self.problems_found)
                    try:
                        # run_nmap gere deja ses propres erreurs et reessais.
                        self.run_nmap(target, scan_type)
                    except Exception as exc:
                        self.log(f"[-] Echec sur {target} : {exc}", tag="error")
                        continue
                    found = len(self.problems_found) - before
                    self.log(f"    -> {found} probleme(s) releve(s) sur {target}",
                             tag="warn" if found else "ok")

                self.log(f"\n[MULTI-CIBLES] Termine : {len(targets)} cibles, "
                         f"{len(self.problems_found)} probleme(s) au total.",
                         tag="success")
            finally:
                self.root.after(0, self.stop_loading)

        threading.Thread(target=run_all, daemon=True).start()

    # ============ NOUVELLES ANIMATIONS ============
    
    
    # ============ FIN ANIMATIONS ============

    # =========================================================================
    # PARTIE 6: REST API SERVER - VRAI SERVEUR HTTP (No Dependencies)
    # =========================================================================
    def start_api_server(self):
        """Démarre un VRAI serveur HTTP multithreadé, en local uniquement, avec authentification."""
        import secrets

        self.disable_buttons()
        self.log("[API] Démarrage du serveur de contrôle...", tag="title")

        # Jeton généré à chaque démarrage : évite un serveur de statut accessible
        # sans authentification. Capturé côté thread principal, tout comme la
        # cible courante, pour ne jamais toucher un widget Tk depuis le thread HTTP.
        api_token = secrets.token_hex(16)
        snapshot_target = self.get_target()
        app_ref = self

        class T800RequestHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                if self.headers.get('X-API-Key') != api_token:
                    self.send_response(401)
                    self.send_header('Content-type', 'application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "unauthorized"}).encode('utf-8'))
                    return

                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.end_headers()

                # Récupérer les stats réelles (target figé au démarrage du serveur,
                # les listes sont lues telles quelles -- un léger risque de valeur
                # transitoire pendant qu'un scan écrit dedans, acceptable pour un
                # simple statut de lecture)
                status = {
                    "system": APP_NAME,
                    "status": "ONLINE",
                    "target": snapshot_target,
                    "threats_detected": len(app_ref.problems_found),
                    "scan_history_count": len(app_ref.scan_history)
                }

                # Renvoyer du JSON
                self.wfile.write(json.dumps(status).encode('utf-8'))

            def log_message(self, format, *args):
                return # Silence les logs console du serveur HTTP

        def run_server():
            try:
                server_address = ('127.0.0.1', 5000)
                httpd = HTTPServer(server_address, T800RequestHandler)
                app_ref.log(f"[+] API REST active sur http://127.0.0.1:5000 (accès local uniquement)", tag="success")
                app_ref.log(f"[+] Clé API requise (header X-API-Key) : {api_token}", tag="warn")
                app_ref.log("[+] Endpoint disponible: GET /status", tag="info")
                httpd.serve_forever()
            except Exception as e:
                app_ref.log(f"[-] Erreur démarrage API: {e}", tag="error")

        # Lancer le serveur dans un thread pour ne pas bloquer l'interface
        api_thread = threading.Thread(target=run_server, daemon=True)
        api_thread.start()
        
        self.root.after(0, self.stop_loading)

    # =========================================================================
    # PARTIE 7: ACTIVE DEFENSE SYSTEM - BLOCAGE RÉEL VIA WINDOWS FIREWALL
    # =========================================================================
    def active_defense_block_ip(self, ip_address):
        """Bloque une IP malveillante via le Pare-feu Windows (Admin requis)."""
        if platform.system() != "Windows":
            self.log("[-] Le blocage actif nécessite Windows.", tag="error")
            return

        rule_name = f"T800_BLOCK_{ip_address}"
        self.log(f"[DÉFENSE] Tentative de blocage de {ip_address}...", tag="warn")
        
        # Commande netsh pour bloquer tout le trafic entrant de cette IP
        cmd = [
            "netsh", "advfirewall", "firewall", "add", "rule",
            f"name={rule_name}",
            "dir=in",
            "action=block",
            f"remoteip={ip_address}",
            "enable=yes"
        ]
        
        # Exécution silencieuse
        output = self.run_safe_command(cmd)
        
        if "OK" in output or "Reussite" in output or "success" in output.lower():
            self.log(f"CIBLE NEUTRALISÉE :{ip_address} bloquée au niveau Kernel.", tag="success")
            self.send_notification("Défense Active", f"IP {ip_address} bloquée par le Firewall.")
        else:
            self.log(f"Échec du blocage (Êtes-vous Admin ?) :{output}", tag="error")

    def generate_firewall_rules(self):
        """Bouton Firewall : Applique les règles basées sur les menaces trouvées."""
        self.log("[*] Application des règles de défense...", tag="title")
        
        ips_to_block = set()
        
        # Uniquement les constats ou l'hostilite est etablie (tentative
        # d'intrusion observee, IP en liste noire publique).
        #
        # "CONNEXIONS RÉPÉTÉES" est volontairement EXCLU : c'est une
        # observation a verifier, et bloquer automatiquement la destination
        # couperait un service parfaitement legitime dans la majorite des cas.
        for issue in self.problems_found:
            if issue.get('type') in ['BRUTE FORCE LIVE', 'IP MALVEILLANTE',
                                     'HONEYPOT INTRUSION']:
                host = issue.get('host')
                if host and host != 'LOCAL' and host != '127.0.0.1':
                    ips_to_block.add(host)
        
        if not ips_to_block:
            self.log("Aucune menace critique nécessitant un blocage immédiat.", tag="ok")
            return

        for ip in ips_to_block:
            self.active_defense_block_ip(ip)

    def scan_broadcast(self):
        """Scan réseau complet par broadcast"""
        target = self.get_target()
        self.start_loading()
        threading.Thread(target=lambda: self._scan_broadcast_thread(target), daemon=True).start()
    
    def _scan_broadcast_thread(self, target):
        """Decouvre les hotes REELLEMENT joignables sur le reseau.

        Cette methode annoncait "Sending ARP requests..." puis listait
        192.168.1.1/51/101/151/201 comme ACTIVE -- sans emettre un paquet et
        sans tenir compte de la cible, meme sur un reseau 10.x. Elle croise
        desormais deux sources reelles : la table ARP du systeme et un
        balayage nmap -sn de la cible.
        """
        try:
            self.log("[DECOUVERTE RESEAU] Recherche des hotes joignables", tag="title")

            # --- 1. Voisins deja connus du systeme ----------------------
            self.log("\n[1] TABLE ARP LOCALE", tag="accent")
            voisins = self._discover_local_hosts()
            if voisins:
                for adresse in voisins:
                    self.log(f"   {adresse}", tag="ok")
                self.log(f"   -> {len(voisins)} voisin(s) dans la table ARP.", tag="info")
            else:
                self.log("   Table ARP vide ou illisible.", tag="warn")

            # --- 2. Balayage actif de la cible --------------------------
            self.log("\n[2] BALAYAGE ACTIF DE LA CIBLE", tag="accent")
            actifs = []
            if not target:
                self.log("   Aucune cible renseignee : balayage actif ignore.", tag="warn")
                self.log("   Renseignez une plage (ex. 192.168.1.0/24) pour l activer.", tag="info")
            elif self.nm is None:
                self.log("   Nmap indisponible : balayage actif impossible.", tag="warn")
                self.log("   Seule la table ARP ci-dessus a pu etre consultee.", tag="info")
            else:
                self.log(f"   nmap -sn {target}", tag="info")
                self.nm.scan(target, arguments="-sn -T4")
                for adresse in self.nm.all_hosts():
                    try:
                        if self.nm[adresse].state() != "up":
                            continue
                        nom = self.nm[adresse].hostname() or ""
                    except Exception:
                        continue
                    actifs.append(adresse)
                    suffixe = f" ({nom})" if nom else ""
                    self.log(f"   {adresse}{suffixe}", tag="ok")
                if actifs:
                    self.log(f"   -> {len(actifs)} hote(s) repondant sur {target}.", tag="info")
                else:
                    self.log(f"   Aucun hote n a repondu sur {target}.", tag="warn")

            # --- 3. Synthese --------------------------------------------
            ensemble = sorted(set(voisins) | set(actifs))
            self.log("\n" + "=" * 60, tag="info")
            if ensemble:
                self.log(f"[RESULTAT] {len(ensemble)} hote(s) distinct(s) decouvert(s).",
                         tag="success")
            else:
                self.log("[RESULTAT] Aucun hote decouvert.", tag="warn")

        except Exception as exc:
            self.log(f"ERREUR DECOUVERTE RESEAU : {exc}", tag="warn")
        finally:
            self.root.after(0, self.stop_loading)

    # =========================================================================
    # PARTIE 8: OSINT RÉEL (Whois & DNS) - RECONNAISSANCE PUBLIQUE RÉELLE
    # =========================================================================
    def osint_research(self):
        """Effectue une recherche OSINT réelle sur la cible."""
        target = self.get_target()
        # Nettoyage si format CIDR
        clean_target = target.split('/')[0]
        
        self.disable_buttons()
        self.text_area.delete(1.0, tk.END)
        self.log(f"[OSINT] RECHERCHE DE RENSEIGNEMENT SUR {clean_target}...", tag="title")
        
        def _run_osint():
            try:
                # 1. Résolution DNS étendue (NSlookup)
                self.log("\n[1] ENREGISTREMENTS DNS PUBLICS", tag="accent")
                cmd = ["nslookup", "-type=ANY", clean_target]
                output = self.run_safe_command(cmd)
                # Filtrer les lignes vides
                lines = [l for l in output.split('\n') if l.strip() and "server" not in l.lower()]
                for l in lines[:10]: # Max 10 lignes
                    self.log(f"  {l.strip()}", tag="info")

                # 2. Infos IP via API Publique (ipinfo.io - Gratuit sans clé pour usage basique)
                self.log("\n[2] IDENTIFICATION DU PROPRIÉTAIRE (ISP/ASN)", tag="accent")
                try:
                    resp = requests.get(f"https://ipinfo.io/{clean_target}/json", timeout=5)
                    data = resp.json()
                    
                    self.log(f"  Org/ISP :{data.get('org', 'Inconnu')}", tag="ok")
                    self.log(f"  Location :{data.get('city')}, {data.get('region')} ({data.get('country')})", tag="info")
                    self.log(f"  Hostname :{data.get('hostname', 'N/A')}", tag="info")
                    self.log(f"  Anycast :{data.get('anycast', False)}", tag="info")
                    
                except Exception as e:
                    self.log(f"  [-] API inaccessible : {e}", tag="warn")

                # 3. Traceroute rapide (Chemin réseau)
                self.log("\n[3] TRACEROUTE (Chemin réseau)", tag="accent")
                if platform.system() == "Windows":
                    # -h 5 pour limiter à 5 sauts (plus rapide pour l'exemple)
                    output = self.run_safe_command(["tracert", "-h", "5", "-d", clean_target])
                    for l in output.split('\n'):
                        if clean_target in l or "*" in l or "ms" in l:
                            self.log(f"  {l.strip()}", tag="matrix")

            except Exception as e:
                self.log(f"Erreur OSINT: {e}", tag="error")
            finally:
                self.root.after(0, self.stop_loading)

        threading.Thread(target=_run_osint, daemon=True).start()

    def configure_proxy(self):
        """Non implémenté : configuration proxy"""
        self._not_implemented(
            "CONFIGURATION PROXY",
            "Les scans n'utilisent actuellement aucun proxy ; le support reste à implémenter.")

    def show_logs_viewer(self):
        """Ouvre une vraie visionneuse du fichier de logs, avec rafraichissement."""
        path = os.path.abspath(self.logs_file)

        window = tk.Toplevel(self.root)
        window.title(f"Visionneuse de logs - {path}")
        window.configure(bg=THEME["bg"])
        scale = max(1.0, window.winfo_fpixels("1i") / 96.0)
        window.geometry(f"{int(900 * scale)}x{int(600 * scale)}")

        bar = tk.Frame(window, bg=THEME["bg"])
        bar.pack(fill="x", padx=8, pady=6)
        status = tk.Label(bar, text="", bg=THEME["bg"], fg=THEME["accent"],
                          font=THEME["font_label"])
        status.pack(side=tk.LEFT)

        area = scrolledtext.ScrolledText(window, bg="#050505", fg=THEME["accent"],
                                         font=THEME["font_mono"], wrap="none")
        area.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        def refresh():
            area.config(state="normal")
            area.delete(1.0, tk.END)
            try:
                with open(path, "r", encoding="utf-8", errors="replace") as handle:
                    lines = handle.readlines()
                # Les logs peuvent etre volumineux : on n'affiche que la fin.
                shown = lines[-2000:]
                area.insert(tk.END, "".join(shown))
                area.see(tk.END)
                status.config(text=f"{len(lines)} ligne(s) - {len(shown)} affichee(s)")
            except FileNotFoundError:
                area.insert(tk.END, f"Aucun fichier de log pour l'instant :\n{path}")
                status.config(text="fichier absent")
            except Exception as exc:
                area.insert(tk.END, f"Lecture impossible : {exc}")
                status.config(text="erreur de lecture")
            area.config(state="disabled")

        self.make_button(bar, " RAFRAICHIR", refresh, width=14).pack(side=tk.RIGHT, padx=4)
        refresh()

    def detect_anomalies_ai(self):
        """
        VRAIE IA : Utilise l'algorithme Isolation Forest (Scikit-learn)
        pour détecter les anomalies dans les connexions réseau actives.
        """
        try:
            from sklearn.ensemble import IsolationForest
            import numpy as np
            import psutil
        except ImportError:
            self.log("Erreur: Scikit-learn manquant. (pip install scikit-learn)", tag="error")
            return

        self.disable_buttons()
        self.progress.start(10)
        self.text_area.delete(1.0, tk.END)
        self.log("[AI] INITIALISATION DU MODÈLE NEURAL (Isolation Forest)...", tag="title")
        
        def _run_ai():
            try:
                # 1. Collecte de données réelles (Features)
                # On vectorise : [LocalPort, RemotePort, Status(enum), Type(TCP/UDP)]
                connections = psutil.net_connections()
                data = []
                details_map = [] # Pour retrouver qui est qui après l'analyse

                status_map = {
                    'ESTABLISHED': 1, 'SYN_SENT': 2, 'SYN_RECV': 3, 'FIN_WAIT1': 4,
                    'FIN_WAIT2': 5, 'TIME_WAIT': 6, 'CLOSE': 7, 'CLOSE_WAIT': 8,
                    'LAST_ACK': 9, 'LISTEN': 10, 'CLOSING': 11, 'NONE': 0
                }

                for c in connections:
                    if c.raddr: # On s'intéresse aux connexions distantes
                        l_port = c.laddr.port
                        r_port = c.raddr.port
                        stat_val = status_map.get(c.status, 0)
                        # Vecteur de features pour l'IA
                        data.append([l_port, r_port, stat_val])
                        details_map.append(c)

                if len(data) < 5:
                    self.log("[-] Pas assez de données pour entraîner le modèle.", tag="warn")
                    return

                # 2. Classement, et non detection.
                # L'ancien code fixait a l'avance la part d'anomalies attendue,
                # ce qui IMPOSE au modele d'en designer exactement autant : sur
                # une machine parfaitement saine, il en signalait quand meme
                # 5 %. C'etait un quota, pas une detection. Laisser le modele
                # trancher seul est pire encore (jusqu'a 60 % signalees).
                #
                # Le fond du probleme : un numero de port n'est pas une
                # grandeur ordonnee -- 443 et 445 ne sont pas "proches". Ce
                # modele ne peut donc pas trancher. Ce qu'il sait faire, c'est
                # CLASSER : dire quelles connexions ressemblent le moins aux
                # autres. On s'en tient a cela, et on le dit.
                clf = IsolationForest(contamination="auto", random_state=42)
                clf.fit(data)
                scores = clf.score_samples(data)
                self.log(f"   {len(data)} connexions distantes analysees.", tag="info")

                # Les plus atypiques d'abord (score le plus bas).
                classement = sorted(range(len(data)), key=lambda i: scores[i])
                a_montrer = classement[:5]

                self.log("", tag="info")
                self.log("   Les connexions les plus atypiques de CETTE machine :", tag="accent")
                for rang, i in enumerate(a_montrer, 1):
                    c = details_map[i]
                    self.log(f"   {rang}. score {scores[i]:+.3f}  "
                             f"{c.laddr.ip}:{c.laddr.port} <-> "
                             f"{c.raddr.ip}:{c.raddr.port} [{c.status}]", tag="info")

                self.log("", tag="info")
                self.log("   Ce classement est un ecart statistique, pas un verdict.", tag="warn")
                self.log("   Une machine saine a toujours une connexion « la plus", tag="warn")
                self.log("   atypique » : la premiere ligne n'est donc pas une alerte.", tag="warn")
                self.log("   Aucun constat n'est ajoute au rapport a partir de ce", tag="warn")
                self.log("   classement. A vous d'identifier le processus proprietaire", tag="warn")
                self.log("   si une ligne vous surprend.", tag="warn")

            except Exception as e:
                self.log(f"Erreur IA: {e}", tag="error")
            finally:
                self.root.after(0, self.stop_loading)

        threading.Thread(target=_run_ai, daemon=True).start()

    def scan_ssl_chain(self):
        """Scan chaîne SSL complète"""
        target = self.get_target()
        self.start_loading()
        threading.Thread(target=lambda: self._scan_ssl_chain_thread(target), daemon=True).start()
    
    @staticmethod
    def _decrire_certificat(der):
        """Resume un certificat DER : sujet, emetteur, dates de validite."""
        import ssl as _ssl
        import tempfile
        import os as _os

        pem = _ssl.DER_cert_to_PEM_cert(der)
        descripteur, chemin = tempfile.mkstemp(suffix=".pem")
        try:
            with _os.fdopen(descripteur, "w") as fichier:
                fichier.write(pem)
            infos = _ssl._ssl._test_decode_cert(chemin)
        finally:
            try:
                _os.unlink(chemin)
            except OSError:
                pass

        def aplatir(champ):
            valeurs = []
            for groupe in infos.get(champ, ()):
                for cle, valeur in groupe:
                    if cle in ("commonName", "organizationName"):
                        valeurs.append(valeur)
            return " / ".join(dict.fromkeys(valeurs)) or "(non renseigne)"

        return {
            "sujet": aplatir("subject"),
            "emetteur": aplatir("issuer"),
            "debut": infos.get("notBefore", "?"),
            "fin": infos.get("notAfter", "?"),
        }

    def _scan_ssl_chain_thread(self, target):
        """Recupere et verifie la VRAIE chaine de certificats de target:443.

        Cette methode affichait auparavant "Leaf/Intermediate/Root: Valid"
        puis "Chain Status: COMPLETE" sans ouvrir le moindre socket, quelle
        que soit la cible -- meme une IP sans serveur HTTPS. Elle etablit
        desormais une connexion TLS reelle et calcule chaque conclusion.
        """
        import ssl
        import socket as _socket
        from datetime import datetime as _dt

        hote = (target or "").split("/")[0].strip()
        port = 443
        if hote.count(":") == 1:
            gauche, _, droite = hote.partition(":")
            if droite.isdigit():
                hote, port = gauche, int(droite)

        protocole = None
        chiffrement = None

        try:
            self.log("[CHAINE SSL] Analyse de la chaine de certificats", tag="title")
            if not hote:
                self.log("Aucune cible renseignee.", tag="warn")
                return
            self.log(f"[*] Connexion TLS a {hote}:{port}...", tag="info")

            chaine = []
            verifiee = False
            motif = None

            # --- 1. Poignee de main VERIFIANTE ---------------------------
            try:
                with _socket.create_connection((hote, port), timeout=10) as brut:
                    contexte = ssl.create_default_context()
                    with contexte.wrap_socket(brut, server_hostname=hote) as tls:
                        verifiee = True
                        protocole = tls.version()
                        chiffrement = tls.cipher()
                        chaine = self._lire_chaine(tls)
            except ssl.SSLCertVerificationError as err:
                motif = getattr(err, "verify_message", None) or str(err)
            except (TimeoutError, _socket.timeout):
                self.log(f"[-] Delai depasse : {hote}:{port} ne repond pas en TLS.", tag="warn")
                return
            except (_socket.gaierror, ConnectionRefusedError, ssl.SSLError, OSError) as err:
                self.log(f"[-] Connexion TLS impossible sur {hote}:{port} : {err}", tag="warn")
                return

            # --- 2. Echec de verification : on decrit quand meme ---------
            if not verifiee:
                self.log("[!] Chaine REFUSEE par le magasin de confiance du systeme :", tag="error")
                self.log(f"    {motif}", tag="error")
                try:
                    with _socket.create_connection((hote, port), timeout=10) as brut:
                        permissif = ssl._create_unverified_context()
                        with permissif.wrap_socket(brut, server_hostname=hote) as tls:
                            protocole = tls.version()
                            chiffrement = tls.cipher()
                            chaine = self._lire_chaine(tls)
                except Exception as err:
                    self.log(f"[-] Certificat illisible : {err}", tag="warn")
                    return
                self.problems_found.append({
                    "type": "CHAINE SSL INVALIDE",
                    "host": f"{hote}:{port}",
                    "details": f"Verification refusee : {motif}",
                    "action": "Installer un certificat signe par une autorite reconnue",
                })
            else:
                self.log("[+] Chaine VALIDEE par le magasin de confiance du systeme.", tag="success")

            if protocole:
                self.log(f"    Protocole negocie    : {protocole}", tag="info")
            if chiffrement:
                self.log(f"    Suite de chiffrement : {chiffrement[0]}", tag="info")

            # --- 3. Description de chaque certificat --------------------
            if not chaine:
                self.log("[-] Le serveur n a presente aucun certificat lisible.", tag="warn")
                return

            nombre = len(chaine)
            if nombre == 1:
                self.log("\n[+] 1 certificat presente (cette version de Python "
                         "n expose pas la chaine complete) :", tag="ok")
            else:
                self.log(f"\n[+] Chaine presentee : {nombre} certificats", tag="ok")

            maintenant = _dt.now()
            for rang, der in enumerate(chaine):
                try:
                    infos = self._decrire_certificat(der)
                except Exception as err:
                    self.log(f"  [{rang}] certificat illisible : {err}", tag="warn")
                    continue

                if rang == 0:
                    role = "certificat du serveur"
                elif nombre > 1 and rang == nombre - 1:
                    role = "autorite racine presentee"
                else:
                    role = "autorite intermediaire"

                self.log(f"  [{rang}] {role}", tag="accent")
                self.log(f"      Sujet    : {infos['sujet']}", tag="info")
                self.log(f"      Emetteur : {infos['emetteur']}", tag="info")
                self.log(f"      Validite : {infos['debut']} -> {infos['fin']}", tag="info")

                try:
                    expiration = _dt.strptime(infos["fin"], "%b %d %H:%M:%S %Y %Z")
                except (ValueError, TypeError):
                    continue
                restant = (expiration - maintenant).days
                if restant < 0:
                    self.log(f"      EXPIRE depuis {-restant} jour(s)", tag="error")
                    if rang == 0:
                        self.problems_found.append({
                            "type": "CERTIFICAT EXPIRE",
                            "host": f"{hote}:{port}",
                            "details": f"Expire depuis {-restant} jour(s) ({infos['fin']})",
                            "action": "Renouveler le certificat",
                        })
                elif restant < 30:
                    self.log(f"      Expire dans {restant} jour(s)", tag="warn")
                    if rang == 0:
                        self.problems_found.append({
                            "type": "CERTIFICAT BIENTOT EXPIRE",
                            "host": f"{hote}:{port}",
                            "details": f"Expire dans {restant} jour(s) ({infos['fin']})",
                            "action": "Planifier le renouvellement",
                        })
                else:
                    self.log(f"      Encore valide {restant} jour(s)", tag="success")

        except Exception as exc:
            self.log(f"ERREUR CHAINE SSL : {exc}", tag="warn")
        finally:
            self.root.after(0, self.stop_loading)

    @staticmethod
    def _lire_chaine(tls):
        """Liste des certificats DER presentes par le pair.

        get_verified_chain / get_unverified_chain n existent qu a partir de
        Python 3.13 ; on retombe sinon sur le seul certificat terminal.
        """
        for methode in ("get_verified_chain", "get_unverified_chain"):
            recuperer = getattr(tls, methode, None)
            if recuperer is None:
                continue
            try:
                certificats = recuperer() or []
            except Exception:
                continue
            resultat = []
            for certificat in certificats:
                if isinstance(certificat, (bytes, bytearray)):
                    resultat.append(bytes(certificat))
                else:
                    lecture = getattr(certificat, "public_bytes", None)
                    if lecture is None:
                        continue
                    try:
                        import ssl as _ssl
                        resultat.append(lecture(_ssl.ENCODING_DER))
                    except Exception:
                        try:
                            resultat.append(lecture())
                        except Exception:
                            continue
            if resultat:
                return resultat
        terminal = tls.getpeercert(binary_form=True)
        return [terminal] if terminal else []

    def analyze_processes(self):
        """Analyse RÉELLE des processus en cours d'exécution"""
        try:
            import psutil
        except ImportError:
            self.log("Erreur: 'psutil' manquant.", tag="error")
            return

        self.log("[PROCESS] SCANNING ACTIVE TASKS...", tag="title")
        self.matrix_processing(lines=4, duration=0.4)
        
        suspicious_names = ['nc.exe', 'powershell.exe', 'cmd.exe', 'mimikatz.exe', 'wireshark.exe', 'keylogger']
        count = 0
        
        try:
            # En-tête
            self.log(f"{'PID':<8} {'NOM':<25} {'UTILISATEUR':<20} {'STATUT':<10}")
            self.log("-" * 70)

            for proc in psutil.process_iter(['pid', 'name', 'username', 'status']):
                try:
                    pinfo = proc.info
                    pid = pinfo['pid']
                    name = pinfo['name']
                    user = pinfo['username'] or "System"
                    status = pinfo['status']
                    
                    # Détection basique par nom
                    tag = "info"
                    prefix = "  "
                    
                    if name.lower() in suspicious_names:
                        tag = "warn"
                        prefix = ""
                        self.problems_found.append({
                            'type': 'PROCESSUS SUSPECT',
                            'details': f"{name} (PID: {pid}) en cours d'exécution",
                            'action': 'Terminer le processus si non autorisé'
                        })
                    
                    # Afficher seulement les processus utilisateur (pour ne pas spammer avec les processus système)
                    if count < 30 or tag == "warn": # Limite l'affichage aux 30 premiers sauf si alerte
                        self.log(f"{prefix}{str(pid):<8} {name:<25} {str(user)[:20]:<20} {status:<10}", tag=tag)
                        count += 1
                        
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    pass
            
            self.log(f"\n[FIN] {count} processus affichés.", tag="success")

        except Exception as e:
            self.log(f"Erreur Processus: {e}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)

    def setup_slack_notifications(self):
        """Ouvre la configuration des alertes (Slack, Discord, dérive)."""
        self.configure_alerts()

    def setup_webhooks(self):
        """Ouvre la configuration des alertes (Slack, Discord, dérive)."""
        self.configure_alerts()

    def configure_alerts(self):
        """Configure les canaux d'alerte et l'alerte sur dérive réseau.

        Remplace deux boutons qui se contentaient d'ecrire "Enter Slack Webhook
        URL" dans le journal sans rien enregistrer ni envoyer.
        """
        if self.config_manager is None:
            self._not_implemented(
                "ALERTES",
                "La configuration n'a pas pu être chargée.")
            return

        config = self.config_manager

        dialog = tk.Toplevel(self.root)
        dialog.title("Alertes et notifications")
        dialog.configure(bg=THEME["bg"])
        scale = max(1.0, dialog.winfo_fpixels("1i") / 96.0)
        dialog.geometry(f"{int(620 * scale)}x{int(470 * scale)}")
        dialog.resizable(True, True)

        tk.Label(dialog, text="ALERTES", font=THEME["font_title"],
                 bg=THEME["bg"], fg=THEME["fg"]).pack(pady=(14, 2))
        tk.Label(dialog,
                 text="Être prévenu quand le réseau change entre deux scans.",
                 font=THEME["font_label"], bg=THEME["bg"], fg=THEME["btn_muted"],
                 wraplength=int(560 * scale)).pack(pady=(0, 12))

        body = tk.Frame(dialog, bg=THEME["bg"])
        body.pack(fill="both", expand=True, padx=20)

        drift_var = tk.BooleanVar(
            value=bool((config.get("alerts", {}) or {}).get("on_drift")))
        tk.Checkbutton(body,
                       text="Alerter quand un hôte ou un port apparaît",
                       variable=drift_var, bg=THEME["bg"], fg=THEME["fg_text"],
                       selectcolor=THEME["bg_input"], activebackground=THEME["bg"],
                       activeforeground=THEME["fg"], font=THEME["font_label"]
                       ).pack(anchor="w", pady=(0, 12))

        entries = {}

        def channel_row(key, label, hint):
            frame = tk.LabelFrame(body, text=f" {label} ", font=THEME["font_label"],
                                  bg=THEME["bg"], fg=THEME["fg"], bd=1, relief="solid")
            frame.pack(fill="x", pady=6)

            section = config.get(key, {}) or {}
            enabled = tk.BooleanVar(value=bool(section.get("enabled")))
            tk.Checkbutton(frame, text=f"Activer {label}", variable=enabled,
                           bg=THEME["bg"], fg=THEME["fg_text"],
                           selectcolor=THEME["bg_input"], activebackground=THEME["bg"],
                           activeforeground=THEME["fg"], font=THEME["font_label"]
                           ).pack(anchor="w", padx=10, pady=(6, 2))

            tk.Label(frame, text=hint, font=THEME["font_label"], bg=THEME["bg"],
                     fg=THEME["btn_muted"], wraplength=int(520 * scale),
                     justify="left").pack(anchor="w", padx=10)

            url = tk.StringVar(
                value=_secrets.unprotect(section.get("webhook_url", "")))
            tk.Entry(frame, textvariable=url, bg=THEME["bg_input"],
                     fg=THEME["accent"], insertbackground=THEME["fg"],
                     font=THEME["font_mono"], bd=0).pack(
                fill="x", padx=10, pady=(4, 10), ipady=scaled(4))
            entries[key] = (enabled, url)

        channel_row("slack", "Slack",
                    "URL de webhook entrante (Slack → Applications → Incoming Webhooks)")
        channel_row("discord", "Discord",
                    "URL de webhook du salon (Paramètres du salon → Intégrations)")

        tk.Label(body, text="L'envoi par email se configure via le bouton "
                            "CONFIGURER EMAIL.", font=THEME["font_label"],
                 bg=THEME["bg"], fg=THEME["btn_muted"],
                 wraplength=int(560 * scale), justify="left").pack(anchor="w", pady=(8, 0))

        def save():
            config.set("alerts.on_drift", drift_var.get())
            for key, (enabled, url) in entries.items():
                config.set(f"{key}.enabled", enabled.get())
                # URL chiffree via DPAPI avant stockage.
                config.set(f"{key}.webhook_url",
                           _secrets.protect(url.get().strip()))
            active = [k for k, (e, _) in entries.items() if e.get()]
            self.log(f"[ALERTES] Dérive : {'activée' if drift_var.get() else 'désactivée'}"
                     f" — canaux : {', '.join(active) if active else 'aucun'}",
                     tag="success")
            dialog.destroy()

        def test():
            if self.webhook_notifier is None:
                messagebox.showwarning("Indisponible",
                                       "Les notificateurs ne sont pas initialisés.",
                                       parent=dialog)
                return
            save_needed = False
            for key, (enabled, url) in entries.items():
                config.set(f"{key}.enabled", enabled.get())
                config.set(f"{key}.webhook_url",
                           _secrets.protect(url.get().strip()))
                save_needed = True
            if not save_needed:
                return
            title = "Test d'alerte"
            body_text = "Ceci est un message de test envoyé depuis SIPA."
            results = []
            for key, sender in (("slack", self.webhook_notifier.send_slack),
                                ("discord", self.webhook_notifier.send_discord)):
                if not entries[key][0].get():
                    continue
                try:
                    results.append(f"{key} : {'envoyé' if sender(title, body_text) else 'échec'}")
                except Exception as exc:
                    results.append(f"{key} : erreur ({exc})")
            messagebox.showinfo("Test d'alerte",
                                chr(10).join(results) or "Aucun canal activé.",
                                parent=dialog)

        buttons = tk.Frame(dialog, bg=THEME["bg"])
        buttons.pack(side=tk.BOTTOM, fill="x", padx=20, pady=14)
        self.make_button(buttons, "ENREGISTRER", save, width=14).pack(side=tk.RIGHT)
        self.make_button(buttons, "TESTER", test, width=10,
                         color=THEME["btn_tool"]).pack(side=tk.RIGHT, padx=8)
        self.make_button(buttons, "ANNULER", dialog.destroy, width=10,
                         color=THEME["btn_muted"]).pack(side=tk.RIGHT)

    def toggle_dark_mode(self, dark: bool):
        """Bascule entre la palette sombre et la palette claire.

        Cette methode se contentait d'ecrire "Theme will be applied on next
        restart" : rien n'etait enregistre, aucune palette claire n'existait,
        et le redemarrage ne changeait donc rien. La bascule est desormais
        reelle -- la palette est appliquee, enregistree, et l'interface est
        reconstruite immediatement.
        """
        palette = "sombre" if dark else "clair"
        if theme_module.current_palette() == palette:
            self.log(f"[THEME] Palette « {palette} » deja active.", tag="info")
            return

        theme_module.apply_palette(palette)

        # Persistance : la preference doit survivre au redemarrage.
        enregistre = False
        config = getattr(self, "config_manager", None)
        if config is not None:
            try:
                config.set("interface.theme", palette)
                enregistre = True
            except Exception as exc:
                self.log(f"[THEME] Preference non enregistree : {exc}", tag="warn")

        try:
            self.rebuild_interface()
        except Exception as exc:
            self.log(f"[THEME] Reconstruction impossible : {exc}", tag="error")
            self.log("[THEME] Redemarrez SIPA pour appliquer la palette.", tag="warn")
            return

        self.log(f"[THEME] Palette « {palette} » appliquee.", tag="ok")
        if enregistre:
            self.log("[THEME] Preference enregistree dans config.json.", tag="info")

    def show_current_theme(self):
        """Affiche la palette active et la source de cette preference."""
        palette = theme_module.current_palette()
        self.log(f"[THEME] Palette active : « {palette} »", tag="ok")
        config = getattr(self, "config_manager", None)
        if config is not None:
            enregistree = config.get("interface.theme", None)
            if enregistree:
                self.log(f"[THEME] Preference enregistree : « {enregistree} »", tag="info")
            else:
                self.log("[THEME] Aucune preference enregistree (defaut : sombre).",
                         tag="info")

    def rebuild_interface(self):
        """Reconstruit l'interface avec la palette courante.

        Tkinter fige les couleurs au moment de la creation de chaque widget :
        changer de palette impose donc de reconstruire les panneaux. Le
        contenu du journal est preserve et reinjecte.
        """
        journal = ""
        try:
            journal = self.text_area.get("1.0", tk.END)
        except Exception:
            journal = ""

        for enfant in list(self.root.winfo_children()):
            try:
                enfant.destroy()
            except Exception:
                pass

        self.root.configure(bg=THEME["bg"])

        self.setup_styles()
        self.build_header()
        self.build_target()
        self.build_controls()
        self.build_logs()
        self.build_progress()
        self.setup_responsive_design()

        if journal.strip():
            try:
                self.text_area.insert("1.0", journal.rstrip() + "\n")
                self.text_area.see(tk.END)
            except Exception:
                pass

    def show_scan_history(self):
        """Affiche l'historique REEL des scans (base SQLite + session courante)."""
        self.log("\n[HISTORIQUE DES SCANS]", tag="title")

        rows = []
        try:
            with self.db_lock:
                self.cursor.execute(
                    "SELECT timestamp, target_ip, scan_type, vulns_found "
                    "FROM scan_history ORDER BY id DESC LIMIT 20")
                rows = self.cursor.fetchall()
        except Exception as exc:
            self.log(f"[-] Base d'historique illisible : {exc}", tag="error")

        if rows:
            self.log(f"{'DATE':21} {'CIBLE':22} {'TYPE':14} MENACES", tag="accent")
            self.log("-" * 66)
            for timestamp, target, scan_type, vulns in rows:
                self.log(f"{timestamp:21} {str(target):22} {str(scan_type):14} {vulns}",
                         tag="warn" if vulns else "ok")
            self.log(f"\n[+] {len(rows)} scan(s) enregistre(s) en base.", tag="info")
        else:
            self.log("[i] Aucun scan enregistre en base pour le moment.", tag="info")
            self.log("    L'historique se remplit a la fin de chaque scan.", tag="info")

        if self.scan_history:
            self.log(f"\n[SESSION COURANTE] {len(self.scan_history)} scan(s) :", tag="accent")
            for entry in self.scan_history[-10:]:
                self.log(f"  {entry.get('timestamp', '?')} - "
                         f"{entry.get('scan_type', '?')} - {entry.get('target', '?')}",
                         tag="info")

    # ========== FONCTIONNALITÉS COMPLÉMENTAIRES ==========

    def auto_save_scan(self, target: str, scan_type: str, results: dict) -> None:
        """Sauvegarde automatique des résultats de scan"""
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            scan_entry = {
                'timestamp': timestamp,
                'target': target,
                'scan_type': scan_type,
                'results': results
            }
            self.scan_history.append(scan_entry)
            self.scan_results_cache[f"{target}_{scan_type}"] = results

            # La cible vient du champ saisi par l'utilisateur : on la nettoie
            # avant de construire un nom de fichier (traversal / caractères
            # invalides sous Windows sinon).
            import re
            safe_target = re.sub(r'[^A-Za-z0-9._-]', '_', target) or "target"
            safe_scan_type = re.sub(r'[^A-Za-z0-9._-]', '_', scan_type) or "scan"

            with open(f"{safe_target}_scan_{safe_scan_type}.json", 'w') as f:
                json.dump(results, f, indent=2)

            self.log(f"[+] Auto-save: Scan results saved for {target}", tag="ok")
        except Exception as exc:
            self.log(f"[-] Auto-save error: {exc}", tag="warn")

    def send_desktop_notification(self, title: str, message: str, severity: str = "info") -> None:
        """Envoie une notification desktop"""
        try:
            if self.notification_enabled:
                self.notification_queue.append((title, message, severity))
                
                # Colorer selon la sévérité
                if severity == "critical":
                    self.log(f"CRITICAL:{title} - {message}", tag="warn")
                elif severity == "warning":
                    self.log(f"WARNING:{title} - {message}", tag="warn")
                else:
                    self.log(f"ℹ INFO:{title} - {message}", tag="info")
        except Exception as exc:
            self.log(f"Notification error: {exc}", tag="warn")

    def log_rotation(self, max_size_mb: int = 10) -> None:
        """Gère la rotation des logs"""
        try:
            log_path = self.logs_file
            if os.path.exists(log_path):
                size_mb = os.path.getsize(log_path) / (1024 * 1024)
                if size_mb > max_size_mb:
                    backup_name = f"{log_path}.{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    os.rename(log_path, backup_name)
                    self.log(f"[+] Log rotation: {log_path} -> {backup_name}", tag="ok")
        except Exception as exc:
            self.log(f"Log rotation error: {exc}", tag="warn")

    def write_logs_to_file(self, message: str) -> None:
        """Écrit les logs dans un fichier"""
        try:
            with open(self.logs_file, 'a', encoding='utf-8') as f:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                f.write(f"[{timestamp}] {message}\n")
        except Exception as exc:
            print(f"Error writing to log file: {exc}")

    def compare_scan_results(self, target: str, scan_type1: str, scan_type2: str) -> None:
        """Compare les résultats de deux scans"""
        self.log("[*] Comparing scan results...", tag="info")
        
        key1 = f"{target}_{scan_type1}"
        key2 = f"{target}_{scan_type2}"
        
        if key1 in self.scan_results_cache and key2 in self.scan_results_cache:
            result1 = self.scan_results_cache[key1]
            result2 = self.scan_results_cache[key2]
            self.log(f"[+] Comparison: {scan_type1} vs {scan_type2}", tag="ok")
            self.log(f"[+] Differences detected: {len(result1) - len(result2)} entries", tag="info")
        else:
            self.log("[-] Insufficient cached data for comparison", tag="warn")

    # ============= FONCTIONNALITÉS ENTREPRISE ==================================
    # ============ FONCTIONNALITÉS ENTREPRISE ==================================
    # =========================================================================
    
    def show_realtime_graphs(self):
        """Affiche les graphiques temps réel basés sur PSUTIL (Données réelles)"""
        # Vérifier si une fenêtre existe déjà
        if hasattr(self, 'graph_window') and self.graph_window and self.graph_window.winfo_exists():
            self.graph_window.lift()  # Ramener au premier plan
            self.log("[*] Fenêtre graphiques déjà ouverte", tag="info")
            return
        
        # Charger matplotlib dynamiquement
        global plt, Figure, FigureCanvasTkAgg, animation
        plt_temp, Figure_temp, FigureCanvasTkAgg_temp, animation_temp = _load_matplotlib()
        
        if not plt_temp:
            messagebox.showerror("Erreur", "Installez matplotlib: pip install matplotlib")
            return
        
        plt = plt_temp
        Figure = Figure_temp
        FigureCanvasTkAgg = FigureCanvasTkAgg_temp
        animation = animation_temp
        
        import psutil
        
        graph_window = tk.Toplevel(self.root)
        graph_window.title("Surveillance en temps réel")
        graph_window.geometry("1000x700")
        graph_window.configure(bg="black")

        fig = Figure(figsize=(10, 7), facecolor='black')
        ax_net = fig.add_subplot(211, facecolor='#111')
        ax_cpu = fig.add_subplot(212, facecolor='#111')

        # Configuration visuelle
        for ax, title in [(ax_net, "Débit Réseau (Ko/s)"), (ax_cpu, "Utilisation CPU (%)")]:
            ax.set_title(title, color="#FFFFFF")
            ax.grid(True, alpha=0.3)
            ax.tick_params(axis='x', colors="white")
            ax.tick_params(axis='y', colors="white")
            for spine in ax.spines.values(): spine.set_color("#FFFFFF")

        # Variables de stockage
        times = list(range(60))
        sent_data = [0] * 60
        recv_data = [0] * 60
        cpu_data = [0] * 60
        
        # Initialisation lines
        l_sent, = ax_net.plot(times, sent_data, label="Upload", color="red")
        l_recv, = ax_net.plot(times, recv_data, label="Download", color="#00FFFF")
        l_cpu, = ax_cpu.plot(times, cpu_data, label="CPU", color="orange")
        
        ax_net.legend()
        
        # Variable mémoire pour calculer la différence (débit)
        self.last_net_io = psutil.net_io_counters()

        def update(frame):
            # 1. Calculer le débit réel
            current_net_io = psutil.net_io_counters()
            # Différence en Octets / 1024 = Ko
            delta_sent = (current_net_io.bytes_sent - self.last_net_io.bytes_sent) / 1024
            delta_recv = (current_net_io.bytes_recv - self.last_net_io.bytes_recv) / 1024
            
            self.last_net_io = current_net_io
            
            # 2. Mise à jour des listes
            sent_data.append(delta_sent)
            recv_data.append(delta_recv)
            cpu_data.append(psutil.cpu_percent())
            
            if len(sent_data) > 60:
                sent_data.pop(0)
                recv_data.pop(0)
                cpu_data.pop(0)

            # 3. Mise à jour graphique
            l_sent.set_ydata(sent_data)
            l_recv.set_ydata(recv_data)
            l_cpu.set_ydata(cpu_data)
            
            # Ajuster les échelles Y dynamiquement
            ax_net.set_ylim(0, max(max(sent_data), max(recv_data)) + 10)
            ax_cpu.set_ylim(0, 100)
            
            return l_sent, l_recv, l_cpu

        canvas_widget = FigureCanvasTkAgg(fig, master=graph_window)
        canvas_widget.draw()
        canvas_widget.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        ani = animation.FuncAnimation(fig, update, interval=1000, blit=False)
        
        # Nettoyage mémoire à la fermeture
        def on_closing():
            try:
                ani.event_source.stop()  # Arrêter l'animation
                plt.close(fig)  # Fermer explicitement la figure
                self.graph_window = None  # Libérer la référence
                graph_window.destroy()
            except Exception as e:
                graph_window.destroy()
        
        graph_window.protocol("WM_DELETE_WINDOW", on_closing)
        
        # Garder référence pour éviter GC ET pour vérification d'existence
        self.graph_window = graph_window
        graph_window.ani = ani
        
        self.log("[+] Monitoring système réel activé.", tag="ok")
    
    def show_3d_dashboard(self):
        """
        DASHBOARD 3D RÉEL : Visualise les hôtes scannés par Nmap.
        Axe X/Y : Position réseau
        Axe Z : Niveau de risque (Nombre de ports ouverts)
        """
        # Charger d'abord : PLOTLY_AVAILABLE ne passe a True que dans
        # _load_plotly(). Tester le drapeau sans appeler le chargeur
        # rendait ces vues inaccessibles meme avec Plotly installe.
        _load_plotly()
        if not PLOTLY_AVAILABLE:
            messagebox.showerror("Erreur", "Installez Plotly : pip install plotly")
            return

        # Vérification qu'un scan a eu lieu
        if not self.nm or not self.nm.all_hosts():
            messagebox.showwarning("Données Manquantes", "Veuillez d'abord lancer un SCAN (Rapide ou Complet) pour générer les données.")
            return

        self.log("[*] Génération de la topologie 3D réelle...", tag="title")
        
        import plotly.graph_objects as go
        import math

        hosts = self.nm.all_hosts()
        x_vals, y_vals, z_vals = [], [], []
        colors = []
        hover_texts = []

        # Centre (Scanner/Localhost)
        x_vals.append(0)
        y_vals.append(0)
        z_vals.append(0)
        colors.append(0)
        hover_texts.append("Poste d'analyse")

        # Placement des hôtes en cercle autour du scanner
        radius = 10
        for i, host in enumerate(hosts):
            angle = (2 * math.pi * i) / len(hosts)
            
            # Position X/Y
            x_vals.append(radius * math.cos(angle))
            y_vals.append(radius * math.sin(angle))
            
            # Position Z (Hauteur) = Nombre de ports ouverts (Risque)
            try:
                # On compte les ports TCP ouverts
                open_ports = len(self.nm[host].get('tcp', {}))
            except:
                open_ports = 0
            
            z_vals.append(open_ports)
            
            # Couleur basée sur le nombre de ports (Plus rouge = plus de ports)
            colors.append(open_ports)
            
            # Infos bulle
            hostname = self.nm[host].hostname() or "Inconnu"
            hover_texts.append(f"IP: {host}<br>Host: {hostname}<br>Ports Ouverts: {open_ports}")

        # Création du graphe
        fig = go.Figure(data=[go.Scatter3d(
            x=x_vals, y=y_vals, z=z_vals,
            mode='markers+text',
            text=[t.split('<br>')[0] for t in hover_texts], # Affiche juste l'IP
            textposition="top center",
            hovertext=hover_texts,
            hoverinfo='text',
            marker=dict(
                size=12,
                color=colors,
                colorscale='Viridis', 
                opacity=0.9,
                showscale=True,
                colorbar=dict(title="Ports Ouverts (Risque)")
            )
        )])

        # Lignes reliant le centre aux hôtes (Topologie étoile)
        for i in range(1, len(x_vals)):
            fig.add_trace(go.Scatter3d(
                x=[0, x_vals[i]],
                y=[0, y_vals[i]],
                z=[0, z_vals[i]],
                mode='lines',
                line=dict(color='green', width=2),
                hoverinfo='none',
                showlegend=False
            ))

        fig.update_layout(
            title='TOPOLOGIE RÉSEAU 3D (Données Nmap Réelles)',
            scene=dict(
                xaxis_title='Réseau X',
                yaxis_title='Réseau Y',
                zaxis_title='Niveau de Risque (Hauteur)',
                bgcolor='black'
            ),
            paper_bgcolor='black',
            font=dict(color='#FFFFFF')
        )

        # BUG #2: Sauvegarder dans fichier temporaire tracké
        import tempfile
        import webbrowser
        temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.html', encoding='utf-8')
        temp_filename = temp_file.name
        temp_file.close()
        
        fig.write_html(temp_filename)
        self.temp_html_files.append(temp_filename)  # Tracker pour nettoyage
        
        webbrowser.open(temp_filename)
        self.log(f"[+] Dashboard généré : {temp_filename}", tag="ok")
        self.log(f"[*] {len(self.temp_html_files)} fichiers HTML temporaires actifs", tag="info")
    
    def show_network_map(self):
        """Affiche une carte réseau interactive basée sur les VRAIES données Nmap"""
        # Charger d'abord : PLOTLY_AVAILABLE ne passe a True que dans
        # _load_plotly(). Tester le drapeau sans appeler le chargeur
        # rendait ces vues inaccessibles meme avec Plotly installe.
        _load_plotly()
        if not PLOTLY_AVAILABLE:
            messagebox.showerror("Erreur", "Plotly n'est pas installé.\nInstallez-le avec: pip install plotly")
            return
        
        # Vérification critique : A-t-on des données réelles ?
        if not self.nm or len(self.nm.all_hosts()) == 0:
            messagebox.showwarning("Pas de données", "Veuillez d'abord lancer un SCAN (Rapide ou Complet) pour peupler la carte.")
            return

        self.log("[*] Génération de la topologie réelle...", tag="title")
        
        # Construction des nœuds basés sur le scan réel
        # On définit le centre (Vous/Gateway)
        nodes_x = [0.5]
        nodes_y = [0.5]
        node_text = ["SCANNER (Localhost)"]
        node_color = [10] # Rouge pour le scanner
        
        edge_x = []
        edge_y = []
        
        # Récupération des hôtes réels
        hosts = self.nm.all_hosts()
        count = len(hosts)
        
        import math
        radius = 0.4
        
        for i, host in enumerate(hosts):
            # Disposition en cercle autour du scanner
            angle = (2 * math.pi * i) / count
            x = 0.5 + radius * math.cos(angle)
            y = 0.5 + radius * math.sin(angle)
            
            nodes_x.append(x)
            nodes_y.append(y)
            
            # Récupérer info réelle (OS, Hostname, Ports ouverts)
            try:
                hostname = self.nm[host].hostname() or host
                open_ports = len(self.nm[host].get('tcp', {}))
                details = f"{hostname}<br>IP: {host}<br>Ports ouverts: {open_ports}"
                
                # Couleur selon le risque (nombre de ports ouverts)
                color = open_ports * 2
            except:
                details = f"IP: {host}"
                color = 2
            
            node_text.append(details)
            node_color.append(color)
            
            # Créer le lien (Edge) entre Scanner et Hôte
            edge_x.extend([0.5, x, None])
            edge_y.extend([0.5, y, None])

        # Création du graphique Plotly
        edge_trace = go.Scatter(
            x=edge_x, y=edge_y,
            line=dict(width=1, color='#888'),
            hoverinfo='none',
            mode='lines')

        node_trace = go.Scatter(
            x=nodes_x, y=nodes_y,
            mode='markers+text',
            text=[t.split('<br>')[0] for t in node_text], # Juste le nom sur le graphe
            textposition="top center",
            hovertext=node_text, # Détails au survol
            hoverinfo='text',
            marker=dict(
                showscale=True,
                colorscale='YlOrRd',
                reversescale=True,
                color=node_color,
                size=20,
                colorbar=dict(
                    thickness=15,
                    title='Niveau Exposition (Ports)',
                    xanchor='left',
                    titleside='right'
                ),
                line_width=2))

        fig = go.Figure(data=[edge_trace, node_trace],
                        layout=go.Layout(
                            title=f'Topologie Réelle ({count} hôtes détectés)',
                            titlefont_size=16,
                            showlegend=False,
                            hovermode='closest',
                            margin=dict(b=20,l=5,r=5,t=40),
                            paper_bgcolor='black',
                            plot_bgcolor='black',
                            font=dict(color='#FFFFFF'), # Texte blanc
                            xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                            yaxis=dict(showgrid=False, zeroline=False, showticklabels=False))
                        )
        
        filename = f"real_topology_{int(time.time())}.html"
        fig.write_html(filename)
        self.log(f"[+] Carte générée avec succès : {filename}", tag="ok")
        
        import webbrowser
        webbrowser.open(filename)
    
    def toggle_monitoring(self):
        """Active/désactive le mode surveillance 24/7"""
        self.monitoring_active = not self.monitoring_active
        
        if self.monitoring_active:
            self.log("[+] MODE SURVEILLANCE 24/7 ACTIVÉ", tag="ok")
            self.log("[*] Monitoring continu démarré...", tag="info")
            self.log("[*] Détection automatique des menaces activée", tag="info")
            self.set_status("Surveillance continue active")

            # Lancer le monitoring en thread
            threading.Thread(target=self._monitoring_loop, daemon=True).start()
        else:
            self.log("[+] MODE SURVEILLANCE DÉSACTIVÉ", tag="warn")
            self.set_status("Surveillance arrêtée")
    
    def _monitoring_loop(self):
        """
        MONITORING RÉEL 24/7 :
        1. Surveillance des journaux Windows (Security) pour les échecs de login.
        2. Surveillance de l'usage CPU/RAM critique.
        """
        import time
        import psutil
        
        self.log("[INIT] Démarrage des sondes de surveillance réelles...", tag="info")
        last_check_time = datetime.now()

        while self.monitoring_active:
            try:
                # 1. Vérification Santé Système (DoS check)
                cpu = psutil.cpu_percent(interval=1)
                mem = psutil.virtual_memory().percent
                
                if cpu > 90:
                    self.log(f"[ALERTE] CPU CRITIQUE: {cpu}% - Possible Crypto-Miner ou DoS", tag="warn")
                    self.send_notification("Alerte Système", f"CPU à {cpu}%")
                
                # 2. Vérification Logs Sécurité Windows (PowerShell)
                # Cherche l'Event ID 4625 (Logon Failed) dans les 2 dernières minutes
                if platform.system() == "Windows":
                    cmd = "Get-EventLog -LogName Security -InstanceId 4625 -After (Get-Date).AddMinutes(-2) | Measure-Object | Select-Object -ExpandProperty Count"
                    output = self.run_safe_command(["powershell", "-Command", cmd])
                    
                    try:
                        failed_count = int(output.strip())
                        if failed_count > 0:
                            msg = f"BRUTE FORCE DÉTECTÉ : {failed_count} échecs de connexion récents !"
                            self.log(f"[SECURITY] {msg}", tag="error")
                            self.send_notification("ALERTE SÉCURITÉ", msg)
                            self.problems_found.append({'type': 'BRUTE FORCE LIVE', 'details': msg, 'action': 'Bloquer IP'})
                    except ValueError:
                        pass # Pas de logs ou erreur de conversion

                # Pause de 60 secondes entre les cycles pour ne pas surcharger le CPU
                time.sleep(60)

            except Exception as e:
                self.log(f"[Monitoring Error] {e}", tag="warn")
                time.sleep(60)
    
    @staticmethod
    def _peut_emettre_paquets_bruts():
        """Vrai si le processus peut emettre des paquets bruts.

        La fragmentation IP (-f) de nmap l exige : sous Windows cela signifie
        etre administrateur.
        """
        if platform.system() != "Windows":
            try:
                return os.geteuid() == 0
            except AttributeError:
                return False
        try:
            return bool(ctypes.windll.shell32.IsUserAnAdmin())
        except Exception:
            return False

    def _options_furtives(self, args):
        """Remplace le rythme du scan par des options reellement discretes."""
        options = [o for o in args.split() if not o.startswith("-T")]
        options += ["-T1", "--scan-delay", "500ms", "--max-retries", "1",
                    "--randomize-hosts"]
        if self._peut_emettre_paquets_bruts():
            options.append("-f")
        return " ".join(options)

    def toggle_stealth_mode(self):
        """Active/desactive le mode furtif, en disant exactement ce qu il fait.

        Le drapeau self.stealth_mode etait positionne ici mais AUCUN scan ne
        le lisait : le journal annoncait des delais aleatoires et une
        fragmentation IP qui n etaient jamais appliques. run_nmap applique
        desormais reellement ces options.
        """
        self.stealth_mode = not self.stealth_mode

        if self.stealth_mode:
            apercu = self._options_furtives("-sV -T4")
            self.log("[FURTIF] Mode furtif ACTIVE", tag="ok")
            self.log("   Les prochains scans nmap recevront ces options :", tag="info")
            self.log(f"   {apercu}", tag="accent")
            self.log("   -T1 + --scan-delay 500ms : rythme tres lent", tag="info")
            self.log("   --randomize-hosts : ordre des cibles brouille", tag="info")
            if self._peut_emettre_paquets_bruts():
                self.log("   -f : fragmentation IP appliquee", tag="info")
            else:
                self.log("   -f (fragmentation IP) NON applique : il faudrait lancer "
                         "SIPA en administrateur.", tag="warn")
            self.log("   Consequence : les scans seront nettement plus lents.", tag="warn")
            self.set_status("Mode furtif actif")
        else:
            self.log("[FURTIF] Mode furtif DESACTIVE (retour au rythme normal)", tag="warn")
            self.set_status("Mode furtif desactive")
    
    def _discover_local_hosts(self):
        """Liste les hotes reellement vus sur le reseau local (table ARP)."""
        hosts = []
        try:
            output = self.run_safe_command(
                ["arp", "-a"] if platform.system() == "Windows" else ["arp", "-n"])
            for line in (output or "").splitlines():
                match = re.search(r"(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})", line)
                if not match:
                    continue
                ip = match.group(1)
                # Ignorer diffusion et multicast, sans interet pour un inventaire
                if ip.endswith(".255") or ip.startswith(("224.", "239.", "255.")):
                    continue
                if ip not in hosts:
                    hosts.append(ip)
        except Exception as exc:
            self.log(f"[-] Decouverte ARP impossible : {exc}", tag="warn")
        return sorted(hosts)

    def detect_network_changes(self):
        """Détecte les changements sur le réseau"""
        self.log("[*] Analyse des changements réseau...", tag="title")
        
        # Decouverte REELLE des voisins via la table ARP du systeme
        # (auparavant : liste d'hotes codee en dur, donc toujours identique).
        current_hosts = self._discover_local_hosts()
        if not current_hosts:
            self.log("[-] Aucun hote detecte (table ARP vide ou illisible).", tag="warn")
            self.log("    Generez du trafic reseau puis relancez la detection.", tag="info")
            return
        
        if not self.network_baseline:
            # Premier scan - établir la baseline
            self.network_baseline = {'hosts': current_hosts, 'timestamp': datetime.now()}
            self.log("[+] Baseline réseau établie", tag="ok")
            self.log(f"[+] {len(current_hosts)} hosts enregistrés", tag="info")
        else:
            # Comparer avec la baseline
            old_hosts = set(self.network_baseline['hosts'])
            new_hosts = set(current_hosts)
            
            added = new_hosts - old_hosts
            removed = old_hosts - new_hosts
            
            self.log("\n DÉTECTION DE CHANGEMENTS:", tag="title")
            
            if added:
                self.log(f"[+] Nouveaux hosts ({len(added)}):", tag="ok")
                for host in added:
                    self.log(f"    • {host} [NOUVEAU]", tag="accent")
                    self.send_notification("Nouvel appareil", f"Détecté: {host}")
            
            if removed:
                self.log(f"[-] Hosts disparus ({len(removed)}):", tag="warn")
                for host in removed:
                    self.log(f"    • {host} [DISPARU]", tag="warn")
            
            if not added and not removed:
                self.log("[=] Aucun changement détecté", tag="info")
            
            # Mettre à jour la baseline
            self.network_baseline = {'hosts': current_hosts, 'timestamp': datetime.now()}
    
    def compare_scans(self):
        """Differentiel entre les deux derniers scans : hotes et ports.

        L'ancienne version comparait seulement le NOMBRE de constats, et son
        propre commentaire admettait "on va simuler l'extraction structurelle".
        Elle ne pouvait de toute facon jamais s'executer : rien n'alimentait
        l'historique.
        """
        self.log("\n[COMPARAISON DE SCANS]", tag="title")

        if len(self.scan_history) < 2:
            self.log(f"[!] Il faut au moins 2 scans "
                     f"({len(self.scan_history)} enregistre(s) pour l'instant).", tag="warn")
            return

        older, newer = self.scan_history[-2], self.scan_history[-1]
        self.log(f"Entre {older['timestamp']} et {newer['timestamp']}", tag="info")
        if older.get("target") != newer.get("target"):
            self.log(f"[!] Cibles differentes ({older.get('target')} puis "
                     f"{newer.get('target')}) : comparaison indicative.", tag="warn")

        old_hosts = (older.get("results") or {}).get("hosts", {})
        new_hosts = (newer.get("results") or {}).get("hosts", {})

        appeared = sorted(set(new_hosts) - set(old_hosts))
        vanished = sorted(set(old_hosts) - set(new_hosts))
        changes = 0

        for host in appeared:
            changes += 1
            self.log(f"  [+] Nouvel hote : {host}", tag="warn")
            self.problems_found.append({
                "type": "CHANGEMENT RÉSEAU", "host": host, "port": "—",
                "service": "Découverte",
                "details": "Hôte absent du scan précédent",
                "action": "Confirmer qu'il s'agit d'un équipement légitime"})

        for host in vanished:
            changes += 1
            self.log(f"  [-] Hote disparu : {host}", tag="info")

        for host in sorted(set(old_hosts) & set(new_hosts)):
            before = set((old_hosts[host].get("ports") or {}))
            after = set((new_hosts[host].get("ports") or {}))

            for port in sorted(after - before, key=lambda value: int(value)):
                changes += 1
                service = new_hosts[host]["ports"][port].get("service") or "inconnu"
                self.log(f"  [+] {host} : port {port} ({service}) OUVERT depuis",
                         tag="warn")
                self.problems_found.append({
                    "type": "CHANGEMENT RÉSEAU", "host": host, "port": port,
                    "service": service,
                    "details": "Port ouvert depuis le scan précédent",
                    "action": "Vérifier que cette ouverture est voulue"})

            for port in sorted(before - after, key=lambda value: int(value)):
                changes += 1
                service = old_hosts[host]["ports"][port].get("service") or "inconnu"
                self.log(f"  [-] {host} : port {port} ({service}) FERMÉ depuis",
                         tag="success")

            for port in sorted(after & before, key=lambda value: int(value)):
                old_version = old_hosts[host]["ports"][port].get("version", "")
                new_version = new_hosts[host]["ports"][port].get("version", "")
                if old_version and new_version and old_version != new_version:
                    changes += 1
                    self.log(f"  [~] {host}:{port} version {old_version} -> {new_version}",
                             tag="info")

        if changes:
            self.log(f"\n{changes} changement(s) detecte(s).", tag="warn")
        else:
            self.log("\nAucun changement entre les deux scans.", tag="success")

    def replay_scans(self):
        """Rejoue les scans précédents"""
        self.log("[*] Replay de scans précédents...", tag="title")
        
        if not self.scan_history:
            self.log("[!] Aucun historique de scan", tag="warn")
            return
        
        self.log(f"[+] {len(self.scan_history)} scans dans l'historique", tag="ok")
        
        for i, scan in enumerate(self.scan_history[-5:], 1):  # 5 derniers
            self.log(f"\n[{i}] Scan du {scan.get('timestamp', 'N/A')}", tag="info")
            self.log(f"    Cible: {scan.get('target', 'N/A')}", tag="accent")
            self.log(f"    Type: {scan.get('scan_type', 'N/A')}", tag="accent")
            time.sleep(0.5)
    
    def capture_screenshot(self):
        """Capture d'écran de l'application"""
        if not PIL_AVAILABLE:
            messagebox.showerror("Erreur", "PIL/Pillow n'est pas installé.\nInstallez-le avec:\npip install pillow")
            return
        
        self.log("[*] Capture d'écran...", tag="info")
        
        try:
            # Capturer l'écran
            screenshot = ImageGrab.grab()
            
            # Nom du fichier
            filename = f"screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            screenshot.save(filename)
            
            self.log(f"[+] Screenshot sauvegardé: {filename}", tag="ok")
            self.send_notification("Screenshot", f"Sauvegardé: {filename}")
            
        except Exception as e:
            self.log(f"[!] Erreur capture: {e}", tag="warn")
    
    def test_notification(self):
        """Test de notification Windows"""
        self.send_notification("Test Notification", "Les notifications sont opérationnelles!")
    
    def send_notification(self, title, message):
        """Envoie une notification Windows"""
        if not TOAST_AVAILABLE:
            self.log(f"[Notification] {title}: {message}", tag="info")
            return
        
        try:
            if self.toaster:
                self.toaster.show_toast(
                    title,
                    message,
                    duration=5,
                    icon_path=None,
                    threaded=True
                )
                self.log(f"Notification envoyée:{title}", tag="ok")
        except Exception as e:
            self.log(f"[!] Erreur notification: {e}", tag="warn")
    
    def toggle_sound_alerts(self):
        """Active/désactive les sons d'alerte"""
        self.sound_enabled = not self.sound_enabled
        
        if self.sound_enabled:
            self.log("SONS D'ALERTE ACTIVÉS", tag="ok")
            self.set_status("Alertes sonores activées")
            self.play_alert_sound()  # Test
        else:
            self.log("SONS D'ALERTE DÉSACTIVÉS", tag="warn")
            self.set_status("Alertes sonores désactivées")
    
    def play_alert_sound(self):
        """Joue un son d'alerte"""
        if not self.sound_enabled:
            return
        
        try:
            # Jouer un beep système
            import winsound
            winsound.Beep(1000, 500)  # 1000 Hz pendant 500ms
        except:
            pass
    
    def configure_email(self):
        """Configure l'envoi d'emails automatiques"""
        self.log("[*] Configuration Email...", tag="title")
        
        # Dialogue de configuration
        dialog = tk.Toplevel(self.root)
        dialog.title("Configuration Email")
        dialog.geometry("500x400")
        dialog.configure(bg=THEME["bg"])
        
        tk.Label(dialog, text="Configuration des Rapports Email", 
                font=("Consolas", 13, "bold"), bg=THEME["bg"], fg=THEME["fg"]).pack(pady=10)
        
        # Champs
        frame = tk.Frame(dialog, bg=THEME["bg"])
        frame.pack(padx=20, pady=10, fill="both", expand=True)
        
        fields = [
            ("SMTP Server:", "smtp.gmail.com"),
            ("SMTP Port:", "587"),
            ("Email:", "votre.email@gmail.com"),
            ("Password:", ""),
            ("Destinataires:", "dest1@mail.com,dest2@mail.com")
        ]
        
        entries = {}
        for i, (label, default) in enumerate(fields):
            tk.Label(frame, text=label, bg=THEME["bg"], fg=THEME["accent"],
                    font=("Consolas", 10)).grid(row=i, column=0, sticky="w", pady=5)
            
            entry = tk.Entry(frame, width=30, bg="#1a1a1a", fg=THEME["fg"],
                           font=("Consolas", 9), show="*" if "Password" in label else "")
            entry.grid(row=i, column=1, padx=10, pady=5)
            entry.insert(0, default)
            entries[label] = entry
        
        def save_config():
            try:
                port = int(entries["SMTP Port:"].get())
            except ValueError:
                messagebox.showerror("Port invalide",
                                     "Le port SMTP doit être un nombre (ex. 587).",
                                     parent=dialog)
                return

            recipients = [r.strip() for r in
                          entries["Destinataires:"].get().split(",") if r.strip()]
            # Mot de passe chiffre via la DPAPI Windows avant stockage : il
            # n'est plus lisible en clair dans config.json.
            password = _secrets.protect(entries["Password:"].get())

            if self.config_manager is not None:
                # On ecrit dans la config REELLEMENT lue par EmailNotifier
                # (auparavant sauvegarde dans un dict que personne ne lisait).
                cfg = self.config_manager
                cfg.set("email.enabled", True)
                cfg.set("email.smtp_server", entries["SMTP Server:"].get())
                cfg.set("email.smtp_port", port)
                cfg.set("email.sender_email", entries["Email:"].get())
                cfg.set("email.sender_password", password)
                cfg.set("email.recipients", recipients)

            self.email_config = {
                'smtp_server': entries["SMTP Server:"].get(),
                'smtp_port': port,
                'sender': entries["Email:"].get(),
                'recipients': recipients,
            }
            protected = _secrets.is_protected(password)
            self.log(f"[+] Configuration email enregistrée "
                     f"(mot de passe {'chiffré' if protected else 'en clair'}).",
                     tag="ok")
            dialog.destroy()
        
        tk.Button(dialog, text="Sauvegarder", command=save_config,
                 bg=THEME["fg"], fg=THEME["bg"], font=("Consolas", 10, "bold")).pack(pady=10)
        
        tk.Button(dialog, text="Test Email", command=self.send_test_email,
                 bg=THEME["accent"], fg="#000000", font=("Consolas", 10, "bold")).pack(pady=5)
    
    def send_test_email(self):
        """Envoie un email de test"""
        self.log("[*] Envoi email de test...", tag="info")
        
        if not self.email_config['sender']:
            self.log("[!] Configuration email incomplète", tag="warn")
            return
        
        try:
            msg = MIMEMultipart()
            msg['From'] = self.email_config['sender']
            msg['To'] = ', '.join(self.email_config['recipients'])
            msg['Subject'] = f"{APP_NAME} — courriel de test"
            
            body = """
            Ceci est un courriel de test envoyé par SIPA.
            
            Les rapports automatiques sont configurés.
            
            Timestamp: {}
            """.format(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Connexion SMTP
            server = smtplib.SMTP(self.email_config['smtp_server'], self.email_config['smtp_port'])
            server.starttls()
            server.login(self.email_config['sender'], self.email_config['password'])
            server.send_message(msg)
            server.quit()
            
            self.log("[+] Email envoyé avec succès!", tag="ok")

        except Exception as e:
            self.log(f"[!] Erreur envoi email: {e}", tag="warn")

    def check_threat_intelligence(self):
        """Vérifie si l'IP cible est dans une liste noire publique"""
        import requests
        
        target = self.get_target()
        # Nettoyage si c'est une plage IP
        if '/' in target: target = target.split('/')[0]
            
        self.log("[THREAT INTEL] VÉRIFICATION RÉPUTATION IP...", tag="title")
        self.log(f"Cible: {target}", tag="info")
        self.matrix_processing(lines=3, duration=0.3)
        
        # URL d'une liste noire connue (Emerging Threats Compromised IPs)
        blocklist_url = "https://rules.emergingthreats.net/blockrules/compromised-ips.txt"
        
        try:
            self.log("Téléchargement de la base de données EmergingThreats...", tag="info")
            response = requests.get(blocklist_url, timeout=5)
            
            if response.status_code == 200:
                bad_ips = response.text.splitlines()
                
                if target in bad_ips:
                    self.log(f"ALERTE: L'IP{target} EST LISTÉE COMME MALVEILLANTE !", tag="error")
                    self.problems_found.append({
                        'type': 'IP MALVEILLANTE',
                        'host': target,
                        'details': 'Présente dans la liste EmergingThreats',
                        'action': 'BLOQUER TOUT TRAFIC'
                    })
                else:
                    self.log(f"L'IP{target} n'est pas dans la liste noire actuelle.", tag="success")
                    self.log(f"(Base de données vérifiée: {len(bad_ips)} IPs connues)", tag="info")
            else:
                self.log("Erreur de connexion à la base de données.", tag="warn")
                
        except Exception as e:
            self.log(f"Erreur Threat Intel: {e}", tag="error")
        finally:
            self.root.after(0, self.stop_loading)
    
    def open_sandbox(self):
        """Non implemente : aucune isolation reelle n'est mise en place."""
        self._not_implemented(
            "SANDBOX",
            "Aucune isolation n'est realisee. Une vraie sandbox demande un "
            "conteneur ou une machine virtuelle dediee (Windows Sandbox, Docker).")

    # ========== INTÉGRATIONS EXTERNES ==========
    
    def _not_implemented(self, feature_name, reason):
        """Signale sans ambiguïté une fonctionnalité annoncée mais non implémentée."""
        self.log(f"\n[{feature_name}]", tag="title")
        self.log("FONCTIONNALITÉ NON IMPLÉMENTÉE", tag="warn")
        self.log(f"   {reason}", tag="info")
        self.log("   Aucune donnée n'a été collectée ni envoyée.", tag="info")

    def sync_tenable(self):
        """Non implémenté : synchronisation Tenable.io"""
        self._not_implemented(
            "TENABLE",
            "Nécessite un compte Tenable.io et l'implémentation de son API REST.")

    def sync_qualys(self):
        """Non implémenté : synchronisation Qualys VMDR"""
        self._not_implemented(
            "QUALYS",
            "Nécessite un compte Qualys VMDR et l'implémentation de son API REST.")

    def sync_defender(self):
        """Non implémenté : synchronisation Microsoft Defender for Endpoint"""
        self._not_implemented(
            "MICROSOFT DEFENDER",
            "Nécessite un tenant Defender for Endpoint et l'implémentation de son API REST.")

    def start_agent_mode(self):
        """Non implémenté : mode agent distribué"""
        self._not_implemented(
            "MODE AGENT",
            "Prévu pour des scans planifiés sur plusieurs machines ; nécessite un déploiement distribué.")

# =========================================================================
# PARTIE 11: DEPENDENCY CHECKER - AUTO-INSTALLATION DES MODULES
# =========================================================================
def check_and_install_dependencies():
    """Signale les dependances Python absentes, sans rien installer.

    Le nom d'import et le nom du paquet PyPI differaient pour python-nmap
    (module `nmap`) et Pillow (module `PIL`) : importlib.util.find_spec les
    voyait donc toujours comme absents, et CHAQUE lancement declenchait un
    `pip install python-nmap pillow` non sollicite de paquets deja presents.
    On corrige la correspondance, et on se contente d'informer : modifier
    l'environnement Python de quelqu'un a son insu n'a pas a etre fait ici.
    """
    required_modules = {
        "nmap": "python-nmap",
        "psutil": "psutil",
        "scapy": "scapy",
        "requests": "requests",
        "matplotlib": "matplotlib",
        "plotly": "plotly",
        "reportlab": "reportlab",
        "openpyxl": "openpyxl",
        "PIL": "Pillow",
        "win10toast": "win10toast",
        "sklearn": "scikit-learn",
    }

    import importlib.util

    print("[INIT] Verification des dependances Python...")
    missing = []
    for import_name, package in required_modules.items():
        try:
            found = importlib.util.find_spec(import_name) is not None
        except (ImportError, ValueError):
            found = False
        if not found:
            missing.append(package)

    if missing:
        print(f"[!] Dependances absentes : {', '.join(missing)}")
        print(f"    Pour les installer : pip install {' '.join(missing)}")
        print("    Ou tout d'un coup  : pip install -r requirements.txt")
        print("    L'application demarre quand meme ; chaque fonction")
        print("    concernee signalera elle-meme le module manquant.")
    else:
        print("[INIT] Toutes les dependances Python sont presentes.")
    return missing

if __name__ == "__main__":
    import sys
    import tkinter.simpledialog

    # Mode console : des qu'un argument est fourni, on n'ouvre pas l'interface.
    # Permet une tache planifiee Windows, un serveur sans ecran ou un script.
    if len(sys.argv) > 1:
        from sipa_core.cli import run as _run_cli
        sys.exit(_run_cli(AuditIA_Ultimate))
    
    # --- PROTECTION BOUCLE INFINIE (CRUCIAL) ---
    # On vérifie si on est dans un .EXE (frozen) ou un script .PY
    is_frozen = getattr(sys, 'frozen', False)

    # Si on est en mode développement (.py), on installe les libs manquantes
    if not is_frozen:
        try:
            # On suppose que la fonction check_and_install_dependencies existe plus haut
            check_and_install_dependencies()
        except NameError:
            pass # La fonction n'est pas définie, pas grave
    else:
        # Si on est en mode .EXE, on NE FAIT RIEN. 
        # PyInstaller a déjà inclus les librairies.
        print("[INIT] Mode Exécutable Détecté. Bypass de l'installation PIP interne.")

    # --- DÉPLOIEMENT AUTO (Nmap/Docker) ---
    # On garde l'installation de Nmap et Docker car ce sont des logiciels externes
    try:
        installer = AutoInstaller()
        installer.run_check_sequence()
    except Exception as e:
        print(f"[BOOTSTRAP ERROR] {e}")
        time.sleep(2)
    
    # --- CHARGEMENT CONFIGURATION ---
    # ConfigManager charge config.json (RÉEL, pas simulé)
    config_manager = ConfigManager("config.json")
    print(f"[CONFIG] Fichier de configuration chargé avec succès")
    
    # --- INITIALISATION DES COMPOSANTS RÉELS ---
    # Audit Logger (GDPR compliant)
    audit_logger = AuditLogger(config_manager)
    audit_logger.log_action('SYSTEM_STARTUP', 'Application démarrée')
    
    # Email Notifier (RÉEL smtp)
    email_notifier = EmailNotifier(config_manager)
    
    # Webhook Notifier (RÉEL http)
    webhook_notifier = WebhookNotifier(config_manager)
    
    # PDF Report Generator (RÉEL reportlab)
    pdf_generator = PDFReportGenerator(config_manager)
    
    # Scheduler (RÉEL APScheduler)
    scan_scheduler = ScanScheduler(config_manager, audit_logger, email_notifier)
    
    # Scan Queue (RÉEL ThreadPoolExecutor)
    scan_queue = ScanQueue(config_manager)
    
    print("[INIT] Tous les composants initialisés (RÉEL, PAS SIMULÉ)")
    
    # --- LANCEMENT DE L'INTERFACE ---
    
    # Détecter la langue système
    def detect_system_language():
        try:
            system_locale = locale.getlocale()[0]
            if not system_locale: system_locale = 'en_US'
            lang_code = system_locale.split('_')[0].upper() if system_locale else 'FR'
            lang_map = {'FR': 'FR', 'EN': 'EN', 'ES': 'ES', 'DE': 'DE', 'IT': 'IT', 'pt': 'ES', 'zh': 'EN', 'ja': 'EN'}
            return lang_map.get(lang_code, 'EN')
        except: return 'EN'
    
    # --- DÉBUT DU BLOC MAGIQUE 4K ---
    try:
        import ctypes
        # Active la haute résolution (High DPI) pour rendre le texte net en 4K
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        # Fallback pour Windows 7 ou versions antérieures
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except:
            pass
    # --- FIN DU BLOC MAGIQUE 4K ---

    root = tk.Tk()

    # NE PAS forcer 'tk scaling' : le processus est DPI-aware, Tk calcule deja
    # le bon facteur depuis la resolution reelle (ex. 2.667 sur un ecran a 200%).
    # L'ecraser avec GetScaleFactorForDevice (2.0) rendait tout trop petit.

    # Pas de transparence : sur un fond noir, meme 3% laisse transparaitre les
    # fenetres du dessous et rend l'interface illisible.
    try:
        root.attributes('-alpha', 1.0)
    except Exception:
        pass

    # Géométrie adaptative et centrage écran
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    target_w = max(1100, min(int(screen_w * 0.82), 1600))
    target_h = max(850, min(int(screen_h * 0.86), 1100))
    pos_x = (screen_w - target_w) // 2
    pos_y = (screen_h - target_h) // 2
    root.geometry(f"{target_w}x{target_h}+{pos_x}+{pos_y}")

    root.withdraw()
    
    # Fenêtre de licence
    def show_license_window():
        """Affiche la fenêtre de licence au démarrage"""
        license_window = tk.Toplevel(root)
        license_window.title("Contrat de licence")
        # Taille adaptee au DPI et bornee a l'ecran : "950x750" en pixels bruts
        # etait minuscule (et le contenu debordait) sur un ecran haute densite.
        _lic_scale = max(1.0, license_window.winfo_fpixels('1i') / 96.0)
        _lic_w = min(int(950 * _lic_scale), int(root.winfo_screenwidth() * 0.9))
        _lic_h = min(int(750 * _lic_scale), int(root.winfo_screenheight() * 0.9))
        license_window.geometry(
            f"{_lic_w}x{_lic_h}+{(root.winfo_screenwidth() - _lic_w) // 2}"
            f"+{max(0, (root.winfo_screenheight() - _lic_h) // 2)}")
        license_window.configure(bg=THEME["bg"])
        license_window.resizable(True, True)
        
        # Détecter la langue
        system_lang = detect_system_language()
        
        # Header
        header = tk.Frame(license_window, bg=THEME["fg"], height=50)
        header.pack(fill="x", padx=0, pady=0)
        
        title_label = tk.Label(header, text=f"{APP_NAME} — CONTRAT DE LICENCE",
                              font=("Consolas", 12, "bold"),
                              bg=THEME["fg"], fg=THEME["bg"])
        title_label.pack(pady=10)
        
        # Language selector
        lang_frame = tk.Frame(license_window, bg=THEME["bg"])
        lang_frame.pack(fill="x", padx=10, pady=5)
        
        lang_label = tk.Label(lang_frame, text="Language / Langue / Idioma:",
                             bg=THEME["bg"], fg=THEME["accent"], font=("Consolas", 9))
        lang_label.pack(side=tk.LEFT, padx=5)
        
        selected_lang = tk.StringVar(value=LANGUAGES.get(system_lang, 'English'))
        lang_combo = ttk.Combobox(lang_frame, textvariable=selected_lang,
                                 values=list(LANGUAGES.values()), state="readonly", width=15)
        lang_combo.pack(side=tk.LEFT, padx=5)
        
        # Text area with license
        text_frame = tk.Frame(license_window, bg=THEME["bg"])
        text_frame.pack(padx=10, pady=10, fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        license_text = tk.Text(text_frame, width=80, height=10, bg="#050505",
                              fg=THEME["accent"], font=("Consolas", 8),
                              yscrollcommand=scrollbar.set)
        license_text.pack(side=tk.LEFT, fill="both", expand=True)
        scrollbar.config(command=license_text.yview)
        
        # Update license text based on language selection
        def update_license_text(*args):
            selected = selected_lang.get()
            for lang_code, lang_name in LANGUAGES.items():
                if lang_name == selected:
                    license_text.config(state="normal")
                    license_text.delete(1.0, tk.END)
                    license_text.insert(1.0, LICENSE_TEXT.get(lang_code, LICENSE_TEXT['EN']))
                    license_text.config(state="disabled")
                    break
        
        lang_combo.bind("<<ComboboxSelected>>", update_license_text)
        
        # Initial license display
        license_text.insert(1.0, LICENSE_TEXT.get(system_lang, LICENSE_TEXT['EN']))
        license_text.config(state="disabled")
        
        # Checkbox acceptance
        check_frame = tk.Frame(license_window, bg=THEME["bg"])
        check_frame.pack(side=tk.BOTTOM, fill="x", padx=10, pady=5)
        
        accept_var = tk.BooleanVar(value=False)
        
        # Get text based on detected language
        license_trans = TRANSLATIONS.get(system_lang, TRANSLATIONS['EN'])
        accept_text = license_trans.get('license_accept', 'I accept the license agreement')
        accept_btn_text = license_trans.get('license_accept_btn', 'ACCEPT')
        decline_btn_text = license_trans.get('license_decline', 'Decline')
        
        accept_check = tk.Checkbutton(check_frame, text=accept_text,
                                     variable=accept_var, bg=THEME["bg"],
                                     fg=THEME["accent"], font=("Consolas", 9),
                                     activebackground=THEME["bg"], activeforeground=THEME["fg"])
        accept_check.pack(anchor="w", padx=10)
        
        # Buttons
        btn_frame = tk.Frame(license_window, bg=THEME["bg"])
        btn_frame.pack(side=tk.BOTTOM, fill="x", padx=10, pady=10)
        
        def accept_license():
        
            # Memoriser l'acceptation : sans cela le contrat revenait a
        
            # chaque demarrage.
        
            try:
        
                config_manager.set("interface.licence_acceptee", True)
        
            except Exception:
        
                pass
            if accept_var.get():
                # Récupérer la langue sélectionnée
                for lang_code, lang_name in LANGUAGES.items():
                    if lang_name == selected_lang.get():
                        root.default_language = lang_code
                        break
                license_window.destroy()
                root.deiconify()  # Montrer l'application principale
            else:
                messagebox.showwarning("Warning", "You must accept the license to use this software!")
        
        def decline_license():
            root.destroy()
        
        accept_btn = tk.Button(btn_frame, text=accept_btn_text, command=accept_license,
                              bg=THEME["fg"], fg=THEME["bg"], font=("Consolas", 10, "bold"),
                              width=15, padx=20, pady=5)
        accept_btn.pack(side=tk.RIGHT, padx=5)
        
        decline_btn = tk.Button(btn_frame, text=decline_btn_text, command=decline_license,
                               bg=THEME["bg_secondary"], fg=THEME["fg"], font=("Consolas", 10),
                               width=15, padx=20, pady=5)
        decline_btn.pack(side=tk.RIGHT, padx=5)
        
        # Empêcher la fermeture sans accepter
        def on_closing():
            decline_license()
        
        license_window.protocol("WM_DELETE_WINDOW", on_closing)
        license_window.grab_set()

    # Palette enregistree : elle doit etre appliquee AVANT que le moindre
    # widget soit cree, Tkinter figeant les couleurs a la construction.
    _palette = (config_manager.get("interface.theme", "sombre") or "sombre")
    theme_module.apply_palette(_palette)

    # Langue enregistree : elle prime sur la langue detectee du systeme.
    _langue = config_manager.get("interface.langue", None)

    # Configuration racine
    root.default_language = _langue or detect_system_language()
    root.title("NETWORK AUDIT")
    root.configure(bg=THEME["bg"])
    
    # Création de l'application
    app = AuditIA_Ultimate(root)

    # Le scheduler etait cree ici puis jamais transmis a l'interface : il ne
    # pouvait donc lancer aucun scan. On le relie et on l'active si configure.
    app.scan_scheduler = scan_scheduler
    app.config_manager = config_manager
    app.email_notifier = email_notifier
    app.webhook_notifier = webhook_notifier

    # Base de connaissances des appareils : registre des fabricants (~39 000
    # entrees) et regles de classification. Rafraichie au plus une fois par mois.
    try:
        from sipa_core.knowledge import KnowledgeBase
        app.knowledge = KnowledgeBase(log=app.log)
        threading.Thread(target=app.knowledge.refresh_oui, daemon=True).start()
    except Exception as _kb_exc:
        app.log(f"[CONNAISSANCES] Base indisponible : {_kb_exc}", tag="warn")

    # Base de vulnerabilites locale (alimentee depuis NVD), si activee en config.
    if (config_manager.get("cve", {}) or {}).get("enabled", True):
        try:
            from sipa_core.cve import CVEDatabase
            app.cve_db = CVEDatabase(
                api_key=(config_manager.get("nvd", {}) or {}).get("api_key"),
                log=app.log)
        except Exception as _cve_exc:
            app.log(f"[CVE] Base indisponible : {_cve_exc}", tag="warn")
    scan_scheduler.scan_callback = app._run_scheduled_scan
    scan_scheduler.start_scheduler()
    app.current_language = root.default_language
    app.lang_combo.set(LANGUAGES[app.current_language])
    
    # Hack pour Nmap après installation fraîche
    if app.nm is None:
        app.init_nmap()
    
    # --- LANCEMENT FINAL ---
    # Assure-toi que la fonction show_license_window est bien accessible ici
    # Sinon, utilise root.deiconify() directement
    # Le contrat n'est presente qu'au PREMIER lancement : il etait reaffiche
    # a chaque demarrage alors que USAGE_GUIDE.md annonce l'inverse.
    if config_manager.get("interface.licence_acceptee", False):
        root.deiconify()
    else:
        try:
            show_license_window()
        except NameError:
            root.deiconify()  # licence introuvable : on lance l'app directement
    
    root.mainloop()
